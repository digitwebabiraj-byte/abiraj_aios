# -*- coding: utf-8 -*-
"""
FMP -> standard merge file (read-only). Parses the 'Combined' sheet of FMP's xlsx output.
Env override: FMP_SRC (xlsx path). Does not touch the FMP task.
"""
import os, re, glob, json
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
BASE = os.path.abspath(os.path.join(HERE, "..", "..", "projects"))
SRC = os.environ.get("FMP_SRC") or glob.glob(os.path.join(BASE, "PRJ-2026-020*", "evidence", "final_outputs", "**", "*fast_moving*.xlsx"), recursive=True)[0]
OUT = os.path.join(HERE, "fmp_merge.json")

# (key, source-header, display-name, role, type)
SPEC = [
    ("rank",    "Overall Rank",     "Overall Rank",     "metric", "num"),
    ("sku",     "SKU",              "SKU",             "id",     "text"),
    ("title",   "Product Name",     "Product Name",    "id",     "text"),
    ("category","Category",         "Category",        "id",     "text"),
    ("f_amz",   "Amazon sold Qty",  "Amazon sold Qty",  "metric", "num"),
    ("f_ebay",  "eBay sold Qty",    "eBay sold Qty",    "metric", "num"),
    ("f_shop",  "Shopify sold Qty", "Shopify sold Qty", "metric", "num"),
    ("f_units", "Total Units Sold", "Total Units Sold", "metric", "num"),
    ("f_rev",   "Total Revenue",    "Total Revenue (€)","metric", "money"),
    ("f_stock", "Current Stock",    "Current Stock",   "metric", "num"),
    ("f_cover", "Stock Cover Days", "Stock Cover Days","metric", "num"),
    ("f_dec",   "Final Decision",   "Final Decision",  "metric", "text"),
]
COLUMNS = [{"key": k, "name": n, "role": r, "type": t} for (k, _s, n, r, t) in SPEC]

def clean(s):
    return "" if s is None else str(s).replace("�", "").replace("â‚¬", "").strip()
def num(x):
    try: return float(x)
    except (TypeError, ValueError): return None

wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
ws = wb["Combined"] if "Combined" in wb.sheetnames else wb[wb.sheetnames[-1]]
grid = list(ws.iter_rows(values_only=True))
hdr_i = next(i for i, r in enumerate(grid) if r and any(clean(c) == "SKU" for c in r))
header = [clean(c) for c in grid[hdr_i]]

def col_idx(src):
    for i, h in enumerate(header):
        if h == src or h.startswith(src):
            return i
    return None
IDX = {k: col_idx(s) for (k, s, _n, _r, _t) in SPEC}
missing = [k for k, v in IDX.items() if v is None]
if missing:
    raise SystemExit("FMP: columns not found: %s (headers=%s)" % (missing, header))

# as_of from preamble date, else file mtime
asof = ""
for r in grid[:hdr_i]:
    for c in r:
        m = re.search(r"(20\d\d-\d\d-\d\d)", str(c) if c else "")
        if m: asof = m.group(1)
if not asof:
    import datetime
    asof = datetime.date.fromtimestamp(os.path.getmtime(SRC)).isoformat()

rows = []
for r in grid[hdr_i + 1:]:
    sku = clean(r[IDX["sku"]]) if IDX["sku"] < len(r) else ""
    if not sku or sku.upper().startswith("TOTAL"):
        continue
    row = {}
    for (k, _s, _n, _r, t) in SPEC:
        v = r[IDX[k]] if IDX[k] < len(r) else None
        row[k] = num(v) if t in ("num", "money") else clean(v)
    rows.append(row)

out = {"task": "FMP", "label": "Fast Moving", "owner": "Mahima",
       "join_key": "sku", "as_of": asof, "columns": COLUMNS, "rows": rows}
json.dump(out, open(OUT, "w", encoding="utf-8"), ensure_ascii=False)
print("wrote", OUT)
print("task FMP |", len(rows), "rows |", len(COLUMNS), "columns | as_of", asof)

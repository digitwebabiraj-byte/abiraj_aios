# -*- coding: utf-8 -*-
"""
SMP -> standard merge file (read-only). Parses the 'Combined' sheet of SMP's xlsx output.
Env override: SMP_SRC (xlsx path). Does not touch the SMP task.
"""
import os, re, glob, json
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
BASE = os.path.abspath(os.path.join(HERE, "..", "..", "projects"))
SRC = os.environ.get("SMP_SRC") or glob.glob(os.path.join(BASE, "PRJ-2026-022*", "evidence", "final_outputs", "**", "*slow_moving*.xlsx"), recursive=True)[0]
OUT = os.path.join(HERE, "smp_merge.json")

SPEC = [
    ("sku",      "SKU",              "SKU",              "id",     "text"),
    ("title",    "Product Name",     "Product Name",     "id",     "text"),
    ("s_stock",  "Stock Qty",        "Stock Qty",        "metric", "num"),
    ("s_last",   "Last Sale Date",   "Last Sale Date",   "metric", "text"),
    ("s_30",     "Sold Qty (30",     "Sold Qty (30 Days)","metric", "num"),
    ("s_90",     "Sold Qty (90",     "Sold Qty (90 Days)","metric", "num"),
    ("s_days",   "Days Without Sale","Days Without Sale","metric", "text"),
    ("s_reason", "Reason",           "Reason",           "metric", "text"),
    ("s_action", "Action",           "Action",           "metric", "text"),
]
COLUMNS = [{"key": k, "name": n, "role": r, "type": t} for (k, _s, n, r, t) in SPEC]

def clean(s):
    return "" if s is None else str(s).replace("�", "").strip()
def num(x):
    try: return float(x)
    except (TypeError, ValueError): return None

wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
ws = wb["Combined"] if "Combined" in wb.sheetnames else wb[wb.sheetnames[-1]]
grid = list(ws.iter_rows(values_only=True))
hdr_i = next(i for i, r in enumerate(grid) if r and any(clean(c) == "SKU" for c in r))
header = [clean(c) for c in grid[hdr_i]]

def col_idx(src):
    for i, h in enumerate(header):
        if h == src or h.startswith(src):
            return i
    return None
IDX = {k: col_idx(s) for (k, s, _n, _r, _t) in SPEC}
missing = [k for k, v in IDX.items() if v is None]
if missing:
    raise SystemExit("SMP: columns not found: %s (headers=%s)" % (missing, header))

asof = ""
for r in grid[:hdr_i]:
    for c in r:
        m = re.search(r"(20\d\d-\d\d-\d\d)", str(c) if c else "")
        if m: asof = m.group(1)
if not asof:
    import datetime
    asof = datetime.date.fromtimestamp(os.path.getmtime(SRC)).isoformat()

rows = []
for r in grid[hdr_i + 1:]:
    sku = clean(r[IDX["sku"]]) if IDX["sku"] < len(r) else ""
    if not sku or sku.upper().startswith("TOTAL"):
        continue
    row = {}
    for (k, _s, _n, _r, t) in SPEC:
        v = r[IDX[k]] if IDX[k] < len(r) else None
        row[k] = num(v) if t in ("num", "money") else clean(v)
    rows.append(row)

out = {"task": "SMP", "label": "Slow / No-Moving", "owner": "Mahima",
       "join_key": "sku", "as_of": asof, "columns": COLUMNS, "rows": rows}
json.dump(out, open(OUT, "w", encoding="utf-8"), ensure_ascii=False)
print("wrote", OUT)
print("task SMP |", len(rows), "rows |", len(COLUMNS), "columns | as_of", asof)

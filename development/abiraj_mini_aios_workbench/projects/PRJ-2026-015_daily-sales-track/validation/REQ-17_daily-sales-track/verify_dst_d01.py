# -*- coding: utf-8 -*-
"""
Independent verification harness for REQ-17-D01 (Daily Sales Track).

Deliberately does NOT import build_dst_d01 or dst_d01_rows. It holds its own reference figures,
obtained from a SEPARATE per-marketplace re-derivation against the live `ledsone` database, plus a
whole-channel aggregate with no GROUP BY at all, and checks the shipped workbook against them.

Formulas are recalculated with LibreOffice before reading, so derived cells are verified as
EVALUATED VALUES, not merely as formula strings.

Usage:  python verify_dst_d01.py
"""

import os
import shutil
import subprocess
import sys
import tempfile

from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.abspath(os.path.join(
    HERE, "..", "..", "evidence", "final_outputs", "REQ-17_daily-sales-track",
    "REQ-17-D01_daily_sales_track.xlsx"))

SOFFICE = r"C:\Program Files\LibreOffice\program\soffice.exe"

EXPECTED_HEADERS = [
    "Account", "Market", "Currency", "Date", "Today's Sales", "Yesterday Sales",
    "Sales Diff", "Sales Growth %", "Same Day LY Sales",
    "Today's Orders", "Yesterday Orders", "Order Growth %", "Same Day LY Orders",
    "Units Sold", "Avg Order Value", "Active Listing",
    "AH Listing", "AH Listing Sales", "AH Sales Trend",
    "PH Listing", "PH Listing Sales", "PH Sales Trend", "Account Sales Trend",
    "AH Holder",
]

# Independent per-marketplace re-derivation, live ledsone, 2026-07-23.
# "Account|Market" -> (sales_r1, sales_r2, sales_ly, orders_r1, orders_r2, orders_ly)
REF_TOTALS = {
    "LEDSone UK|UK":             (837.93, 863.36, 1233.51, 42, 49, 78),
    "LEDSone UK|Germany":        (306.58, 61.33, 534.56, 6, 4, 22),
    "LEDSone UK|France":         (0.0, 26.33, 0.0, 0, 1, 0),
    "LEDSone UK|Ireland":        (0.0, 5.81, 0.0, 0, 1, 0),
    "ElectricalSone UK|UK":      (397.95, 444.60, 497.19, 21, 28, 34),
    "ElectricalSone UK|Germany": (45.48, 159.54, 105.46, 4, 9, 6),
    "Sunsone|UK":                (261.26, 198.00, 259.76, 14, 10, 13),
    "Sunsone|Germany":           (32.29, 78.15, 152.48, 2, 4, 12),
    "LEDSone DE|Germany":        (149.50, 241.79, 212.66, 11, 10, 11),
    "Huetten Lampen DE|Germany": (397.19, 256.76, 394.90, 19, 11, 23),
    "Coventry Lights|UK":        (202.68, 280.92, 113.61, 8, 15, 5),
    "Vintage Interior|UK":       (157.59, 75.78, 82.87, 8, 8, 10),
    "DC Transformer|UK":         (41.99, 57.70, 27.34, 4, 2, 3),
    "Retro LED|UK":              (0.0, 65.78, 314.72, 0, 2, 6),
    "Neighbour Market|US":       (0.0, 29.62, 0.0, 0, 1, 0),
    "Lighting Sone|UK":          (0.0, 16.58, 0.0, 0, 1, 0),
    "Homin GmbH|Germany":        (152.91, 28.98, 0.0, 3, 2, 0),
}
# Money is NEVER blended. Independent per-currency aggregate, live ledsone, 2026-07-23,
# obtained by joining order_management.order_info.currency.
REF_BY_CCY = {
    "GBP": {"r1": 1899.40, "r2": 2002.72, "ly": 2529.00, "o_r1": 97, "o_r2": 115},
    "EUR": {"r1": 1083.95, "r2": 858.69, "ly": 1400.06, "o_r1": 45, "o_r2": 42},
    "USD": {"r1": 0.0, "r2": 29.62, "ly": 0.0, "o_r1": 0, "o_r2": 1},
}
REF_ORDERS = {"r1": 142, "r2": 158}
# listings.market_place_id_mapping, verified 2026-07-23
SITE_CCY = {"UK": "GBP", "Germany": "EUR", "France": "EUR", "Ireland": "EUR",
            "Austria": "EUR", "Italy": "EUR", "Spain": "EUR", "Netherlands": "EUR",
            "US": "USD", "Canada": "CAD"}
REF_ROWS = 30
REF_ACTIVE_TOTAL = 14606

# The external anchor: Thinesh's own Seller Hub screen for 22 Jul 2026.
SELLER_HUB_ANCHOR = ("LEDSone UK|UK", 837.93)

TREND_BAND = 0.05
ERRORS = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NULL!", "#NUM!", "Err:")

results = []


def check(name, ok, detail=""):
    results.append((name, bool(ok), detail))


def recalc(src):
    """Round-trip through LibreOffice so formula results are cached in the file.

    Input and output must be different directories or LibreOffice refuses to overwrite
    its own source and exits 1.
    """
    tmp = tempfile.mkdtemp(prefix="dst_verify_")
    in_dir = os.path.join(tmp, "in")
    out_dir = os.path.join(tmp, "out")
    os.makedirs(in_dir)
    os.makedirs(out_dir)
    work = os.path.join(in_dir, os.path.basename(src))
    shutil.copy2(src, work)
    proc = subprocess.run(
        [SOFFICE, "--headless", "--norestore", "--convert-to", "xlsx",
         "--outdir", out_dir, work],
        capture_output=True, timeout=240)
    out = os.path.join(out_dir, os.path.basename(src))
    if proc.returncode != 0 or not os.path.isfile(out):
        raise RuntimeError("LibreOffice recalc failed (rc={0}): {1}".format(
            proc.returncode, (proc.stderr or b"").decode("utf-8", "replace")[:400]))
    return out, tmp


def expected_trend(cur, prev):
    if prev == 0 and cur == 0:
        return ""
    if prev == 0:
        return "Up"
    g = (cur - prev) / prev
    return "Up" if g > TREND_BAND else ("Down" if g < -TREND_BAND else "Stable")


def main():
    if not os.path.isfile(XLSX):
        print("MISSING WORKBOOK: {0}".format(XLSX))
        return 1

    work, tmp = recalc(XLSX)
    wb = load_workbook(work, data_only=True)
    try:
        ws = wb["Daily Sales Track"]

        hdrs = [ws.cell(1, c).value for c in range(1, len(EXPECTED_HEADERS) + 1)]
        check("V1 24 headers exact and in order", hdrs == EXPECTED_HEADERS,
              "" if hdrs == EXPECTED_HEADERS else "got {0}".format(hdrs))
        check("V2 row count = 30 account x marketplace rows", ws.max_row == REF_ROWS + 1,
              "max_row={0}".format(ws.max_row))

        rows = {}
        for r in range(2, ws.max_row + 1):
            rows["{0}|{1}".format(ws.cell(r, 1).value, ws.cell(r, 2).value)] = r
        check("V3 every reference row present", set(REF_TOTALS).issubset(set(rows)),
              "missing {0}".format(sorted(set(REF_TOTALS) - set(rows))))

        mism = []
        for kk, r in rows.items():
            exp = REF_TOTALS.get(kk, (0.0, 0.0, 0.0, 0, 0, 0))
            for col, e, label in ((5, exp[0], "s_r1"), (6, exp[1], "s_r2"), (9, exp[2], "s_ly"),
                                  (10, exp[3], "o_r1"), (11, exp[4], "o_r2"), (13, exp[5], "o_ly")):
                got = ws.cell(r, col).value or 0
                if abs(float(got) - e) > 0.005:
                    mism.append("{0}.{1}: {2} != {3}".format(kk, label, got, e))
        check("V4 all 180 measured cells match the per-marketplace re-derivation",
              not mism, "; ".join(mism[:6]))

        bad, act_total = [], 0
        for kk, r in rows.items():
            active = ws.cell(r, 16).value or 0
            ah = ws.cell(r, 17).value or 0
            ph = ws.cell(r, 20).value or 0
            act_total += active
            if ah + ph != active:
                bad.append("{0} AH+PH={1} != Active={2}".format(kk, ah + ph, active))
        check("V5 AH + PH = Active on every row", not bad, "; ".join(bad))
        check("V5b active listings total matches the independent count",
              act_total == REF_ACTIVE_TOTAL,
              "got {0} want {1}".format(act_total, REF_ACTIVE_TOTAL))

        diffbad, growbad, aovbad = [], [], []
        for kk, r in rows.items():
            exp = REF_TOTALS.get(kk, (0.0, 0.0, 0.0, 0, 0, 0))
            s1, s2, o1 = exp[0], exp[1], exp[3]
            got = ws.cell(r, 7).value
            if abs(float(got or 0) - (s1 - s2)) > 0.005:
                diffbad.append("{0}: {1}".format(kk, got))
            got = ws.cell(r, 8).value
            e = "" if s2 == 0 else (s1 - s2) / s2
            if e == "":
                if got not in (None, ""):
                    growbad.append("{0}: expected blank got {1}".format(kk, got))
            elif got is None or abs(float(got) - e) > 1e-6:
                growbad.append("{0}: {1} != {2}".format(kk, got, e))
            got = ws.cell(r, 15).value
            e = "" if o1 == 0 else s1 / o1
            if e == "":
                if got not in (None, ""):
                    aovbad.append("{0}: expected blank got {1}".format(kk, got))
            elif got is None or abs(float(got) - e) > 0.005:
                aovbad.append("{0}: {1} != {2}".format(kk, got, e))
        check("V6 Sales Diff formula evaluates correctly", not diffbad, "; ".join(diffbad))
        check("V7 Sales Growth % evaluates correctly (blank when prior = 0)",
              not growbad, "; ".join(growbad))
        check("V8 Avg Order Value evaluates correctly (blank when 0 orders)",
              not aovbad, "; ".join(aovbad))

        trendbad = []
        for kk, r in rows.items():
            exp = REF_TOTALS.get(kk, (0.0, 0.0, 0.0, 0, 0, 0))
            want = expected_trend(exp[0], exp[1])
            got = ws.cell(r, 23).value or ""
            got_word = got.split()[-1] if got else ""
            if got_word != want:
                trendbad.append("{0}: '{1}' != '{2}'".format(kk, got_word, want))
        check("V9 Account Sales Trend evaluates to the expected band",
              not trendbad, "; ".join(trendbad))

        ei = wb["Engine Inputs"]
        misalign = [r for kk, r in rows.items()
                    if (ws.cell(r, 1).value, ws.cell(r, 2).value, ws.cell(r, 3).value)
                    != (ei.cell(r, 1).value, ei.cell(r, 2).value, ei.cell(r, 3).value)]
        check("V10 Engine Inputs rows align 1:1 with the report", not misalign,
              "rows {0}".format(misalign))

        ccybad = []
        for kk, r in rows.items():
            site = ws.cell(r, 2).value
            want = SITE_CCY.get(site)
            got = ws.cell(r, 3).value
            if got != want:
                ccybad.append("{0}: {1} != {2}".format(kk, got, want))
        check("V15 per-row currency matches listings.market_place_id_mapping",
              not ccybad, "; ".join(ccybad))

        symbad = []
        for kk, r in rows.items():
            code = ws.cell(r, 3).value
            fmt = ws.cell(r, 5).number_format
            want = {"GBP": "£", "EUR": "€", "USD": "$", "CAD": "CA$"}[code]
            if want not in fmt:
                symbad.append("{0}: {1} lacks {2}".format(kk, fmt, want))
        check("V16 money cells carry the row's OWN currency symbol",
              not symbad, "; ".join(symbad))

        k = wb["KPI Summary"]
        kpi = {}
        for i in range(6, 40):
            lbl = k.cell(i, 1).value
            if lbl:
                kpi[lbl] = k.cell(i, 2).value
        kbad = []
        for ccy, ref in REF_BY_CCY.items():
            for lbl, key in (("Total Sales Today", "r1"), ("Total Sales Yesterday", "r2"),
                             ("Same Day LY Sales", "ly")):
                name = "{0} — {1}".format(lbl, ccy)
                got = kpi.get(name)
                if got is None or abs(float(got) - ref[key]) > 0.005:
                    kbad.append("{0}={1} want {2}".format(name, got, ref[key]))
        for lbl, want in (("Total Orders", REF_ORDERS["r1"]),
                          ("Yesterday Orders", REF_ORDERS["r2"]),
                          ("Rows (account x marketplace)", REF_ROWS)):
            if int(kpi.get(lbl) or 0) != want:
                kbad.append("{0}={1}".format(lbl, kpi.get(lbl)))
        check("V11 KPI totals reconcile PER CURRENCY to the live aggregate",
              not kbad, "; ".join(kbad[:6]))

        blended = [l for l in kpi
                   if l in ("Total Sales Today", "Total Sales Yesterday",
                            "Same Day LY Sales", "Average Order Value", "Overall Growth")]
        check("V17 no blended cross-currency money total anywhere on the KPI sheet",
              not blended, "found {0}".format(blended))

        hbad = [kk for kk, r in rows.items() if not str(ws.cell(r, 24).value or "").strip()]
        check("V13 AH Holder present on every row", not hbad, "; ".join(hbad))

        akey, aval = SELLER_HUB_ANCHOR
        got = ws.cell(rows[akey], 5).value if akey in rows else None
        check("V14 SELLER HUB ANCHOR: LEDSone UK / UK = 837.93",
              got is not None and abs(float(got) - aval) < 0.005, "got {0}".format(got))

        errs = []
        for sh in wb.worksheets:
            for row in sh.iter_rows():
                for c in row:
                    if isinstance(c.value, str) and any(e in c.value for e in ERRORS):
                        errs.append("{0}!{1}={2}".format(sh.title, c.coordinate, c.value))
        check("V12 zero formula errors anywhere in the workbook", not errs, "; ".join(errs[:8]))

    finally:
        wb.close()
        shutil.rmtree(tmp, ignore_errors=True)

    width = max(len(n) for n, _o, _d in results)
    print("\nREQ-17-D01 verification — {0}\n".format(os.path.basename(XLSX)))
    for name, ok, detail in results:
        print("  [{0}] {1}{2}".format("PASS" if ok else "FAIL", name.ljust(width + 2),
                                      "" if ok else "  -> " + detail))
    failed = [n for n, o, _d in results if not o]
    print("\n  {0}/{1} checks passed.\n".format(len(results) - len(failed), len(results)))
    return 1 if failed else 0


if __name__ == "__main__":
    sys.exit(main())

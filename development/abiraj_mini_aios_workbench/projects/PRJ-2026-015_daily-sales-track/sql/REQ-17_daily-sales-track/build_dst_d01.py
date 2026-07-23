# -*- coding: utf-8 -*-
"""
REQ-17-D01 — Daily Sales Track (DST) workbook generator.

PRJ-2026-015_daily-sales-track  ·  project_code = dst

DATA SOURCE LOCK (owner instruction 2026-07-23): every figure is retrieved from the raw `ledsone`
Postgres via the Ledsone DB MCP (https://mcp.ledsone.co.uk/mcp), with query patterns from the AIOS
knowledge base (https://docs.ledsone.co.uk/mcp). The warehouse `order_management_copy` is OUT OF
SCOPE for data retrieval.

GRAIN: one row per ACCOUNT x MARKETPLACE (decision F reversed 2026-07-23 — see dst_d01_rows.py).

The governed dataset lives in `dst_d01_rows.py`, retrieved read-only through that MCP. The SQL that
produced every block is recorded verbatim in SQL_USED below, so a scheduled run (REQ-17-D02) can
execute exactly these statements over psycopg2 and feed the same render_workbook().

Read-only. Writes one .xlsx and one .json. Touches no live table.
"""

import io
import json
import os
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from dst_d01_rows import ROWS, UNASSIGNED

# ---------------------------------------------------------------------------
# Anchor — decision B (owner, 2026-07-23)
#   run date R          -> the morning the report is generated
#   R-1 "Today's ..."   -> the day being reported
#   R-2 "Yesterday ..." -> the comparison day
#   LY                  -> same CALENDAR date one year before R-1 (decision C)
# ---------------------------------------------------------------------------
RUN_DATE = date(2026, 7, 23)
D_R1 = date(2026, 7, 22)
D_R2 = date(2026, 7, 21)
D_LY = date(2025, 7, 22)

TREND_BAND = 0.05  # decision E — PROVISIONAL, editable on the Config sheet

SQL_USED = {
    "daily_totals_by_market": """
SELECT ss.name AS account, mp.name AS site,
  ROUND(SUM(o.total) FILTER (WHERE o.order_date::date=DATE '2026-07-22'),2) AS s_r1,
  ROUND(SUM(o.total) FILTER (WHERE o.order_date::date=DATE '2026-07-21'),2) AS s_r2,
  ROUND(SUM(o.total) FILTER (WHERE o.order_date::date=DATE '2025-07-22'),2) AS s_ly,
  COUNT(DISTINCT o.id) FILTER (WHERE o.order_date::date=DATE '2026-07-22') AS o_r1,
  COUNT(DISTINCT o.id) FILTER (WHERE o.order_date::date=DATE '2026-07-21') AS o_r2,
  COUNT(DISTINCT o.id) FILTER (WHERE o.order_date::date=DATE '2025-07-22') AS o_ly
FROM order_management.orders o
JOIN order_management.sub_source ss ON ss.id=o.sub_source_id AND ss.source_id=2
LEFT JOIN order_management.market_place mp ON mp.id::text = o.market_place
WHERE o.status<>'Cancelled'
GROUP BY 1,2""",
    "units_and_ah_ph_by_market": """
WITH ph AS (SELECT DISTINCT ref_id FROM staff.ph_category_products WHERE source_id=2)
SELECT ss.name AS account, mp.name AS site, o.order_date::date AS d,
       SUM(CAST(oii.item_quantity AS INT)) AS units,
       ROUND(SUM(CASE WHEN ph.ref_id IS NOT NULL
             THEN CAST(oii.item_quantity AS INT)*CAST(oii.item_price AS NUMERIC) ELSE 0 END),2) AS ph_s,
       ROUND(SUM(CASE WHEN ph.ref_id IS NULL
             THEN CAST(oii.item_quantity AS INT)*CAST(oii.item_price AS NUMERIC) ELSE 0 END),2) AS ah_s
FROM order_management.orders o
JOIN order_management.sub_source ss ON ss.id=o.sub_source_id AND ss.source_id=2
LEFT JOIN order_management.market_place mp ON mp.id::text = o.market_place
JOIN order_management.order_item_info oii ON oii.order_id=o.id
LEFT JOIN ph ON ph.ref_id = oii.item_id
WHERE o.status<>'Cancelled' AND o.order_date::date IN (DATE '2026-07-22', DATE '2026-07-21')
GROUP BY 1,2,3""",
    "listings_ah_ph_by_market": """
WITH ph AS (SELECT DISTINCT ref_id FROM staff.ph_category_products WHERE source_id=2)
SELECT ss.name AS account, el.site,
       COUNT(DISTINCT el.item_id) AS active,
       COUNT(DISTINCT el.item_id) FILTER (WHERE ph.ref_id IS NOT NULL) AS ph_l,
       COUNT(DISTINCT el.item_id) FILTER (WHERE ph.ref_id IS NULL)     AS ah_l
FROM listings.ebay_listings el
JOIN order_management.sub_source ss ON ss.id = el.sub_source AND ss.source_id = 2
LEFT JOIN ph ON ph.ref_id = el.item_id
WHERE el.is_ended = 0 AND el.all_list = 1     -- AIOS KB: all_list=1, never is_child/is_parent
  AND el.site IS NOT NULL
GROUP BY 1,2""",
}

HEADERS = [
    "Account", "Market", "Date", "Today's Sales (£)", "Yesterday Sales (£)",
    "Sales Diff (£)", "Sales Growth %", "Same Day LY Sales (£)",
    "Today's Orders", "Yesterday Orders", "Order Growth %", "Same Day LY Orders",
    "Units Sold", "Avg Order Value (£)", "Active Listing",
    "AH Listing", "AH Listing Sales", "AH Sales Trend",
    "PH Listing", "PH Listing Sales", "PH Sales Trend", "Account Sales Trend",
    "AH Holder",
]
N_COLS = len(HEADERS)

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(bold=True, size=13, color="1F3864")
EDIT_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MONEY = '#,##0.00'
PCT = '0.00%'
INT = '#,##0'


def _trend_formula(cur_ref, prev_ref):
    """Up / Down / Stable against the editable band on Config!$B$2.

    Both periods zero -> blank (nothing to compare, and 'Stable' would imply trading).
    Prior zero, current positive -> Up (growth % is undefined, the direction is not).
    """
    band = "Config!$B$2"
    return (
        '=IF(AND({p}=0,{c}=0),"",'
        'IF({p}=0,"\U0001F4C8 Up",'
        'IF(({c}-{p})/{p}>{b},"\U0001F4C8 Up",'
        'IF(({c}-{p})/{p}<-{b},"\U0001F4C9 Down","➡ Stable"))))'
    ).format(c=cur_ref, p=prev_ref, b=band)


def render_workbook(out_path):
    wb = Workbook()

    # ---------------- Sheet 1 — Daily Sales Track ----------------
    ws = wb.active
    ws.title = "Daily Sales Track"

    for i, h in enumerate(HEADERS, start=1):
        c = ws.cell(1, i, h)
        c.fill, c.font, c.border = HDR_FILL, HDR_FONT, BOX
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r, row in enumerate(ROWS, start=2):
        ws.cell(r, 1, row["display"])
        ws.cell(r, 2, row["site"])
        ws.cell(r, 3, D_R1).number_format = 'd-mmm-yyyy'
        ws.cell(r, 4, row["s_r1"]).number_format = MONEY
        ws.cell(r, 5, row["s_r2"]).number_format = MONEY
        ws.cell(r, 6, "=D{0}-E{0}".format(r)).number_format = MONEY
        ws.cell(r, 7, '=IF(E{0}=0,"",(D{0}-E{0})/E{0})'.format(r)).number_format = PCT
        ws.cell(r, 8, row["s_ly"]).number_format = MONEY
        ws.cell(r, 9, row["o_r1"]).number_format = INT
        ws.cell(r, 10, row["o_r2"]).number_format = INT
        ws.cell(r, 11, '=IF(J{0}=0,"",(I{0}-J{0})/J{0})'.format(r)).number_format = PCT
        ws.cell(r, 12, row["o_ly"]).number_format = INT
        ws.cell(r, 13, row["units_r1"]).number_format = INT
        ws.cell(r, 14, '=IF(I{0}=0,"",D{0}/I{0})'.format(r)).number_format = MONEY
        ws.cell(r, 15, row["active"]).number_format = INT
        ws.cell(r, 16, row["ah_l"]).number_format = INT
        ws.cell(r, 17, row["ah_r1"]).number_format = MONEY
        ws.cell(r, 18, _trend_formula("Q{0}".format(r), "'Engine Inputs'!D{0}".format(r)))
        ws.cell(r, 19, row["ph_l"]).number_format = INT
        ws.cell(r, 20, row["ph_r1"]).number_format = MONEY
        ws.cell(r, 21, _trend_formula("T{0}".format(r), "'Engine Inputs'!C{0}".format(r)))
        ws.cell(r, 22, _trend_formula("D{0}".format(r), "E{0}".format(r)))
        c = ws.cell(r, 23, row["holder"])
        if row["holder"] == UNASSIGNED:
            c.font = Font(italic=True, color="808080")
        for col in range(1, N_COLS + 1):
            ws.cell(r, col).border = BOX

    last = len(ROWS) + 1
    ws.auto_filter.ref = "A1:{0}{1}".format(get_column_letter(N_COLS), last)
    ws.freeze_panes = "D2"
    widths = [21, 12, 12, 15, 16, 13, 13, 18, 13, 15, 13, 16, 11, 16, 13, 12, 15, 15, 12, 15, 15, 18, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 34

    # ---------------- Sheet 2 — KPI Summary ----------------
    k = wb.create_sheet("KPI Summary")
    k["A1"] = "Daily Sales Track — KPI Summary"
    k["A1"].font = TITLE_FONT
    k["A2"] = "Reporting day"
    k["B2"] = D_R1
    k["B2"].number_format = 'd-mmm-yyyy'
    k["A3"] = "Generated"
    k["B3"] = RUN_DATE
    k["B3"].number_format = 'd-mmm-yyyy'

    for i, h in enumerate(["KPI", "Value"], start=1):
        c = k.cell(5, i, h)
        c.fill, c.font, c.border = HDR_FILL, HDR_FONT, BOX

    rng = "'Daily Sales Track'"
    kpis = [
        ("Rows (account x marketplace)", "=COUNTA({0}!A2:A{1})".format(rng, last), INT),
        ("Total Sales Today", "=SUM({0}!D2:D{1})".format(rng, last), MONEY),
        ("Total Sales Yesterday", "=SUM({0}!E2:E{1})".format(rng, last), MONEY),
        ("Overall Growth", "=IF(B8=0,\"\",(B7-B8)/B8)", PCT),
        ("Total Orders", "=SUM({0}!I2:I{1})".format(rng, last), INT),
        ("Yesterday Orders", "=SUM({0}!J2:J{1})".format(rng, last), INT),
        ("Order Growth", "=IF(B11=0,\"\",(B10-B11)/B11)", PCT),
        ("Total Units Sold", "=SUM({0}!M2:M{1})".format(rng, last), INT),
        ("Average Order Value", "=IF(B10=0,\"\",B7/B10)", MONEY),
    ]
    for i, (label, formula, fmt) in enumerate(kpis, start=6):
        k.cell(i, 1, label).border = BOX
        c = k.cell(i, 2, formula)
        c.number_format, c.border = fmt, BOX
    k.column_dimensions["A"].width = 28
    k.column_dimensions["B"].width = 18

    # ---------------- Sheet 3 — Config ----------------
    cfg = wb.create_sheet("Config")
    cfg["A1"] = "Editable engine configuration"
    cfg["A1"].font = TITLE_FONT
    cfg["A2"] = "Trend band (+/-)"
    cfg["B2"] = TREND_BAND
    cfg["B2"].number_format = PCT
    cfg["B2"].fill = EDIT_FILL
    cfg["B2"].border = BOX
    cfg["C2"] = ("PROVISIONAL - decision E is OPEN. Change this cell and every trend column "
                 "on the report re-evaluates. Above +band = Up, below -band = Down, else Stable.")
    notes = [
        "",
        "Evidence for the default: the source sample brackets the band - +6.91% shows 'Up' and "
        "+3.89% shows 'Stable', so the cut sits between them. 5% is the nearest round value.",
        "",
        "MEASURED WARNING (2026-07-23, last 30 days, live): day-over-day account sales are far more "
        "volatile than this band assumes. Median absolute daily move by account:",
        "   led_sone 15.6%  |  so_926407 18.5%  |  electricalsone 40.4%  |  ledsonede 48.1%  |  "
        "lighting_sone 62.6%",
        "At +/-5% only 6.5% of account-days read 'Stable'; at +/-10% just 14.6%; at +/-20% only 25.5%.",
        "CONSEQUENCE: whatever band is chosen, most rows will read Up or Down every day.",
        "RECOMMENDATION: compare against the SAME WEEKDAY LAST WEEK rather than the previous day. "
        "Daily retail has a strong weekday rhythm; removing it makes 'Stable' meaningful.",
    ]
    for i, n in enumerate(notes, start=3):
        cfg.cell(i, 1, n)
    cfg.column_dimensions["A"].width = 120

    # ---------------- Sheet 4 — Engine Inputs ----------------
    ei = wb.create_sheet("Engine Inputs")
    # Headers MUST stay on row 1: data rows are referenced POSITIONALLY by the trend formulas and
    # must align 1:1 with 'Daily Sales Track' rows 2..N. Do not sort or insert rows on either sheet.
    for i, h in enumerate(["Account", "Market",
                           "PH Sales {0}".format(D_R2.isoformat()),
                           "AH Sales {0}".format(D_R2.isoformat())], start=1):
        c = ei.cell(1, i, h)
        c.fill, c.font, c.border = HDR_FILL, HDR_FONT, BOX
    for r, row in enumerate(ROWS, start=2):
        ei.cell(r, 1, row["display"]).border = BOX
        ei.cell(r, 2, row["site"]).border = BOX
        ei.cell(r, 3, row["ph_r2"]).number_format = MONEY
        ei.cell(r, 4, row["ah_r2"]).number_format = MONEY
    for col, w in zip("ABCD", [21, 12, 22, 22]):
        ei.column_dimensions[col].width = w
    ei["F1"] = ("Prior-day (R-2) AH / PH sales - these drive the AH Sales Trend and PH Sales Trend "
                "columns. Rows 2..{0} align 1:1 with 'Daily Sales Track' rows 2..{0}; do not sort "
                "or insert rows on either sheet independently.".format(len(ROWS) + 1))
    ei["F1"].font = Font(bold=True, color="C00000")

    # ---------------- Sheet 5 — Data Notes ----------------
    dn = wb.create_sheet("Data Notes")
    dn["A1"] = "REQ-17-D01 Daily Sales Track — sources, definitions, assumptions and gaps"
    dn["A1"].font = TITLE_FONT
    lines = [
        "",
        "SOURCE (locked by owner instruction 2026-07-23):",
        "  Raw ledsone Postgres via the Ledsone DB MCP (https://mcp.ledsone.co.uk/mcp), read-only.",
        "  Query patterns from the AIOS knowledge base (https://docs.ledsone.co.uk/mcp).",
        "  The warehouse order_management_copy is OUT OF SCOPE for data retrieval.",
        "",
        "GRAIN: one row per ACCOUNT x MARKETPLACE (30 rows). CHANGED 2026-07-23.",
        "  It was one row per account. A Seller Hub check on 22 Jul showed LEDSone UK at GBP 837.93",
        "  while the account row read GBP 1,144.51 - both correct, because the account row combined",
        "  UK (837.93) and Germany (306.58). Seller Hub reports per marketplace, so this report now",
        "  does too: EVERY ROW TIES TO ONE SELLER HUB SCREEN.",
        "  Consequence: an account appears on several rows (LEDSone UK sells on 10 sites).",
        "",
        "ANCHOR (decision B): a report generated on date R reports R-1 as \"Today\" and R-2 as \"Yesterday\".",
        "  Generated {0}.  \"Today\" = {1}.  \"Yesterday\" = {2}.  Same Day LY = {3}.".format(
            RUN_DATE.isoformat(), D_R1.isoformat(), D_R2.isoformat(), D_LY.isoformat()),
        "  The column headers still read \"Today's\" / \"Yesterday\" - the dates above are what they mean.",
        "",
        "SALES (decision M): every order PLACED on the day, excluding status 'Cancelled' only.",
        "  Sales   = SUM(order_management.orders.total)   (order grain - one row per order)",
        "  Orders  = COUNT(DISTINCT orders.id)",
        "  Units   = SUM(CAST(order_item_info.item_quantity AS INT))",
        "  Why not 'Completed' only: orders reach Completed about 2 days after purchase. On {0}".format(D_R1.isoformat()),
        "  only 36 of 142 orders (25.4%) had matured. Filtering on Completed would have reported",
        "  GBP 928.58 instead of GBP 2,983.35 - a 69% understatement that reads as a crash.",
        "",
        "!! DOES NOT TIE TO EBPD (REQ-13). That monthly dashboard counts Completed only, and is built",
        "   from the warehouse mirror, which diverges from live ledsone by roughly 0.2-0.4%.",
        "",
        "SCOPE (decision G): all eBay accounts. Universe = every account x site with live listings.",
        "",
        "AH / PH (decision A): AH is the PH remainder. A live listing with no portfolio-holder category",
        "  assignment belongs to the account holder, so AH Listing + PH Listing = Active Listing on every row.",
        "  Live totals: 14,606 listings = 2,750 PH + 11,856 AH across 30 account x marketplace rows.",
        "  Listing filter is all_list = 1 AND is_ended = 0, per the AIOS knowledge base, which states",
        "  explicitly: do not use is_child / is_parent combinations (decision K).",
        "  !! ONE live listing carries a NULL site and cannot be placed in a marketplace. It is excluded,",
        "     which is why the total is 14,606 here against 14,607 at account grain.",
        "",
        "!! ACTIVE LISTING IS UNDERSTATED BY ROUGHLY 5-6% - DO NOT QUOTE IT AGAINST SELLER HUB.",
        "   Measured 2026-07-23 against eBay Seller Hub for led_sone:",
        "     eBay active, UK site      3,033   |  this report  2,843   (-190)",
        "     eBay active, all sites    6,883   |  this report  6,510   (-373)",
        "   Cause: the ledsone listings mirror flags a listing is_ended=1 when its end_date passes, but",
        "   eBay AUTO-RENEWS Good-Til-Cancelled listings and the flag is not cleared until a full",
        "   re-sync. The mirror holds 3,450 UK item_ids for led_sone and calls 606 ended; if eBay is",
        "   right only ~417 are. It is NOT a sync outage (68,033 rows updated in the last 24h) and NOT",
        "   a date bug (only 1 listing is flagged ended with a future end_date).",
        "   This is a defect in the listings sync, outside this report. Sales / orders / units / AOV are",
        "   unaffected - they come from the orders table and are externally verified.",
        "   AH Listing and PH Listing split the same understated total, so their PROPORTIONS hold but",
        "   their absolute counts inherit the same shortfall.",
        "",
        "   CORRECTION 2026-07-23: an earlier draft of this note said the gap to REQ-16 (ESNM, 11,156)",
        "   was 'a corrected definition plus a wider scope'. That was WRONG about the definition.",
        "   Measured: all_list=1 gives 14,607 and is_child=0 gives 14,602 across all sites - 5 listings",
        "   apart, 0.03%. On UK+DE they are 11,175 vs 11,176. The two rules are effectively identical;",
        "   the gap to ESNM is ~99.97% SCOPE (all sites here vs UK+DE there), not definition.",
        "   all_list=1 is still the filter used, because the AIOS knowledge base mandates it.",
        "",
        "!! AH/PH SALES ARE LINE-LEVEL and will NOT sum exactly to Today's Sales. Today's Sales is the",
        "   order grand total (includes postage, discounts); AH and PH sales are SUM(qty x item price)",
        "   attributed per listing. MEASURED: +GBP 25.35 channel-wide, or 0.85%. Postage and discount,",
        "   not an error.",
        "",
        "!! UNITS IS DEFINITION-SENSITIVE. Units Sold = SUM(item_quantity), the AIOS knowledge base's",
        "   canonical choice. order_item_info also carries real_qty; using it gives 222 rather than 223",
        "   for this day - exactly 1 of 152 order lines differs. REQ-16 used COALESCE(real_qty,",
        "   item_quantity). Flagged for the Business Validator; not a defect.",
        "",
        "SAME DAY LY (decision C): the same CALENDAR date one year earlier, so the weekday differs",
        "  ({0} was a {1}; {2} was a {3}). Day-of-week effects are present in this comparison.".format(
            D_R1.isoformat(), D_R1.strftime("%A"), D_LY.isoformat(), D_LY.strftime("%A")),
        "",
        "UNITS (decision H): R-1 only. There is deliberately no prior-day or last-year unit figure.",
        "",
        "SNAPSHOT (decision I): this output REPLACES the previous one each morning. No history is kept.",
        "",
        "TRENDS (decision E - STILL OPEN): band is editable on the Config sheet, provisionally +/-5%.",
        "  Both periods zero renders blank, not 'Stable' - 'Stable' would imply trading took place.",
        "",
        "ZERO vs BLANK: 0 means measured zero. Blank means not applicable or nothing to compare.",
        "",
        "REMOVED: 'Best Seller' was dropped on the owner's instruction 2026-07-23 (decision D).",
        "",
        "ADDED: 'Market' (col 2) and 'AH Holder' (col 23) are NOT in the source sheet. Market comes from",
        "  the grain change; AH Holder was added on owner request so a reader can see who owns a falling",
        "  row without leaving the sheet. AH Holder is a MANUAL map - no database holds an",
        "  account-to-staff assignment - and must be maintained by hand.",
        "  Sunsone's two holders now sit on their own rows: Powsteena on Sunsone/UK, Sivajitha on",
        "  Sunsone/Germany. Thinesh originally wrote 'Jarshini', which matches nobody (staff.users holds",
        "  Jarsini id 91 AND Jasmini id 84); he confirmed Jarsini on 2026-07-23.",
        "",
        "STILL OPEN: E (trend band) - J (recipients and delivery time) - O (possible duplicate: a",
        "  Portfolio Holder dashboard app already exists in the ph_dashboard database).",
    ]
    for i, line in enumerate(lines, start=2):
        dn.cell(i, 1, line)
    dn.column_dimensions["A"].width = 118

    wb.save(out_path)
    return out_path


def write_governed_json(path):
    payload = {
        "deliverable": "REQ-17-D01",
        "project": "PRJ-2026-015_daily-sales-track",
        "project_code": "dst",
        "generated": RUN_DATE.isoformat(),
        "grain": "account x marketplace",
        "anchor": {"today": D_R1.isoformat(), "yesterday": D_R2.isoformat(),
                   "same_day_last_year": D_LY.isoformat()},
        "source": {"database": "ledsone",
                   "mcp": "https://mcp.ledsone.co.uk/mcp",
                   "knowledge_base": "https://docs.ledsone.co.uk/mcp",
                   "read_only": True},
        "trend_band": TREND_BAND,
        "sql": SQL_USED,
        "rows": ROWS,
    }
    with io.open(path, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, indent=2, ensure_ascii=False)
    return path


if __name__ == "__main__":
    here = os.path.dirname(os.path.abspath(__file__))
    out_dir = os.path.abspath(os.path.join(
        here, "..", "..", "evidence", "final_outputs", "REQ-17_daily-sales-track"))
    if not os.path.isdir(out_dir):
        os.makedirs(out_dir)
    xlsx = render_workbook(os.path.join(out_dir, "REQ-17-D01_daily_sales_track.xlsx"))
    js = write_governed_json(os.path.join(out_dir, "dst_d01_data.json"))
    print("workbook: {0}".format(xlsx))
    print("dataset : {0}".format(js))

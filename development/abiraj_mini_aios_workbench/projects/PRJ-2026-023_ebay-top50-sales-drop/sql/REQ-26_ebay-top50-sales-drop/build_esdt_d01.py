#!/usr/bin/env python3
"""
REQ-26-D01 — eBay UK Top 50 Sales Drop (esdt) · builder
PRJ-2026-023 · requester Kobiga · account ELECTRICALSONE (eBay UK)

Read-only. Connects to the live raw ledsone DB via LED_* env creds, runs the canonical
enriched ranking SQL, derives the report columns, and writes:
  - REQ-26-D01_ebay_top50_sales_drop.xlsx   (Notes & Method / Top 50 Sales Drop / Diagnostics)
  - esdt_payload.json                        (snapshot of the fetched rows, for audit/repro)

Confirmed business rules (Kobiga, 2026-08-12):
  grain = SKU; period = last 30d vs previous 30d; scope = ELECTRICALSONE eBay UK
  (orders.sub_source_id=22, market_place='23', status='Completed').
  Thresholds = PDF §6; Reason/Action ladder = PDF §8; rank by £ loss desc, tie Drop% desc, Top 50.
"""
import os, json, datetime as dt
import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = os.path.normpath(os.path.join(
    HERE, "..", "..", "evidence", "final_outputs", "REQ-26_ebay-top50-sales-drop"))
XLSX = os.path.join(OUT_DIR, "REQ-26-D01_ebay_top50_sales_drop.xlsx")
PAYLOAD = os.path.join(HERE, "esdt_payload.json")

SQL = r"""
WITH lines AS (
  SELECT oi.item_sku AS sku, oi.item_id, oi.item_title, o.order_date::date AS d, o.id AS oid,
         COALESCE(NULLIF(oi.item_quantity,'')::numeric,0) AS qty,
         COALESCE(NULLIF(oi.item_price,'')::numeric,0)    AS price
  FROM order_management.orders o
  JOIN order_management.order_item_info oi ON oi.order_id=o.id
  WHERE o.sub_source_id=22 AND o.market_place='23' AND o.status='Completed'
    AND o.order_date>=CURRENT_DATE-60 AND oi.item_sku IS NOT NULL AND oi.item_sku<>''
),
agg AS (
  SELECT sku,
    ROUND(SUM(CASE WHEN d>=CURRENT_DATE-30 THEN qty*price ELSE 0 END)::numeric,2) cur_sales,
    ROUND(SUM(CASE WHEN d>=CURRENT_DATE-60 AND d<CURRENT_DATE-30 THEN qty*price ELSE 0 END)::numeric,2) prev_sales,
    SUM(CASE WHEN d>=CURRENT_DATE-30 THEN qty ELSE 0 END)::int cur_units,
    SUM(CASE WHEN d>=CURRENT_DATE-60 AND d<CURRENT_DATE-30 THEN qty ELSE 0 END)::int prev_units
  FROM lines GROUP BY sku
),
top AS (
  SELECT sku, prev_sales, cur_sales, cur_units, prev_units,
         ROW_NUMBER() OVER (ORDER BY (prev_sales-cur_sales) DESC, ((prev_sales-cur_sales)/prev_sales) DESC) rank
  FROM agg WHERE prev_sales>0 AND cur_sales<prev_sales
  ORDER BY (prev_sales-cur_sales) DESC, ((prev_sales-cur_sales)/prev_sales) DESC LIMIT 50
),
rep AS (
  SELECT DISTINCT ON (l.sku) l.sku, l.item_id, l.item_title
  FROM lines l JOIN top t ON t.sku=l.sku ORDER BY l.sku, l.d DESC, l.oid DESC
),
bridge AS ( SELECT DISTINCT l.sku, l.item_id FROM lines l JOIN top t ON t.sku=l.sku WHERE l.item_id<>'' ),
trf AS (
  SELECT b.sku,
    SUM(CASE WHEN td.date>=CURRENT_DATE-30 THEN td.impressions ELSE 0 END) cur_impr,
    SUM(CASE WHEN td.date>=CURRENT_DATE-60 AND td.date<CURRENT_DATE-30 THEN td.impressions ELSE 0 END) prev_impr,
    SUM(CASE WHEN td.date>=CURRENT_DATE-30 THEN td.ebay_views ELSE 0 END) cur_views,
    SUM(CASE WHEN td.date>=CURRENT_DATE-60 AND td.date<CURRENT_DATE-30 THEN td.ebay_views ELSE 0 END) prev_views,
    SUM(CASE WHEN td.date>=CURRENT_DATE-30 THEN td.quantity_sold ELSE 0 END) cur_qsold
  FROM bridge b JOIN business_reports.ebay_traffic_data td ON td.item_id=b.item_id AND td.site_code='EBAY-GB'
  WHERE td.date>=CURRENT_DATE-60 GROUP BY b.sku
),
ppc AS (
  SELECT b.sku,
    SUM(CASE WHEN lp.date>=CURRENT_DATE-30 THEN lp.sale_amount_listing_currency ELSE 0 END) cur_ppc_sales,
    SUM(CASE WHEN lp.date>=CURRENT_DATE-30 THEN lp.ad_fees_listing_currency ELSE 0 END) cur_ppc_spend,
    SUM(CASE WHEN lp.date>=CURRENT_DATE-60 AND lp.date<CURRENT_DATE-30 THEN lp.ad_fees_listing_currency ELSE 0 END) prev_ppc_spend,
    SUM(CASE WHEN lp.date>=CURRENT_DATE-60 AND lp.date<CURRENT_DATE-30 THEN lp.sale_amount_listing_currency ELSE 0 END) prev_ppc_sales
  FROM bridge b JOIN ebay_campaigns.listing_performance lp ON lp.ebay_listing_id::text=b.item_id
  WHERE lp.date>=CURRENT_DATE-60 GROUP BY b.sku
),
stk AS (
  SELECT p.sku, SUM(COALESCE(s.stock,0))::int stock_uk
  FROM inventory.products p JOIN inventory.local_inventory_current_stock_location_wise s ON s.inventory_id=p.id
  WHERE s.warehouse_location='UK' GROUP BY p.sku
),
img AS (
  SELECT DISTINCT ON (item_id) item_id, main_image_url
  FROM listings.ebay_listings
  WHERE main_image_url IS NOT NULL AND main_image_url<>''
  ORDER BY item_id, is_parent DESC NULLS LAST
)
SELECT t.rank, t.sku, rep.item_id, rep.item_title AS product, img.main_image_url AS image,
       t.prev_sales::float, t.cur_sales::float, (t.cur_sales-t.prev_sales)::float loss_gbp,
       ROUND(100.0*(t.cur_sales-t.prev_sales)/t.prev_sales,1)::float drop_pct,
       t.prev_units, t.cur_units,
       COALESCE(trf.cur_impr,0) cur_impr, COALESCE(trf.prev_impr,0) prev_impr,
       COALESCE(trf.cur_views,0) cur_views, COALESCE(trf.prev_views,0) prev_views,
       COALESCE(trf.cur_qsold,0) cur_qsold,
       COALESCE(ppc.cur_ppc_sales,0)::float cur_ppc_sales, COALESCE(ppc.cur_ppc_spend,0)::float cur_ppc_spend,
       COALESCE(ppc.prev_ppc_spend,0)::float prev_ppc_spend, COALESCE(ppc.prev_ppc_sales,0)::float prev_ppc_sales,
       COALESCE(stk.stock_uk,0) stock_uk
FROM top t
LEFT JOIN rep ON rep.sku=t.sku
LEFT JOIN trf ON trf.sku=t.sku
LEFT JOIN ppc ON ppc.sku=t.sku
LEFT JOIN stk ON stk.sku=t.sku
LEFT JOIN img ON img.item_id=rep.item_id
ORDER BY t.rank;
"""

COLS = ["rank","sku","item_id","product","image","prev_sales","cur_sales","loss_gbp","drop_pct",
        "prev_units","cur_units","cur_impr","prev_impr","cur_views","prev_views","cur_qsold",
        "cur_ppc_sales","cur_ppc_spend","prev_ppc_spend","prev_ppc_sales","stock_uk"]

def fetch():
    conn = psycopg2.connect(host=os.environ["LED_PGHOST"], port=os.environ.get("LED_PGPORT",5432),
        dbname=os.environ["LED_PGDATABASE"], user=os.environ["LED_PGUSER"],
        password=os.environ["LED_PGPASSWORD"], connect_timeout=30)
    try:
        cur = conn.cursor(); cur.execute(SQL)
        rows = [dict(zip(COLS, r)) for r in cur.fetchall()]
        cur.close(); return rows
    finally:
        conn.close()

def priority(mag):
    if mag >= 50: return "🔴 Critical"
    if mag >= 30: return "🟠 High"
    if mag >= 15: return "🟡 Medium"
    return "🟢 Stable"

def diagnose(r):
    """PDF §8 reason/flag ladder → (Reason, Action)."""
    imp_c, imp_p = r["cur_impr"], r["prev_impr"]
    v_c, v_p = r["cur_views"], r["prev_views"]
    ctr_c = (v_c/imp_c) if imp_c else 0.0
    ctr_p = (v_p/imp_p) if imp_p else 0.0
    cvr_c = (r["cur_units"]/v_c) if v_c else 0.0
    cvr_p = (r["prev_units"]/v_p) if v_p else 0.0
    sp_c, sp_p = r["cur_ppc_spend"], r["prev_ppc_spend"]
    if r["stock_uk"] == 0:
        return "Out of stock", "Stock — restock urgently"
    if imp_p > 0 and imp_c < imp_p*0.85:
        return "Visibility / SEO drop (impressions ↓)", "SEO Review — title, keywords, item specifics"
    if ctr_p > 0 and ctr_c < ctr_p*0.85:
        return "CTR drop — listing appeal", "Main Image + Title Review"
    if cvr_p > 0 and cvr_c < cvr_p*0.85:
        return "Conversion drop — price / offer", "Price & Offer Review"
    if sp_p > 0 and sp_c < sp_p*0.85:
        return "PPC pull-back (ad spend ↓)", "PPC Review — budget & bids"
    return "Broad sales decline", "Review listing, price & PPC"

def derive(r):
    imp, v = r["cur_impr"], r["cur_views"]
    ctr = round(100.0*v/imp, 2) if imp else None            # % (views / impressions)
    cvr = round(100.0*r["cur_units"]/v, 2) if v else None    # % (units sold / views)
    roas = round(r["cur_ppc_sales"]/r["cur_ppc_spend"], 2) if r["cur_ppc_spend"] > 0 else None
    mag = abs(r["drop_pct"])
    reason, action = diagnose(r)
    return dict(ctr=ctr, cvr=cvr, roas=roas, priority=priority(mag), reason=reason, action=action)

# ---------- styling ----------
HDR_FILL = PatternFill("solid", fgColor="1F3B57")
HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(bold=True, size=16, color="1F3B57")
THIN = Side(style="thin", color="D6DCE4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
PRIO_FILL = {"🔴 Critical":"FDE0E0","🟠 High":"FCEBD6","🟡 Medium":"FCF6D6","🟢 Stable":"E3F4E3"}

HEADERS = ["Rank","SKU","Item ID","Product","Previous Sales (£)","Current Sales (£)","Loss (£)",
           "Drop %","CTR %","CVR %","ROAS","Stock","Priority","Reason","Action"]

def build_xlsx(rows, meta):
    wb = Workbook()
    # ---- Notes & Method ----
    ns = wb.active; ns.title = "Notes & Method"
    ns["A1"] = "eBay UK — Top 50 Sales Drop (ELECTRICALSONE)"; ns["A1"].font = TITLE_FONT
    notes = [
        "", f"Deliverable: REQ-26-D01   ·   Project: PRJ-2026-023 (esdt)   ·   Requester: Kobiga",
        f"Generated: {meta['generated']}   ·   Source: live raw ledsone DB (read-only)",
        "",
        "WHAT THIS IS",
        "  The 50 eBay UK SKUs (account ELECTRICALSONE) whose sales fell the most this period vs the",
        "  previous equal period, with the traffic / conversion / advertising / stock context needed to act.",
        "",
        "CONFIRMED RULES (Kobiga, 2026-08-12)",
        "  • Grain: one row per SKU (individual product code).",
        f"  • Period: CURRENT = last 30 days [{meta['cur_from']} → {meta['cur_to']}]; ",
        f"            PREVIOUS = the 30 days before [{meta['prev_from']} → {meta['prev_to']}].",
        "  • Scope: ELECTRICALSONE, eBay UK only  (orders.sub_source_id=22, market_place='23', status='Completed').",
        "  • Included only if it SOLD in the previous period AND sold LESS this period (drops only).",
        "  • Ranked by absolute £ loss (largest first); ties broken by larger Drop %. Top 50.",
        "",
        "COLUMNS",
        "  Previous / Current Sales (£) = SUM(item price × qty), completed eBay-UK orders, in each window.",
        "  Loss (£) = Current − Previous (negative).   Drop % = (Current − Previous) / Previous × 100.",
        "  CTR % = eBay views ÷ impressions (organic, EBAY-GB, current window).",
        "  CVR % = units sold ÷ eBay views (current window).",
        "  ROAS  = PPC sales ÷ PPC spend (current window); 'n/a' = SKU not advertised / £0 ad spend.",
        "  Stock = current UK-warehouse stock for the SKU.",
        "",
        "PRIORITY (PDF §6)   🔴 Critical ≥50%   🟠 High 30–49.99%   🟡 Medium 15–29.99%   🟢 Stable <15%",
        "",
        "REASON / ACTION (PDF §8 diagnosis ladder — provisional defaults, first match wins):",
        "  Out of stock → restock ·  Impressions ↓ → SEO review ·  CTR ↓ → main image/title review ·",
        "  Conversion ↓ → price/offer review ·  PPC spend ↓ → PPC review ·  else → general review.",
        "",
        "TRACEABILITY / HONESTY",
        "  Every figure is live from the raw ledsone DB; nothing is copied from the source mock-up.",
        "  'n/a' means no truthful value (e.g. unadvertised SKU) — never a guessed number.",
        "  Traffic/PPC are listing-level; for multi-SKU listings the listing's traffic is attributed to the SKU.",
        "",
        "STATUS: draft on confirmed defaults — pending Kobiga review. Not published to ph_task, not automated.",
    ]
    for i, line in enumerate(notes, start=2):
        ns.cell(row=i, column=1, value=line)
    ns.column_dimensions["A"].width = 108

    # ---- Top 50 Sales Drop ----
    ws = wb.create_sheet("Top 50 Sales Drop")
    ws.cell(row=1, column=1, value="eBay UK — Top 50 Sales Drop · ELECTRICALSONE").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value=f"Last 30 days vs previous 30 days · generated {meta['generated']} · live ledsone data").font = Font(italic=True, color="5A6B7B", size=10)
    hr = 4
    for c, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=hr, column=c, value=h)
        cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = CENTER; cell.border = BORDER
    for i, r in enumerate(rows):
        d = derive(r); rr = hr+1+i
        vals = [r["rank"], r["sku"], str(r["item_id"] or ""), r["product"],
                round(r["prev_sales"],2), round(r["cur_sales"],2), round(r["loss_gbp"],2), r["drop_pct"]/100.0,
                (d["ctr"]/100.0 if d["ctr"] is not None else "n/a"),
                (d["cvr"]/100.0 if d["cvr"] is not None else "n/a"),
                (d["roas"] if d["roas"] is not None else "n/a"),
                r["stock_uk"], d["priority"], d["reason"], d["action"]]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=rr, column=c, value=v); cell.border = BORDER
            cell.alignment = LEFT if c in (2,3,4,14,15) else CENTER
        ws.cell(row=rr, column=5).number_format = '£#,##0.00'
        ws.cell(row=rr, column=6).number_format = '£#,##0.00'
        ws.cell(row=rr, column=7).number_format = '£#,##0.00'
        ws.cell(row=rr, column=8).number_format = '0.0%'
        for cc in (9,10):
            if isinstance(ws.cell(row=rr, column=cc).value, float):
                ws.cell(row=rr, column=cc).number_format = '0.00%'
        fill = PRIO_FILL.get(d["priority"])
        if fill:
            ws.cell(row=rr, column=13).fill = PatternFill("solid", fgColor=fill)
    widths = [6,30,16,52,16,16,12,9,9,9,8,8,13,30,34]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A{hr}:{get_column_letter(len(HEADERS))}{hr+len(rows)}"

    # ---- Diagnostics (audit inputs) ----
    ds = wb.create_sheet("Diagnostics")
    dh = ["Rank","SKU","Item ID","Cur Impr","Prev Impr","Cur Views","Prev Views","Cur Units","Prev Units",
          "PPC Sales (cur)","PPC Spend (cur)","PPC Spend (prev)","PPC Sales (prev)","Stock UK"]
    for c, h in enumerate(dh, start=1):
        cell = ds.cell(row=1, column=c, value=h); cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = CENTER
    for i, r in enumerate(rows):
        vals = [r["rank"], r["sku"], str(r["item_id"] or ""), r["cur_impr"], r["prev_impr"], r["cur_views"],
                r["prev_views"], r["cur_units"], r["prev_units"], round(r["cur_ppc_sales"],2),
                round(r["cur_ppc_spend"],2), round(r["prev_ppc_spend"],2), round(r["prev_ppc_sales"],2), r["stock_uk"]]
        for c, v in enumerate(vals, start=1):
            ds.cell(row=2+i, column=c, value=v)
    for c, w in enumerate([6,30,16,11,11,11,11,10,10,14,15,16,15,9], start=1):
        ds.column_dimensions[get_column_letter(c)].width = w
    ds.freeze_panes = "A2"

    os.makedirs(OUT_DIR, exist_ok=True)
    wb.save(XLSX)

def main():
    rows = fetch()
    today = dt.date.today()
    meta = dict(generated=dt.datetime.now().strftime("%Y-%m-%d %H:%M"),
                cur_from=str(today - dt.timedelta(days=30)), cur_to=str(today),
                prev_from=str(today - dt.timedelta(days=60)), prev_to=str(today - dt.timedelta(days=30)))
    with open(PAYLOAD, "w", encoding="utf-8") as f:
        json.dump(dict(meta=meta, rows=rows), f, indent=2, default=str)
    build_xlsx(rows, meta)
    print(f"OK  rows={len(rows)}  xlsx={XLSX}")
    crit = sum(1 for r in rows if abs(r['drop_pct'])>=50)
    print(f"    critical(>=50%)={crit}  rank1={rows[0]['sku']} loss={rows[0]['loss_gbp']:.2f}")

if __name__ == "__main__":
    main()

# -*- coding: utf-8 -*-
"""Build the FINAL Excel to mirror the published HTML: per-marketplace rows, order_total, ON_SITE ad + TACOS."""
import re, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

# --- import the SAME data the published HTML uses (no drift) ---
src=open("build_html_v3.py",encoding="utf-8").read()
Rblock=re.search(r'R = \[.*?\n\]', src, re.S).group(0)
ns={}; exec(Rblock, ns); R=ns['R']
keys=["name","store","mkt","mkc","rev","ord","units","conv","ad","active","newl","stock"]
ROWS=[dict(zip(keys,r)) for r in R]
MLAB={"UK":"UK","Germany":"DE","France":"FR","Italy":"IT","Ireland":"IE","US":"US","Canada":"CA"}
MKTS=["UK","Germany","France","Italy","Ireland","US","Canada"]

# ranks
salesRank={i+1:None for i in range(len(ROWS))}
order=sorted(range(len(ROWS)), key=lambda i:-ROWS[i]['rev'][0])
srank={idx:pos+1 for pos,idx in enumerate(order)}
adidx=[i for i in range(len(ROWS)) if ROWS[i]['ad']]
prank={idx:pos+1 for pos,idx in enumerate(sorted(adidx, key=lambda i:-ROWS[i]['ad'][0]))}

OUT=r"eBay_dash_final.xlsx"
FONT="Arial"
navy=PatternFill("solid",fgColor="1E293B"); slate=PatternFill("solid",fgColor="334155"); teal=PatternFill("solid",fgColor="0D9488")
sub=PatternFill("solid",fgColor="D9E1F2"); card=PatternFill("solid",fgColor="EEF4F3")
green=PatternFill("solid",fgColor="C6EFCE"); yellow=PatternFill("solid",fgColor="FFEB9C"); red=PatternFill("solid",fgColor="FFC7CE")
note=PatternFill("solid",fgColor="FFF2CC")
wf=Font(name=FONT,bold=True,color="FFFFFF"); bold=Font(name=FONT,bold=True); reg=Font(name=FONT); ital=Font(name=FONT,italic=True,size=9,color="595959")
thin=Side(style="thin",color="BFBFBF"); bd=Border(left=thin,right=thin,top=thin,bottom=thin)
ctr=Alignment(horizontal="center",vertical="center",wrap_text=True); lft=Alignment(horizontal="left",vertical="center",wrap_text=True)
GBP='£#,##0'; GBP2='£#,##0.00'; PCT='0.00%'; NUM='#,##0'; DEC2='0.00'

wb=openpyxl.Workbook(); ws=wb.active; ws.title="Dashboard"
def C(coord,val,font=reg,fill=None,align=None,num=None,cm=None,border=True):
    c=ws[coord]; c.value=val; c.font=font
    if fill:c.fill=fill
    c.alignment=align or ctr
    if num:c.number_format=num
    if border:c.border=bd
    if cm:c.comment=Comment(cm,"note")
    return c

C("A1","eBay Account Performance Dashboard — June 2026",Font(name=FONT,bold=True,size=15,color="1E293B"),border=False,align=lft)
C("A2","Rows = account × marketplace · Sales = SUM(order_total), Completed · Ad = eBay Promoted Listings ON_SITE (Priority) only · TACOS = Ad Spend ÷ Revenue · REQ-13-D01",ital,border=False,align=lft)

# section header row 4
r=4
C(f"A{r}","Account",wf,slate,ctr); C(f"B{r}","Market",wf,slate,ctr)
ws.merge_cells(f"C{r}:Q{r}"); C(f"C{r}","SALES",wf,slate,ctr)
ws.merge_cells(f"R{r}:V{r}"); C(f"R{r}","ADVERTISING (ON_SITE)",wf,teal,ctr)
ws.merge_cells(f"W{r}:Z{r}"); C(f"W{r}","LISTINGS & STOCK",wf,slate,ctr)
# column headers row 5
r=5
H={"A":"Account / Store","B":"Mkt","C":"Revenue","D":"LM Rev","E":"LY Rev","F":"Orders","G":"LM","H":"LY","I":"Units","J":"LM","K":"LY",
 "L":"AOV","M":"LM","N":"LY","O":"Conversion","P":"LM","Q":"LY","R":"Ad Spend","S":"Ad Sales","T":"TACOS","U":"Return","V":"PPC Rank",
 "W":"Active","X":"New","Y":"Sales Rank","Z":"Stock"}
for col,t in H.items(): C(f"{col}{r}",t,bold,sub,ctr)
ws["O5"].comment=Comment("Conversion = account conversions / page-views (whole-account eBay traffic, traffic_data which_channel=2).","note")
ws["R5"].comment=Comment("eBay Promoted Listings PRIORITY / ON_SITE campaigns only (record_subtype=ON_SITE). Standard COST_PER_SALE excluded.","note")
ws["S5"].comment=Comment("Ad Sales = eBay-attributed at ON_SITE level (stays under revenue at this scope).","note")
ws["T5"].comment=Comment("TACOS = Ad Spend / total revenue. ACOS/ROAS on attributed sales omitted (attribution overlaps).","note")
ws["X5"].comment=Comment("New Listings = eBay listings created in June (ledsone listings.ebay_listings.created_at, distinct item_id).","note")
ws["Z5"].comment=Comment("Stock = warehouse units for the site's SKUs; shared across a store's marketplace rows (overlaps).","note")

r0=6
def v(x): return "" if x is None else x
for i,d in enumerate(ROWS):
    r=r0+i; rev,od,un,cv,ad=d['rev'],d['ord'],d['units'],d['conv'],d['ad']
    C(f"A{r}",f"{d['name']} ({d['store']})",bold,None,lft)
    C(f"B{r}",MLAB[d['mkt']],bold,None,ctr)
    C(f"C{r}",rev[0],reg,None,ctr,GBP); C(f"D{r}",v(rev[1]),reg,None,ctr,GBP); C(f"E{r}",v(rev[2]),reg,None,ctr,GBP)
    C(f"F{r}",od[0],reg,None,ctr,NUM); C(f"G{r}",v(od[1]),reg,None,ctr,NUM); C(f"H{r}",v(od[2]),reg,None,ctr,NUM)
    C(f"I{r}",un[0],reg,None,ctr,NUM); C(f"J{r}",v(un[1]),reg,None,ctr,NUM); C(f"K{r}",v(un[2]),reg,None,ctr,NUM)
    C(f"L{r}",f"=IF(F{r}=0,0,C{r}/F{r})",reg,None,ctr,GBP2)
    C(f"M{r}",(f"=IF(G{r}=0,0,D{r}/G{r})" if od[1] else ""),reg,None,ctr,GBP2)
    C(f"N{r}",(f"=IF(H{r}=0,0,E{r}/H{r})" if od[2] else ""),reg,None,ctr,GBP2)
    C(f"O{r}",v(cv[0]),reg,None,ctr,PCT); C(f"P{r}",v(cv[1]),reg,None,ctr,PCT); C(f"Q{r}",v(cv[2]),reg,None,ctr,PCT)
    if ad:
        C(f"R{r}",ad[0],reg,None,ctr,GBP); C(f"S{r}",ad[1],reg,None,ctr,GBP)
        C(f"T{r}",f"=IF(C{r}=0,0,R{r}/C{r})",reg,None,ctr,PCT); C(f"U{r}",f"=IF(R{r}=0,0,C{r}/R{r})",reg,None,ctr,DEC2)
        C(f"V{r}",f"#{prank[i]}",reg,None,ctr)
    else:
        for col in ("R","S","T","U","V"): C(f"{col}{r}","—",ital,None,ctr)
    C(f"W{r}",d['active'],reg,None,ctr,NUM); C(f"X{r}",(d['newl'] if d['newl'] else 0),reg,None,ctr,NUM)
    C(f"Y{r}",f"#{srank[i]}",reg,None,ctr); C(f"Z{r}",d['stock'],reg,None,ctr,NUM)
    # RAG
    if rev[1]:
        g=(rev[0]-rev[1])/rev[1]; ws[f"C{r}"].fill=green if g>0.10 else (yellow if g>=0 else red)
    if cv[0] is not None: ws[f"O{r}"].fill=green if cv[0]>0.045 else (yellow if cv[0]>=0.03 else red)
    if ad:
        tac=ad[0]/rev[0]; ret=rev[0]/ad[0]
        ws[f"T{r}"].fill=green if tac<0.12 else (yellow if tac<=0.18 else red)
        ws[f"U{r}"].fill=green if ret>8 else (yellow if ret>=5 else red)

rt=r0+len(ROWS); last=rt-1
C(f"A{rt}","TOTAL — 22 rows",wf,navy,ctr); C(f"B{rt}","",wf,navy,ctr)
for col in ("C","D","E","F","G","H","I","J","K","R","S","W","X","Z"):
    C(f"{col}{rt}",f"=SUM({col}{r0}:{col}{last})",wf,sub,ctr,GBP if col in("C","D","E","R","S") else NUM)
C(f"L{rt}",f"=IF(F{rt}=0,0,C{rt}/F{rt})",wf,sub,ctr,GBP2)
C(f"T{rt}",f"=IF(C{rt}=0,0,R{rt}/C{rt})",wf,sub,ctr,PCT); C(f"U{rt}",f"=IF(R{rt}=0,0,C{rt}/R{rt})",wf,sub,ctr,DEC2)
for col in ("M","N","O","P","Q","V","Y"): C(f"{col}{rt}","",wf,sub,ctr)

ws.column_dimensions["A"].width=26
for col in ["C","D","E","R","S","Z","W"]: ws.column_dimensions[col].width=12
for col in ["B","F","G","H","I","J","K","L","M","N","O","P","Q","T","U","V","X","Y"]: ws.column_dimensions[col].width=10
for rr in range(r0,rt+1): ws.row_dimensions[rr].height=26
ws.freeze_panes="C6"

# ===== Sheet 2: By Marketplace =====
ws2=wb.create_sheet("By Marketplace"); ws=ws2
C("A1","By Marketplace — all accounts combined (June 2026)",Font(name=FONT,bold=True,size=13,color="0D9488"),border=False,align=lft)
hdr=["Marketplace","Revenue","Orders","Units","Ad Spend (ON_SITE)","Ad Sales","TACOS","Return","Rows"]
for j,h in enumerate(hdr): C(f"{get_column_letter(1+j)}3",h,wf,teal,ctr)
MS={m:[0,0,0,0.0,0.0,0] for m in MKTS}
for d in ROWS:
    s=MS[d['mkt']]; s[0]+=d['rev'][0]; s[1]+=d['ord'][0]; s[2]+=d['units'][0]
    if d['ad']: s[3]+=d['ad'][0]; s[4]+=d['ad'][1]
    s[5]+=1
for k,m in enumerate(MKTS):
    r=4+k; s=MS[m]
    C(f"A{r}",f"{MLAB[m]} — {m}",bold,None,lft); C(f"B{r}",round(s[0],2),reg,None,ctr,GBP); C(f"C{r}",s[1],reg,None,ctr,NUM); C(f"D{r}",s[2],reg,None,ctr,NUM)
    C(f"E{r}",(round(s[3],2) if s[3] else "—"),reg,None,ctr,GBP if s[3] else None)
    C(f"F{r}",(round(s[4],2) if s[4] else "—"),reg,None,ctr,GBP if s[4] else None)
    C(f"G{r}",(f"=IF(B{r}=0,0,E{r}/B{r})" if s[3] else "—"),reg,None,ctr,PCT if s[3] else None)
    C(f"H{r}",(f"=IF(E{r}=0,0,B{r}/E{r})" if s[3] else "—"),reg,None,ctr,DEC2 if s[3] else None)
    C(f"I{r}",int(s[5]),reg,None,ctr,NUM)
for col,w in {"A":16,"B":13,"C":10,"D":10,"E":15,"F":12,"G":10,"H":10,"I":8}.items(): ws.column_dimensions[col].width=w

# ===== Sheet 3: Definitions =====
ws3=wb.create_sheet("Definitions"); ws=ws3
C("A1","DEFINITIONS, DATA SOURCES & METHOD",Font(name=FONT,bold=True,size=13,color="1E293B"),border=False,align=lft)
defs=[
 "Reporting month = June 2026 (order_date >= 2026-06-01 AND < 2026-07-01). LM = May 2026, LY = June 2025.",
 "Rows = account (ss_name) × marketplace (order_transaction.market_place). A store sells cross-border; each row = that store's sales to one marketplace's buyers.",
 "Revenue = SUM(order_total) on source_name='EBAY', order_status='Completed' — eBay's settled paid value incl. real postage (NOT item_price*qty, NOT +shipping_template_price). AOV = Revenue/Orders.",
 "Conversion = account conversions / page-views (whole-account eBay traffic, traffic_data which_channel=2). Blank where no traffic.",
 "Ad Spend / Ad Sales = eBay Promoted Listings PRIORITY / ON_SITE campaigns only (ppc_performance joined to ppc where record_subtype='ON_SITE'; join record_id=parent_id). Standard COST_PER_SALE fees excluded per Thinesh.",
 "TACOS = Ad Spend / total revenue (real efficiency). Return = revenue / Ad Spend. ACOS/ROAS on eBay-attributed sales are omitted (one order is attributed to every overlapping campaign, so attributed sales over-count).",
 "Active Listings = distinct ref_id (listing_data, ebay, per marketplace). New Listings = distinct item_id created in June (ledsone DB listings.ebay_listings.created_at). Stock = SUM(inv_final_stock) for the site's SKUs (shared/overlapping).",
 "Sales Rank = rows ranked by June revenue. PPC Rank = ad rows ranked by ON_SITE spend.",
 "Account mapping (Thinesh-confirmed): LEDSONE UK=led_sone, SUNSONE UK=so_926407, Electricalsone UK=electricalsone, LEDSONE DE=ledsonede.",
 "OPEN: orders = COUNT(DISTINCT order_id) (led_sone UK 1,517) — team reference showed 1,619 = COUNT(*) line count; pending confirmation. Conversion RAG threshold (green >4.5%) predates whole-account conversion (~2-3%) — recalibration pending.",
 "Read-only against all warehouse + ledsone tables. Published HTML twin live in tech_team_outputs.ph_task (Thinesh, Jarsini, kobiga, powsteena).",
]
for k,t in enumerate(defs):
    r=3+k; ws.merge_cells(f"A{r}:H{r}"); C(f"A{r}","• "+t,reg,note,lft)
ws.column_dimensions["A"].width=140

wb.save(OUT); print("SAVED",OUT,"| rows:",len(ROWS),"| total row:",rt)

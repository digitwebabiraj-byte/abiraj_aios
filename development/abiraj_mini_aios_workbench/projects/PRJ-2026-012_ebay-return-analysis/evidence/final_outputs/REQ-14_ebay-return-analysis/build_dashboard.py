#!/usr/bin/env python3
"""
Build the eBay Return Analysis dashboard (.xlsx) from the SQL query outputs.

INPUTS (tab-separated, NO header row, NULLs as empty string):
  main.tsv              -> statement 1 of ebay_return_analysis.sql (19 columns, in order)
  reason_breakdown.tsv  -> statement 2 (Return Reason<TAB>Returns<TAB>Pct); Pct col optional

USAGE:
  python build_dashboard.py main.tsv reason_breakdown.tsv output.xlsx

Requires: openpyxl (pip install openpyxl). After writing, recalculate with LibreOffice
so cached formula values populate, e.g. the xlsx skill's scripts/recalc.py.

The 19 columns MUST arrive already labelled by the SQL (friendly Account, mapped
Main Return Reason, Return Rank as "#n"). This script only formats; it does not remap.
"""
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MAIN = sys.argv[1] if len(sys.argv) > 1 else "main.tsv"
BRK  = sys.argv[2] if len(sys.argv) > 2 else "reason_breakdown.tsv"
OUT  = sys.argv[3] if len(sys.argv) > 3 else "eBay_Return_Analysis.xlsx"
PERIOD_LABEL = "June 2026"   # cosmetic: update to match the query's reporting period

HEADERS = ["SKU","Product Title","Account","Orders","Returns","Return Rate",
    "Last Month Returns","Last Year Returns","Refund (£)","Return Cost (£)",
    "Main Return Reason","Return Rank","Negative Feedback","Open Cases","Stock",
    "Ad Spend (£)","Ad Sales (£)","ACOS","ROAS"]

# Number formats — zeros show as 0 / £0.00 / 0.0% (NOT dashes). Genuinely-null cells
# (empty string in the TSV) stay blank on purpose: Return Rate w/ no orders, ACOS/ROAS
# w/ no ad activity.
CUR='£#,##0.00'; PCT='0.0%'; INT='#,##0'; ROAS='0.00"x"'
FMT={3:INT,4:INT,5:PCT,6:INT,7:INT,8:CUR,9:CUR,12:INT,13:INT,14:INT,15:CUR,16:CUR,17:PCT,18:ROAS}

# Static spec content (fixed by the dashboard mockup)
FILTERS=[("Date Range","Today, Yesterday, Last 7 Days, Last 30 Days, This Month, Last Month, Last 90 Days, Last Year, Custom"),
    ("Account","All Accounts, plus each eBay store"),("Return Status","All, Open, Approved, Refunded, Closed, Rejected"),
    ("Return Reason","All Reasons"),("SKU","Search SKU"),("Category","All Categories"),("Brand","All Brands")]
EFF=[("Daily Return Monitoring","15 min","3 min","93% Faster"),("Return Report Generation","20 min","Real-Time","100% Automated"),
    ("Return Reason Analysis","30 min","5 min","94% Faster"),("Refund Cost Analysis","15 min","Instant","100% Automated"),
    ("SKU Return Investigation","15 min","4 min","91% Faster"),("Account Return Review","10 min","3 min","93% Faster")]

def read_tsv(path):
    out=[]
    with open(path, encoding="utf-8") as f:
        for line in f:
            line=line.rstrip("\n")
            if line.strip(): out.append(line.split("\t"))
    return out

def num(v):
    if v in (None,""): return None
    try:
        fv=float(v); return int(fv) if fv.is_integer() else fv
    except ValueError: return v

rows=read_tsv(MAIN)
brk=read_tsv(BRK)   # [label, count, (pct)]

NAVY="1F3A5F"; STEEL="2E5A88"; BAND="F5F8FB"; F="Arial"; ncol=len(HEADERS)
BORDER=Side(style="thin",color="D3DCE6"); thin=Border(left=BORDER,right=BORDER,top=BORDER,bottom=BORDER)
wb=Workbook(); ws=wb.active; ws.title="eBay Return Analysis"; ws.sheet_view.showGridLines=False

ws["A1"]="eBay Return Analysis Dashboard"; ws["A1"].font=Font(name=F,size=16,bold=True,color="FFFFFF")
ws["A2"]=f"Reporting period: {PERIOD_LABEL}  ·  Scope: all eBay stores & marketplaces  ·  One row per variant SKU with a return"
ws["A2"].font=Font(name=F,size=9,color="FFFFFF")
ws["A3"]="Last updated: "+datetime.now().strftime("%Y-%m-%d %H:%M")+"  ·  Source: live Ledsone PostgreSQL  ·  [ Manual Refresh ]"
ws["A3"].font=Font(name=F,size=9,color="FFFFFF")
for r in (1,2,3):
    for c in range(1,ncol+1): ws.cell(r,c).fill=PatternFill("solid",fgColor=NAVY)
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=ncol)
ws.row_dimensions[1].height=24

HDR=5
for i,h in enumerate(HEADERS,start=1):
    c=ws.cell(HDR,i,h); c.font=Font(name=F,size=9,bold=True,color="FFFFFF"); c.fill=PatternFill("solid",fgColor=NAVY)
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=thin
ws.row_dimensions[HDR].height=32

start=HDR+1
for ri,row in enumerate(rows):
    xr=start+ri
    for ci in range(ncol):
        raw=row[ci] if ci<len(row) else ""
        val=raw if ci in (0,1,2,10,11) else num(raw)
        c=ws.cell(xr,ci+1,val); c.font=Font(name=F,size=9,bold=(ci==0)); c.border=thin
        if ci in FMT: c.number_format=FMT[ci]
        c.alignment=Alignment(horizontal="left" if ci in (0,1,2,10) else "center")
    if ri%2==1:
        for ci in range(ncol): ws.cell(xr,ci+1).fill=PatternFill("solid",fgColor=BAND)
last=start+len(rows)-1
for xr in range(start,last+1):
    rr=ws.cell(xr,6).value
    if isinstance(rr,(int,float)) and rr>=0.5:
        ws.cell(xr,6).fill=PatternFill("solid",fgColor="FBE3E1"); ws.cell(xr,6).font=Font(name=F,size=9,bold=True,color="B03A2E")
    if ws.cell(xr,15).value==0:
        ws.cell(xr,15).fill=PatternFill("solid",fgColor="FCEEDB"); ws.cell(xr,15).font=Font(name=F,size=9,bold=True,color="B9770E")

tot=last+1
ws.cell(tot,1,"TOTAL / AVG"); ws.cell(tot,3,"%d SKUs"%len(rows))
for ci in [4,5,8,9,13,15,16]:
    L=get_column_letter(ci+1); ws.cell(tot,ci+1,f"=SUM({L}{start}:{L}{last})")
ws.cell(tot,6,f"=IFERROR(SUM(E{start}:E{last})/SUM(D{start}:D{last}),0)"); ws.cell(tot,6).number_format=PCT
ws.cell(tot,18,f"=IFERROR(P{tot}/Q{tot},0)"); ws.cell(tot,18).number_format=PCT
ws.cell(tot,19,f"=IFERROR(Q{tot}/P{tot},0)"); ws.cell(tot,19).number_format=ROAS
for ci in range(ncol):
    c=ws.cell(tot,ci+1); c.fill=PatternFill("solid",fgColor=STEEL); c.font=Font(name=F,size=9,bold=True,color="FFFFFF"); c.border=thin
    if ci in FMT: c.number_format=FMT[ci]
ws.cell(tot,4).number_format=INT; ws.cell(tot,5).number_format=INT
ws.cell(tot,1).alignment=Alignment(horizontal="left"); ws.cell(tot,3).alignment=Alignment(horizontal="left")

# Return Reason breakdown + Filter Options (mockup layout)
sec=tot+3
ws.cell(sec,1,f"Return Reason Breakdown — {PERIOD_LABEL}").font=Font(name=F,size=12,bold=True,color=NAVY)
ws.cell(sec,5,"Filter Options").font=Font(name=F,size=12,bold=True,color=NAVY)
rh=sec+1
for i,h in enumerate(["Return Reason","Returns","% of Returns"],start=1):
    c=ws.cell(rh,i,h); c.font=Font(name=F,size=9,bold=True,color="FFFFFF"); c.fill=PatternFill("solid",fgColor=NAVY); c.border=thin
    c.alignment=Alignment(horizontal="left" if i==1 else "center")
for i,h in enumerate(["Filter","Options"],start=5):
    c=ws.cell(rh,i,h); c.font=Font(name=F,size=9,bold=True,color="FFFFFF"); c.fill=PatternFill("solid",fgColor=NAVY); c.border=thin
rbs=rh+1
for j,br in enumerate(brk):
    r=rbs+j; label=br[0]; cnt=num(br[1])
    ws.cell(r,1,label).font=Font(name=F,size=9); ws.cell(r,1).border=thin
    ws.cell(r,2,cnt).font=Font(name=F,size=9); ws.cell(r,2).border=thin; ws.cell(r,2).alignment=Alignment(horizontal="center")
    ws.cell(r,3,f"=B{r}/$B${rbs+len(brk)}"); ws.cell(r,3).number_format=PCT; ws.cell(r,3).border=thin; ws.cell(r,3).alignment=Alignment(horizontal="center")
    if j%2==1:
        for cc in (1,2,3): ws.cell(r,cc).fill=PatternFill("solid",fgColor=BAND)
rbt=rbs+len(brk)
ws.cell(rbt,1,"Total"); ws.cell(rbt,2,f"=SUM(B{rbs}:B{rbt-1})"); ws.cell(rbt,3,f"=B{rbt}/$B${rbt}"); ws.cell(rbt,3).number_format=PCT
for cc in (1,2,3):
    ws.cell(rbt,cc).fill=PatternFill("solid",fgColor=STEEL); ws.cell(rbt,cc).border=thin; ws.cell(rbt,cc).font=Font(name=F,size=9,bold=True,color="FFFFFF")
ws.cell(rbt,2).alignment=Alignment(horizontal="center"); ws.cell(rbt,3).alignment=Alignment(horizontal="center")
for j,(fname,opts) in enumerate(FILTERS):
    r=rh+1+j
    ws.cell(r,5,fname).font=Font(name=F,size=9,bold=True); ws.cell(r,5).border=thin; ws.cell(r,5).alignment=Alignment(vertical="top")
    ws.cell(r,6,opts).font=Font(name=F,size=9); ws.cell(r,6).border=thin; ws.cell(r,6).alignment=Alignment(wrap_text=True,vertical="top")
    ws.merge_cells(start_row=r,start_column=6,end_row=r,end_column=ncol)

ke=rbt+3
ws.cell(ke,1,"Return Workflow — Before / After").font=Font(name=F,size=12,bold=True,color=NAVY)
keh=ke+1
for i,h in enumerate(["KPI","Before","After","Improvement"],start=1):
    c=ws.cell(keh,i,h); c.font=Font(name=F,size=9,bold=True,color="FFFFFF"); c.fill=PatternFill("solid",fgColor=NAVY); c.border=thin
    c.alignment=Alignment(horizontal="left" if i==1 else "center")
for j,(a,b,cc,d) in enumerate(EFF):
    r=keh+1+j
    for i,v in enumerate((a,b,cc,d),start=1):
        c=ws.cell(r,i,v); c.font=Font(name=F,size=9,bold=(i==1)); c.border=thin
        c.alignment=Alignment(horizontal="left" if i==1 else "center")
    ws.cell(r,4).font=Font(name=F,size=9,bold=True,color="1E7A46")
    if j%2==1:
        for i in range(1,5): ws.cell(r,i).fill=PatternFill("solid",fgColor=BAND)

for i,w in enumerate([22,52,16,8,8,11,13,13,11,12,20,9,10,9,8,12,12,9,8],start=1):
    ws.column_dimensions[get_column_letter(i)].width=w
ws.freeze_panes="C6"

wb.save(OUT)
print(f"wrote {OUT}: {len(rows)} SKU rows, {len(brk)} reason rows. Now recalc with LibreOffice.")

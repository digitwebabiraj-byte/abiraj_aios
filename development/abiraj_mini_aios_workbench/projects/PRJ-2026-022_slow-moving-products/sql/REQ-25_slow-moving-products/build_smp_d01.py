# -*- coding: utf-8 -*-
"""REQ-25-D01 Slow Moving Products (smp) - Mahima / Germany.
Inverse of Fast Moving #020, built to the SAME house pattern: a styled Excel
(Notes & Method sheet + banded data sheet with title banner, red 'never sold'
highlight) AND a self-contained light-theme HTML dashboard (hero, KPI cards,
search + Reason/Recency filters, sortable sticky-header table, CSV, full screen).

Numbers 100% RAW mcp.ledsone (order_management + inventory) via LED_* env creds.
Reason/Action = PROVISIONAL default rules (documented on the Notes tab / banner)
pending Mahima's sign-off.

Emits (into evidence/final_outputs/REQ-25_slow-moving-products/):
  REQ-25-D01_slow_moving_products.xlsx
  REQ-25-D01_slow_moving_products.html
  + smp_payload.json snapshot (next to this script)
"""
import os, json, datetime, psycopg2
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT  = os.path.abspath(os.path.join(HERE, "..", "..", "evidence", "final_outputs", "REQ-25_slow-moving-products"))

# Germany-stocked SKUs (universe) — sku -> title, stock
SQL_STOCK = r"""
SELECT p.sku, MAX(p.title) title, SUM(COALESCE(s.stock,0))::int stock
FROM inventory.products p
JOIN inventory.local_inventory_current_stock_location_wise s ON s.inventory_id=p.id
WHERE s.warehouse_location='Germany'
GROUP BY p.sku HAVING SUM(COALESCE(s.stock,0))>0;
"""

# Per (SKU, channel) sales: last sale (all-time on that channel) + rolling 30/90d units.
SQL_SALES = r"""
SELECT oi.item_sku sku, ss.source_id chan,
  MAX(o.order_date::date) last_sale,
  SUM(CASE WHEN o.order_date::date>=CURRENT_DATE-30 THEN COALESCE(NULLIF(oi.item_quantity,'')::numeric,0) ELSE 0 END)::int q30,
  SUM(CASE WHEN o.order_date::date>=CURRENT_DATE-90 THEN COALESCE(NULLIF(oi.item_quantity,'')::numeric,0) ELSE 0 END)::int q90
FROM order_management.orders o
JOIN order_management.sub_source ss ON ss.id=o.sub_source_id
JOIN order_management.order_item_info oi ON oi.order_id=o.id
WHERE o.market_place='10' AND o.status='Completed' AND ss.source_id IN (1,2,3)
  AND oi.item_sku IS NOT NULL AND oi.item_sku<>''
GROUP BY oi.item_sku, ss.source_id;
"""
CHAN = {1: "amazon", 2: "ebay", 3: "shopify"}

def clean_name(sku, title):
    if not title or title.strip().lower().startswith("combo default"):
        return sku  # combo placeholder trap -> fall back to SKU
    return title.strip()

def rule_engine(stock, q90, dws):
    """PROVISIONAL default Reason/Action rules (pending Mahima). Returns (reason, action, cls)."""
    if dws is None:
        return ("No sales history (dead stock)",
                "Clearance / liquidate" if stock >= 100 else "Review / delist", "never")
    if q90 == 0 and stock >= 100:
        return ("High stock, no demand in 90 days", "Clearance / bundle", "high")
    if q90 == 0:
        return ("No demand in 90 days", "Create bundle / promote", "dead90")
    return ("Slowing down (no sale in 30 days)", "Improve listing / promote", "slowing")

def _row(sku, title, stock, last_sale, q30, q90, today):
    dws = None if last_sale is None else (today - last_sale).days
    reason, action, cls = rule_engine(stock, q90, dws)
    return {"sku": sku, "title": title, "stock": stock,
            "last_sale": last_sale.isoformat() if last_sale else None,
            "q30": q30, "q90": q90, "dws": dws,
            "reason": reason, "action": action, "cls": cls}

def fetch(today):
    conn = psycopg2.connect(host=os.environ["LED_PGHOST"], user=os.environ["LED_PGUSER"],
        password=os.environ["LED_PGPASSWORD"], dbname=os.environ.get("LED_PGDATABASE","ledsone"),
        port=os.environ.get("LED_PGPORT","5432"), connect_timeout=30)
    try:
        cur = conn.cursor()
        cur.execute(SQL_STOCK); stock_rows = cur.fetchall()
        cur.execute(SQL_SALES); sales_rows = cur.fetchall()
    finally:
        conn.close()

    stk = {sku: (clean_name(sku, title), stock) for sku, title, stock in stock_rows}
    # per-channel sales keyed by (sku, chan); combined aggregate keyed by sku
    perchan = {}                      # (sku,chan) -> [last_sale, q30, q90]
    comb = {}                         # sku -> [last_sale, q30, q90]
    for sku, chan, last_sale, q30, q90 in sales_rows:
        if sku not in stk:            # only SKUs holding German stock
            continue
        perchan[(sku, chan)] = [last_sale, q30, q90]
        c = comb.setdefault(sku, [None, 0, 0])
        if last_sale and (c[0] is None or last_sale > c[0]): c[0] = last_sale
        c[1] += q30; c[2] += q90

    data = {"amazon": [], "ebay": [], "shopify": [], "combined": []}
    # channel tabs: sold on that channel before, but 0 units there in last 30 days
    for (sku, chan), (last_sale, q30, q90) in perchan.items():
        if q30 != 0:
            continue
        title, stock = stk[sku]
        data[CHAN[chan]].append(_row(sku, title, stock, last_sale, q30, q90, today))
    # combined tab: not selling ANYWHERE in last 30d (incl. never-sold dead stock)
    for sku, (title, stock) in stk.items():
        c = comb.get(sku, [None, 0, 0])
        if c[1] != 0:
            continue
        data["combined"].append(_row(sku, title, stock, c[0], c[1], c[2], today))

    for k in data:                    # sort like FMP: biggest tied-up stock first
        data[k].sort(key=lambda r: (-r["stock"], -(r["dws"] if r["dws"] is not None else 10**9)))
    return data

# ============================== EXCEL ==============================
FONT = "Arial"
TITLE_F   = Font(name=FONT, size=14, bold=True, color="FFFFFF")
SUB_F     = Font(name=FONT, size=9,  italic=True, color="4B2A6A")
HEAD_F    = Font(name=FONT, size=10, bold=True, color="FFFFFF")
CELL_F    = Font(name=FONT, size=10)
HEAD_FILL = PatternFill("solid", fgColor="4C1D95")   # deep violet
TITLE_FILL= PatternFill("solid", fgColor="6D28D9")
BAND_FILL = PatternFill("solid", fgColor="F3EEFC")
NEVER_FILL= PatternFill("solid", fgColor="FDE2E4")   # never-sold highlight
thin = Side(style="thin", color="D6C9F0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CEN = Alignment(horizontal="center", vertical="center")
LEFT= Alignment(horizontal="left", vertical="center")
INT = '#,##0'

HEADERS = ["SKU","Product Name","Stock Qty","Last Sale Date","Sold Qty (30 Days)",
           "Sold Qty (90 Days)","Days Without Sale","Reason","Action"]

def data_sheet(wb, sheet_name, title_txt, sub_txt, rows):
    ws = wb.create_sheet(sheet_name)
    ncols = len(HEADERS)
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=ncols)
    t = ws.cell(row=1,column=1,value=title_txt)
    t.font=TITLE_F; t.fill=TITLE_FILL; t.alignment=Alignment(horizontal="left",vertical="center")
    ws.row_dimensions[1].height=24
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=ncols)
    ws.cell(row=2,column=1, value=sub_txt).font=SUB_F
    hr = 3
    for i,h in enumerate(HEADERS,1):
        c = ws.cell(row=hr,column=i,value=h); c.font=HEAD_F; c.fill=HEAD_FILL; c.alignment=CEN; c.border=BORDER
    r = hr+1
    center_cols = {3,5,6,7}
    for row in rows:
        vals = [row["sku"], row["title"], row["stock"], row["last_sale"] or "",
                row["q30"], row["q90"],
                ("Never" if row["dws"] is None else row["dws"]),
                row["reason"], row["action"]]
        for i,v in enumerate(vals,1):
            cell = ws.cell(row=r,column=i,value=v); cell.border=BORDER; cell.font=CELL_F
            if i in (3,5,6) or (i==7 and isinstance(v,int)): cell.number_format=INT; cell.alignment=CEN
            elif i in center_cols: cell.alignment=CEN
            else: cell.alignment=LEFT
        band = NEVER_FILL if row["dws"] is None else (BAND_FILL if (r-hr)%2==0 else None)
        if band:
            for i in range(1,ncols+1): ws.cell(row=r,column=i).fill=band
        r += 1
    ws.freeze_panes=ws.cell(row=hr+1,column=1)
    ws.auto_filter.ref=f"A{hr}:{get_column_letter(ncols)}{r-1}"
    widths=[16,46,11,14,15,15,16,32,26]
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
    return ws

def notes_sheet(wb, data, meta):
    ws=wb.create_sheet("Notes & Method",0)
    ws.column_dimensions['A'].width=26; ws.column_dimensions['B'].width=110
    def put(a,b,head=False):
        r=ws.max_row+1 if (ws.max_row>1 or ws.cell(1,1).value) else 1
        ca=ws.cell(row=r,column=1,value=a)
        if head: ca.font=Font(name=FONT,size=12,bold=True,color="4C1D95")
        else: ca.font=Font(name=FONT,size=10,bold=True,color="4B2A6A")
        cb=ws.cell(row=r,column=2,value=b); cb.alignment=Alignment(wrap_text=True,vertical="top")
        ws.row_dimensions[r].height=max(15,14*(1+len(str(b))//95))
    rows = data["combined"]
    never=sum(1 for r in rows if r["dws"] is None); z90=sum(1 for r in rows if r["q90"]==0)
    put("REQ-25-D01 Slow Moving Products",
        "Products holding stock in the Germany warehouse that are NOT selling. Inverse of Fast Moving "
        "Products (#020). Prepared for Mahima.", head=True)
    put("Tabs","Shopify DE / Amazon DE / eBay DE = SKUs that sold on that channel before but 0 units there "
               "in the last 30 days (sales figures are channel-only). Combined = SKUs with 0 units sold on ANY "
               "channel in the last 30 days, including never-sold dead stock (sales figures all-channel).")
    put("Scope","Germany (DE), 3 channels (Amazon=1 / eBay=2 / Shopify=3). Order status = Completed. "
                 "One row per SKU; on channel tabs the sales are that channel's, on Combined they are summed.")
    put("Row counts",
        f"Shopify DE {len(data['shopify']):,}  ·  Amazon DE {len(data['amazon']):,}  ·  "
        f"eBay DE {len(data['ebay']):,}  ·  Combined {len(rows):,}. Sorted by Stock Qty descending.")
    put("Slow-moving rule [DEFAULT – confirm]",
        "A SKU is 'slow moving' if it holds German stock (>0) AND sold 0 units in the last 30 days "
        "(on the channel for a channel tab; on any channel for Combined).")
    put("Windows","Last 30 Days / Last 90 Days sales are rolling windows ending yesterday. Last Sale Date "
                  "and Days Without Sale use ALL-TIME history.")
    put("Data sources","RAW mcp.ledsone DB (read-only). Sales/units: order_management.orders + "
        "order_item_info + sub_source (source_id 1/2/3), market_place='10' (Germany), status='Completed'. "
        "Stock: inventory.products + inventory.local_inventory_current_stock_location_wise, "
        "warehouse_location='Germany'.")
    put("Days Without Sale","= today − Last Sale Date (all-time). 'Never' = no completed sale on record "
                            "for that SKU.")
    put("Reason [DEFAULT RULE – confirm]",
        "Never sold → 'No sales history (dead stock)'; 90-day sales = 0 & stock ≥ 100 → 'High "
        "stock, no demand in 90 days'; 90-day sales = 0 & stock < 100 → 'No demand in 90 days'; sold in "
        "90d but not last 30d → 'Slowing down (no sale in 30 days)'.")
    put("Action [DEFAULT RULE – confirm]",
        "Never sold & stock ≥ 100 → Clearance / liquidate; never sold & stock < 100 → Review / "
        "delist; high stock no demand → Clearance / bundle; no demand → Create bundle / promote; "
        "slowing → Improve listing / promote.")
    put("Known caveats","(1) Reason/Action thresholds above are Claude's documented DEFAULTS, NOT yet agreed "
        "by Mahima. (2) Combo SKUs whose inv title reads 'Combo Default Title.' fall back to the SKU as the "
        "name. (3) Stock is live 'as of today', not as-of any window. (4) This report has no money column.")
    put("Rows in red",f"Highlighted rows have NEVER sold on record (Combined: {never:,} of {len(rows):,}; "
                      f"{z90:,} have zero 90-day sales).")
    return ws

CHAN_LABEL = {"amazon": "Amazon DE", "ebay": "eBay DE", "shopify": "Shopify DE"}

def build_xlsx(data, meta):
    wb=Workbook(); wb.remove(wb.active)
    g = meta["generated"]
    for key in ("shopify", "amazon", "ebay"):
        rows = data[key]; lbl = CHAN_LABEL[key]
        data_sheet(wb, lbl, f"Slow Moving Products – {lbl} (Germany)",
            (f"SKUs holding German stock that sold on {lbl} before but 0 units there in the last 30 days"
             f"  |  {len(rows):,} products  |  Sorted by Stock Qty desc  |  "
             f"Sales figures are {lbl}-only  |  Live data pulled {g}"), rows)
    rows = data["combined"]
    data_sheet(wb, "Combined", "Slow Moving Products – Combined / All Channels (Germany)",
        (f"SKUs holding German stock with 0 units sold on ANY channel in the last 30 days (incl. never-sold "
         f"dead stock)  |  {len(rows):,} products  |  Sorted by Stock Qty desc  |  Sales figures are all-channel"
         f"  |  Live data pulled {g}"), rows)
    notes_sheet(wb, data, meta)       # Notes & Method inserted at index 0
    path=os.path.join(OUT,"REQ-25-D01_slow_moving_products.xlsx"); wb.save(path); return path

# ============================== HTML ==============================
def build_html(data, meta):
    def pack(rows):
        return [{"sku":r["sku"],"title":r["title"] or "","stock":r["stock"],
                 "last":r["last_sale"] or "","q30":r["q30"],"q90":r["q90"],
                 "dws":r["dws"],"reason":r["reason"],"action":r["action"],"cls":r["cls"]} for r in rows]
    blob=json.dumps({k:pack(v) for k,v in data.items()}, ensure_ascii=False)
    html=r"""<!doctype html><html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Slow Moving Products — Germany · REQ-25-D01</title>
<style>
:root{--bg:#f6f4fb;--card:#fff;--ink:#241b33;--muted:#6b5f7d;--line:#ece5f6;
--brand:#6d28d9;--brand2:#8b5cf6;--rose:#e11d48;--amber:#f59e0b;
--never:#e11d48;--neverbg:#fff0f3;--zebra:#faf7fe;--hover:#f3ecff;
--shadow:0 10px 34px rgba(60,20,90,.10),0 2px 8px rgba(60,20,90,.05);--radius:14px;}
*{box-sizing:border-box}html,body{margin:0}
body{font-family:'Segoe UI',Roboto,Helvetica,Arial,system-ui,sans-serif;color:var(--ink);
background:radial-gradient(1000px 460px at 8% -10%,rgba(139,92,246,.16),transparent 55%),
radial-gradient(900px 460px at 100% -6%,rgba(225,29,72,.10),transparent 52%),var(--bg);
background-attachment:fixed;min-height:100vh;padding:8px clamp(8px,1.4vw,18px);-webkit-font-smoothing:antialiased}
.wrap{width:100%;margin:0}
header.hero{background:linear-gradient(115deg,#4c1d95 0%,#6d28d9 50%,#8b5cf6 110%);color:#fff;border-radius:11px;
padding:7px 16px;box-shadow:0 6px 18px rgba(109,40,217,.24);display:flex;align-items:center;justify-content:space-between;
gap:12px;flex-wrap:wrap;animation:drop .5s ease}
.hero h1{margin:0;font-size:16px;font-weight:800}
.hero .sub{font-size:11px;opacity:.9;margin-top:2px}
.badges{display:flex;gap:6px;flex-wrap:wrap;margin-top:4px}
.pill{background:rgba(255,255,255,.16);border:1px solid rgba(255,255,255,.30);padding:2px 9px;border-radius:999px;font-size:10.5px;font-weight:600}
.fsbtn{background:rgba(255,255,255,.18);border:1px solid rgba(255,255,255,.42);color:#fff;border-radius:8px;
padding:6px 12px;font-size:11.5px;font-weight:700;cursor:pointer;transition:.2s}.fsbtn:hover{background:rgba(255,255,255,.32)}
.kpis{display:grid;grid-template-columns:repeat(4,1fr);gap:8px;margin:7px 0}
.kpi{border-radius:9px;padding:6px 12px;color:#fff;box-shadow:var(--shadow);min-height:44px;
display:flex;flex-direction:column;justify-content:center;animation:rise .45s ease both}
.kpi.k0{background:linear-gradient(135deg,#5b21b6,#8b5cf6)}.kpi.k1{background:linear-gradient(135deg,#7c3aed,#a78bfa)}
.kpi.k2{background:linear-gradient(135deg,#b45309,#f59e0b)}.kpi.k3{background:linear-gradient(135deg,#be123c,#fb7185)}
.kpi:nth-child(2){animation-delay:.05s}.kpi:nth-child(3){animation-delay:.1s}.kpi:nth-child(4){animation-delay:.15s}
.kpi .lab{font-size:9.5px;font-weight:700;text-transform:uppercase;letter-spacing:.4px;opacity:.9}
.kpi .val{font-size:18px;font-weight:800;font-variant-numeric:tabular-nums}
.kpi .s{font-size:9.5px;opacity:.85}
.tabs{display:flex;gap:6px;flex-wrap:wrap;margin:0 0 6px}
.tab{background:var(--card);border:1px solid var(--line);color:var(--muted);padding:6px 15px;border-radius:9px;
font-weight:700;font-size:12.5px;cursor:pointer;transition:.18s;box-shadow:0 1px 4px rgba(60,20,90,.05)}
.tab:hover{color:var(--brand);border-color:var(--brand2)}
.tab.active{background:linear-gradient(120deg,var(--brand),var(--brand2));color:#fff;border-color:transparent;box-shadow:0 8px 18px rgba(109,40,217,.32)}
.tab .n{opacity:.7;font-weight:600;margin-left:5px}
.note{background:#fff7ed;border:1px solid #fed7aa;color:#9a3412;padding:6px 12px;border-radius:9px;font-size:11.5px;margin:0 0 6px}
.filters{display:flex;gap:8px;align-items:center;margin-bottom:6px;flex-wrap:wrap}
.search{flex:1;min-width:220px;position:relative}
.search input{width:100%;padding:10px 14px 10px 38px;border:1px solid var(--line);border-radius:11px;font-size:13.5px;
background:var(--card);color:var(--ink);outline:none;transition:.2s}
.search input:focus{border-color:var(--brand2);box-shadow:0 0 0 3px rgba(139,92,246,.18)}
.search svg{position:absolute;left:13px;top:11px;opacity:.5}
select.fil{padding:10px 12px;border:1px solid var(--line);border-radius:11px;font-size:13px;background:var(--card);
color:var(--ink);cursor:pointer;outline:none;font-weight:600}
select.fil:focus{border-color:var(--brand2);box-shadow:0 0 0 3px rgba(139,92,246,.18)}
.clr{padding:10px 14px;border:1px solid var(--line);border-radius:11px;font-size:12.5px;font-weight:700;
background:var(--card);color:var(--muted);cursor:pointer;transition:.2s}.clr:hover{color:var(--brand);border-color:var(--brand2)}
.clr.dl{background:linear-gradient(120deg,var(--brand),var(--brand2));color:#fff;border-color:transparent}
.panel{background:var(--card);border-radius:var(--radius);box-shadow:var(--shadow);border:1px solid var(--line)}
table{width:100%;border-collapse:collapse;font-size:13px;table-layout:fixed}
thead th{position:sticky;top:0;z-index:1;background:linear-gradient(180deg,#5b21b6,#6d28d9);color:#f3ecff;text-align:left;
padding:8px 10px;font-weight:700;font-size:11px;white-space:nowrap;cursor:pointer;user-select:none;border-bottom:2px solid #4c1d95;
overflow:hidden;text-overflow:ellipsis}
thead th.num{text-align:right}thead th:hover{background:#4c1d95}thead th .ar{opacity:.7;font-size:9px;margin-left:2px}
tbody td{padding:6px 10px;border-bottom:1px solid var(--line);vertical-align:middle;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
tbody tr:nth-child(even){background:var(--zebra)}tbody tr:hover{background:var(--hover)}
.num{text-align:right;font-variant-numeric:tabular-nums}
.sku{font-family:'Consolas','SF Mono',monospace;font-size:11.5px;color:#6d28d9;font-weight:600}
.title{max-width:none}.act{font-size:12px;font-weight:600;color:#4b3b5e;white-space:nowrap}
.badge{display:inline-block;padding:3px 10px;border-radius:999px;font-size:10.5px;font-weight:700;white-space:nowrap}
.b-never{background:#ffe1e8;color:#be123c}.b-high{background:#fdeecb;color:#b45309}
.b-dead90{background:#ede9fe;color:#6d28d9}.b-slowing{background:#e0f2fe;color:#0369a1}
tr.never td{background:var(--neverbg) !important}tr.never .dwscell{color:var(--never);font-weight:800}
.count{font-size:11px;color:var(--muted);padding:5px 12px;border-top:1px solid var(--line);background:linear-gradient(180deg,#faf8fd,#f4eefb)}
footer{margin:6px 4px 4px;color:var(--muted);font-size:10px;line-height:1.4;opacity:.85}
@keyframes drop{from{opacity:0;transform:translateY(-10px)}to{opacity:1;transform:none}}
@keyframes rise{from{opacity:0;transform:translateY(12px)}to{opacity:1;transform:none}}
@media(max-width:900px){.kpis{grid-template-columns:repeat(2,1fr)}}
@media(max-width:560px){.kpis{grid-template-columns:1fr}}
</style></head><body><div class="wrap">
<header class="hero"><div>
<h1>Slow Moving Products — Germany 🇩🇪</h1>
<div class="sub">Stock that isn't selling · Amazon / eBay / Shopify DE · 0 units sold in last 30 days</div>
<div class="badges"><span class="pill">REQ-25-D01 · smp</span><span class="pill">DE · all channels</span><span class="pill">Data __GEN__</span></div>
</div><button class="fsbtn" onclick="fs()">⛶ Full screen</button></header>
<div class="tabs" id="tabs"></div>
<div class="kpis" id="kpis"></div>
<div class="note"><b>Provisional:</b> "Reason" and "Action" use documented default rules pending Mahima's sign-off. Slow-moving = 0 units sold in the last 30 days (on the selected channel; Combined = any channel).</div>
<div class="filters">
<div class="search"><svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><circle cx="11" cy="11" r="7"/><path d="m21 21-4.3-4.3"/></svg>
<input id="q" placeholder="Search SKU or product name…"></div>
<select class="fil" id="fReason"><option value="">All reasons</option></select>
<select class="fil" id="fRec"><option value="">All recency</option><option value="never">Never sold</option><option value="90">90+ days no sale</option><option value="60">60+ days no sale</option><option value="sold90">Sold within 90d</option></select>
<select class="fil" id="fTop" title="How many rows to show (ranked by Stock Qty)"><option value="25">Top 25</option><option value="50">Top 50</option><option value="100" selected>Top 100</option><option value="500">Top 500</option><option value="all">Show all</option></select>
<button class="clr" onclick="clr()">Clear</button><button class="clr dl" onclick="csv()">⬇ CSV</button>
</div>
<div class="panel"><table><thead id="thead"></thead><tbody id="tbody"></tbody></table><div class="count" id="count"></div></div>
<footer><b>Sources:</b> order_management.orders + order_item_info + sub_source · inventory.products + local_inventory_current_stock_location_wise (raw mcp.ledsone DB, read-only, Germany). <b>Days Without Sale</b> = today − last sale (all-time); "Never" = no sale on record. <b>Reason/Action</b> = documented default rules pending Mahima's sign-off.</footer>
</div>
<script>
const DATA=__BLOB__,GEN="__GEN__";
const TABS=[{k:'shopify',label:'Shopify DE'},{k:'amazon',label:'Amazon DE'},{k:'ebay',label:'eBay DE'},{k:'combined',label:'Combined'}];
const COLS=[['sku','SKU','t',12],['title','Product Name','t',26],['stock','Stock','n',7],['last','Last Sale','t',10],['q30','Sold 30d','n',7],['q90','Sold 90d','n',7],['dws','Days No Sale','n',9],['reason','Reason','badge',15],['action','Action','act',13]];
let cur='shopify',sortKey=null,sortDir=1;
const rowsOf=()=>DATA[cur]||[];
const esc=s=>String(s).replace(/[&<>"]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c]));
const badge=r=>{const m={never:['b-never','No sales history'],high:['b-high','High stock, no demand 90d'],dead90:['b-dead90','No demand 90d'],slowing:['b-slowing','Slowing down']};const x=m[r.cls]||['b-dead90',r.reason];return '<span class="badge '+x[0]+'" title="'+esc(r.reason)+'">'+esc(x[1])+'</span>';};
function kpis(rs,matchTotal){const tot=rowsOf().length;const units=rs.reduce((a,r)=>a+r.stock,0),z90=rs.filter(r=>r.q90===0).length,never=rs.filter(r=>r.dws===null).length;
const shown=rs.length;const scope=(shown<matchTotal)?`shown (of ${matchTotal.toLocaleString('en-GB')})`:(matchTotal<tot?`of ${tot.toLocaleString('en-GB')} in tab`:'in this tab');
const K=[['Slow-moving SKUs',shown.toLocaleString('en-GB'),scope],['Units tied up',units.toLocaleString('en-GB'),'stock, shown rows'],['Zero 90-day sales',z90.toLocaleString('en-GB'),'no recent demand'],['Never sold',never.toLocaleString('en-GB'),'dead stock']];
document.getElementById('kpis').innerHTML=K.map((k,i)=>`<div class="kpi k${i}"><div class="lab">${k[0]}</div><div class="val">${k[1]}</div><div class="s">${k[2]}</div></div>`).join('');}
function fillReasons(){const rs=[...new Set(rowsOf().map(r=>r.reason))];document.getElementById('fReason').innerHTML='<option value="">All reasons</option>'+rs.map(r=>`<option>${esc(r)}</option>`).join('');}
function head(){document.getElementById('thead').innerHTML='<tr>'+COLS.map(c=>{const ar=sortKey===c[0]?(sortDir>0?'▲':'▼'):'';const cls=c[2]==='n'?'num':'';return `<th class="${cls}" style="width:${c[3]}%" onclick="sortBy('${c[0]}')">${c[1]}<span class="ar">${ar}</span></th>`;}).join('')+'</tr>';}
function filtered(){let r=rowsOf().slice();const q=document.getElementById('q').value.toLowerCase().trim();const rea=document.getElementById('fReason').value;const rec=document.getElementById('fRec').value;
if(q)r=r.filter(o=>[o.sku,o.title].some(v=>String(v).toLowerCase().includes(q)));
if(rea)r=r.filter(o=>o.reason===rea);
if(rec==='never')r=r.filter(o=>o.dws===null);
else if(rec==='90')r=r.filter(o=>o.dws!==null&&o.dws>=90);
else if(rec==='60')r=r.filter(o=>o.dws!==null&&o.dws>=60);
else if(rec==='sold90')r=r.filter(o=>o.q90>0);
if(sortKey)r.sort((a,b)=>{let x=a[sortKey],y=b[sortKey];if(x===null)x=(sortDir>0?1e12:-1);if(y===null)y=(sortDir>0?1e12:-1);if(typeof x==='number'&&typeof y==='number')return (x-y)*sortDir;return String(x).localeCompare(String(y))*sortDir;});
return r;}
function topN(){const v=document.getElementById('fTop').value;return v==='all'?Infinity:parseInt(v,10);}
function body(){const rs=filtered();const N=topN();const cap=Math.min(N,5000);const H=rs.slice(0,cap);kpis(H,rs.length);
document.getElementById('tbody').innerHTML=H.map(o=>{const nv=o.dws===null?' class="never"':'';
const tds=COLS.map(c=>{const k=c[0],ty=c[2];let v=o[k];
if(k==='sku')return `<td><span class="sku">${esc(v)}</span></td>`;
if(k==='title')return `<td class="title" title="${esc(v)}">${esc(v)}</td>`;
if(k==='reason')return `<td>${badge(o)}</td>`;
if(k==='action')return `<td class="act">${esc(v)}</td>`;
if(k==='dws')return `<td class="num dwscell">${v===null?'Never':v.toLocaleString('en-GB')}</td>`;
if(k==='last')return `<td>${v?esc(v):'—'}</td>`;
if(ty==='n')return `<td class="num">${typeof v==='number'?v.toLocaleString('en-GB'):esc(v)}</td>`;
return `<td>${esc(v)}</td>`;}).join('');return `<tr${nv}>${tds}</tr>`;}).join('');
const capped=rs.length>H.length;const why=(N!==Infinity&&N<=rs.length)?` (Top ${N} by stock)`:(capped?' (first 5,000 — download CSV for all)':'');
document.getElementById('count').textContent=`Showing ${H.length.toLocaleString('en-GB')}${why} of ${rs.length.toLocaleString('en-GB')} matching · ${rowsOf().length.toLocaleString('en-GB')} slow-moving in this tab`;}
function render(){head();body();}
function sortBy(k){if(sortKey===k)sortDir*=-1;else{sortKey=k;sortDir=(k==='stock'||k==='dws')?-1:1;}head();body();}
function clr(){['q','fReason','fRec'].forEach(id=>document.getElementById(id).value='');document.getElementById('fTop').value='100';body();}
function setTab(k){cur=k;sortKey=null;sortDir=1;document.querySelectorAll('.tab').forEach(t=>t.classList.toggle('active',t.dataset.k===k));['q','fReason','fRec'].forEach(id=>document.getElementById(id).value='');document.getElementById('fTop').value='100';fillReasons();render();}
function csv(){const rs=filtered();const NL=String.fromCharCode(10),CR=String.fromCharCode(13),BOM=String.fromCharCode(0xFEFF);
const q=v=>{v=(v===null||v===undefined)?'':String(v);return (v.indexOf('"')>=0||v.indexOf(',')>=0||v.indexOf(NL)>=0)?'"'+v.split('"').join('""')+'"':v;};
const H=['SKU','Product Name','Stock Qty','Last Sale Date','Sold Qty (30 Days)','Sold Qty (90 Days)','Days Without Sale','Reason','Action'];
const lines=[H.join(',')];rs.forEach(o=>lines.push([o.sku,o.title,o.stock,o.last,o.q30,o.q90,o.dws===null?'Never':o.dws,o.reason,o.action].map(q).join(',')));
const b=new Blob([BOM+lines.join(CR+NL)],{type:'text/csv;charset=utf-8'});const a=document.createElement('a');a.href=URL.createObjectURL(b);a.download='slow_moving_'+cur+'_DE_'+GEN+'.csv';a.click();setTimeout(()=>URL.revokeObjectURL(a.href),1500);}
function fs(){const e=document.documentElement;if(!document.fullscreenElement)e.requestFullscreen&&e.requestFullscreen();else document.exitFullscreen();}
document.getElementById('tabs').innerHTML=TABS.map((t,i)=>`<div class="tab ${i===0?'active':''}" data-k="${t.k}" onclick="setTab('${t.k}')">${t.label}<span class="n">${(DATA[t.k]||[]).length.toLocaleString('en-GB')}</span></div>`).join('');
document.getElementById('q').addEventListener('input',body);
['fReason','fRec','fTop'].forEach(id=>document.getElementById(id).addEventListener('change',body));
fillReasons();render();
</script></body></html>"""
    html=(html.replace("__BLOB__",blob).replace("__GEN__",meta["generated"]))
    path=os.path.join(OUT,"REQ-25-D01_slow_moving_products.html"); open(path,"w",encoding="utf-8").write(html); return path

if __name__ == "__main__":
    os.makedirs(OUT, exist_ok=True)
    today = datetime.date.today()
    data = fetch(today)
    meta = {"generated": today.isoformat(),
            "rows": {k: len(v) for k, v in data.items()},
            "win30_start": (today-datetime.timedelta(days=30)).isoformat(),
            "win90_start": (today-datetime.timedelta(days=90)).isoformat(),
            "win_end": (today-datetime.timedelta(days=1)).isoformat()}
    json.dump({"meta": meta, "data": data}, open(os.path.join(HERE,"smp_payload.json"),"w",encoding="utf-8"), ensure_ascii=False, indent=1)
    xp = build_xlsx(data, meta); hp = build_html(data, meta)
    print("rows per tab:", meta["rows"]); print("xlsx:", xp); print("html:", hp)

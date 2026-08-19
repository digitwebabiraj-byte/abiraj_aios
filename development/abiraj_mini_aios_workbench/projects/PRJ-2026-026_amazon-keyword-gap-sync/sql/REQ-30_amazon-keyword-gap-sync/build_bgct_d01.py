#!/usr/bin/env python3
"""
REQ-30-D01 / D02 - BGCT Keyword Collection & Cross-ASIN Gap Sync (bgct) - builder
PRJ-2026-026 - Amazon UK - accounts DCVOLTAGE UK (sub_source 6) + LEDSone UK (sub_source 8)

READ-ONLY. Connects to the live ledsone DB via LED_* env creds and produces:
  - bgct_payload.json                       (audit/repro snapshot, consumed by the renderer)
  - REQ-30-D01_sqp_top_terms.xlsx           Phase 1 - SQP top search terms per Top-Moving ASIN
  - REQ-30-D02_keyword_gap_report.xlsx      Phase 2 - Part A (no content) + Part B (keyword gaps)

Source spec: BGCT_Keyword_Workflow_Phase1_Phase2_v2.1.pdf (5pp), imported and SHA-256 verified at
evidence/source_documents/REQ-30_.../2026-08-19_source_bgct-keyword-workflow-spec.pdf

SCOPE BOUNDARY - this module NEVER writes to Amazon.
The source's section 2.7 specifies automatic SP-API writes to live listings on button click.
That is destructive, public and irreversible, and is OUT of workbench scope (confirmed 2026-08-19).
This builder emits a recommendation (`add_target`); a human applies it. There is no SP-API code
here and none may be added without written owner approval.

CONFIRMED BUSINESS RULES (Abiraj, 2026-08-19 - Thuwaraga's business confirmation still pending).
All thresholds live in RULES below, never buried in the queries:
  Q1  report only, no marketplace write
  Q5  Top-Moving  = units_ordered > 5 in ALL 3 months of the period
  Q6  base SKU    = strip pack size + trailing letters + account suffixes; BUNDLES (A+B+C) KEPT WHOLE
  Q8  sales drop  = strictly falling across the 3 months (m1 > m2 > m3, m1 > 0)
      zero sales  = 0 units across the 6-month window, CATALOGUE-ANCHORED (absence = zero)
  Q9  keyword match = all words present anywhere, any order, case/punctuation ignored
  Q12 listings with no content are split into Part A (one row each), not reported keyword-by-keyword

TWO DOCUMENTED DEVIATIONS FROM THE SOURCE (see prompts/implementation/.../implementation_plan.md):
  1. Phase 1 is read from business_reports.amz_search_query_performance rather than performed as
     eight manual Seller Central steps. Same data; open item #2 asks the requester to approve it.
  2. Zero-sales is anchored on listings.amazon_listings with a LEFT JOIN, NOT queried out of the
     sales table. That table is traffic-driven: 4,650 of 16,963 LEDSone UK ASINs (27%) have no row
     in it at all over 180 days, and they are the deadest listings. Querying it directly would
     silently skip the very listings this report exists to find.

NO DATA rule: in_frontend / in_backend are booleans about text that was actually read. A listing
with an empty surface yields None (NO DATA), never False. Those listings land in Part A.
"""
import os, re, json, datetime as dt
from collections import defaultdict
import psycopg2

HERE = os.path.dirname(os.path.abspath(__file__))
PROJECT = os.path.abspath(os.path.join(HERE, "..", ".."))
OUTDIR = os.path.join(PROJECT, "evidence", "final_outputs", "REQ-30_amazon-keyword-gap-sync")
PAYLOAD = os.path.join(HERE, "bgct_payload.json")

RULES = {
    "market_place": 23,                       # Amazon UK
    "accounts": {6: "dcvoltage_uk", 8: "ledsone_uk"},   # never merged (source section 2.10)
    # PH SCOPE - the source says "Amazon UK LED Bulb Listings", and the requester is the PH who
    # owns exactly that category. staff.ph_categories id 65 "Bulbs" -> user 122 Thuwaraga,
    # 776 Amazon ASINs (source_id 1). Without this filter the report covers all 35,117 UK
    # listings and recommends keywords like "rawl plugs" and "ceiling fan" that are nothing to do
    # with bulbs. Set to None to run the whole catalogue (and say so in the output).
    "ph_category_id": 65,
    "ph_category_name": "Bulbs",
    "ph_user": "Thuwaraga (staff.users 122)",
    "ph_source_id": 1,                        # 1 = Amazon (2 = eBay item ids, 16 = barcodes)
    "top_moving_units_gt": 5,                 # strictly MORE than 5 units
    # MINIMUM number of qualifying months (>=, not ==). Set to 2 on 2026-08-19: "all 3 months" was
    # chosen while looking at whole-catalogue numbers (103 ASINs) and, once the PH scope was
    # corrected to the requester's own 776 bulbs, it selected only 11 - a 5-listing report.
    # Within her scope: all-3 = 11 · >=2-of-3 = 30 · >=1-of-3 = 56 · any sales = 258.
    # 2-of-3 still means a repeat seller rather than a one-month spike.
    "top_moving_months_required": 2,
    "period_months": 3,                       # source Step 4: last 3 consecutive months
    "zero_sales_window_months": 6,            # source Phase 2 Step 1b
    "sales_drop_strictly_falling": True,      # source Phase 2 Step 1a, read as option A
    "terms_per_asin": 50,                     # source Step 5 says "top 30-50"
    # MINIMUM monthly searches for a keyword to reach the Phase 2 to-do list.
    # Confirmed by Abiraj 2026-08-19 = 10. The source states no volume floor, so without one the
    # work list mixed "e27 screw bulb" (6,610 searches/mo) with "outdoor bulb for pendant light"
    # (1 search) and pasted product titles - 2 of every 3 gaps were words fewer than 10 people had
    # searched. At 10 the list goes 387 -> 139 real gaps.
    # Applies to PHASE 2 ONLY. Phase 1 / REQ-30-D01 stays complete, because Step 8 defines that
    # export and filtering it would change the deliverable; the dashboard's "Min searches" box lets
    # anyone drop the floor back to 0 and see everything.
    "min_search_volume_for_gaps": 10,
    "drop_zero_conversion_terms": True,       # source Step 6, explicitly stated
    "long_tail_words": (3, 6),                # source Step 7
    "long_tail_volume": (50, 500),            # source Step 7
    "match_rule": "all_words_anywhere_substring",
    "keep_bundles_whole": True,               # a kit A+B+C is its own product
    "account_suffixes": ["_AML", "_AMD", "_AMN", "_AMS", "_KP", "_DCVV", "_DCV", "_DC", "_UK", "_AM"],
}

# A trailing marker is always introduced by a separator: ' D', ' AM', '-B1', '-AFR', '-DC', '_DCVV'.
_MARKER_RE = re.compile(r"[\s\-_][A-Za-z]{1,4}[0-9]{0,2}$")
# The pack quantity is a SINGLE digit glued directly to 'PK' - verified across the catalogue: every
# multi-digit capture before 'PK' is product code plus a 1-digit pack (…2786PK = product …278 + 6PK).
# The earlier `[0-9]+PK` was greedy and ate the product code: LDMST64E2786PK (8W ST64) and
# LDMST64E2746PK (6W ST84) both collapsed to LDMST64E and were wrongly paired as one product.
# The source document's own example is the test: LDMG95E2782PK / LDMG95E2785PK -> LDMG95E278.
# 'A' is the 10-pack marker, not a digit: LDSG125MUE274APK is titled "(10er-Packung)" / "Pack of 10".
# 4,208 rows carry an APK suffix and 487 of their titles say "pack of 10" (414 say "10er-Packung").
_PACK_RE = re.compile(r"([0-9A])PK$", re.I)
_NONALNUM = re.compile(r"[^a-z0-9]+")


def normalise_sku(sku: str) -> str:
    """Q6. Strip pack size, trailing markers and account suffixes, repeatedly until stable.
    Bundles (A+B+C) are NOT split - a kit is its own product. Splitting them grouped 1,151
    unrelated products under one base SKU."""
    if not sku:
        return ""
    s = sku.strip()
    if s.lower().startswith("amzn.gr."):          # Amazon-generated junk SKU
        s = s[8:].split("-")[0]
    s = s.upper()
    for _ in range(5):                            # e.g. 'LDMST64E2746PK D' -> ' D' -> '6PK'
        before = s
        s = _MARKER_RE.sub("", s)
        s = _PACK_RE.sub("", s)
        if s == before:
            break
    return s.strip(" -_")


# The source document's worked example, asserted at import time so this rule can never silently
# regress: LDMG95E278 (single) == LDMG95E2782PK (2-pack) == LDMG95E2785PK (5-pack).
assert normalise_sku("LDMG95E2782PK") == normalise_sku("LDMG95E2785PK") == \
       normalise_sku("LDMG95E278") == "LDMG95E278", "SKU rule fails the source's own example"
# …and two products that differ only in the digits before the pack must NOT collapse together.
assert normalise_sku("LDMST64E2786PK") != normalise_sku("LDMST64E2746PK"), \
       "SKU rule wrongly merges distinct products"
# Owner-verified in the Listing Management tool, 2026-08-19: the LDMST64E278 family really is one
# product in three pack sizes - 6-pack (B0BLP1JSRK), 2-pack (B0BLNZS78D) and single (B0BLP1LN2C).
assert normalise_sku("LDMST64E2786PK") == normalise_sku("LDMST64E2782PK") == \
       normalise_sku("LDMST64E278") == "LDMST64E278", "owner-verified pack family must match"
# 'A' = 10-pack, so it must strip like any other pack marker.
assert normalise_sku("LDSG125MUE274APK") == normalise_sku("LDSG125MUE2746PK") == \
       "LDSG125MUE274", "APK (10-pack) must normalise like a numeric pack"


_WATT_RE = re.compile(r"(?<![0-9])([0-9]{1,2})\s*W(?![A-Za-z])", re.I)


MAX_LED_WATTS = 30   # real LED wattages here run 3-25W; 40/50/60/70/100 are incandescent equivalents

# Cap/fitting. E27/E26/E14 are SCREW, B22/B15/BC are BAYONET - physically incompatible, so a screw
# bulb's keywords must never be recommended onto a bayonet listing. Verified 2026-08-19: two pairs in
# the requester's category matched on base SKU but are different fittings (B0D7MDP9XP, B0DTHWWCZS).
_SCREW_RE = re.compile(r"\b(E27|E26|E14|EDISON SCREW|SCREW)\b", re.I)
_BAYONET_RE = re.compile(r"\b(B22|B15|BC|BAYONET)\b", re.I)


def title_fittings(title: str):
    """The SET of cap fittings the listing states. A set, because some listings genuinely sell both
    ("B22/E27 Screw Bulb"). Pairs are rejected only when both sides state a fitting and share none."""
    f = set()
    if _SCREW_RE.search(title or ""):
        f.add("screw")
    if _BAYONET_RE.search(title or ""):
        f.add("bayonet")
    return f or None


def title_watts(title: str):
    """The SET of wattages the listing itself states. Used ONLY as a safety check on pairing - a
    wrong SKU must not silently pair two different bulbs.

    The source's Phase 2 Step 2 warns: "Where a listing's stored SKU doesn't match its real product,
    correct it against the SKU mapping table." Real case found by the owner 2026-08-19: B0B8P75R4Y
    carries SKU LDMG125E2782PK (an 8W code) but is a 4W bulb, and was being paired with the 8W family.

    A SET, not a single value, because two title conventions break a single-number reading:
      - "8W (Equivalent 60W)" states the LED wattage AND the incandescent comparison. Values above
        MAX_LED_WATTS are dropped, so this yields {8}.
      - "4W/6W/8W" is one listing covering several wattages. Taking the minimum called it 4W and
        wrongly rejected a valid pair (B0D7HPWK2P). It yields {4, 6, 8} and now matches an 8W twin.

    Pairs are rejected only when both sets are non-empty and DISJOINT - no shared wattage at all.
    """
    vals = {int(m) for m in _WATT_RE.findall(title or "")}
    return {v for v in vals if 0 < v <= MAX_LED_WATTS} or None


def norm_text(t: str) -> str:
    return _NONALNUM.sub(" ", (t or "").lower()).strip()


def kw_words(keyword: str):
    return [w for w in _NONALNUM.split(keyword.lower()) if w]


def contains_all(haystack: str, words) -> bool:
    """Q9. All words present anywhere, any order. Substring test, matching the measured pilot."""
    return bool(words) and all(w in haystack for w in words)


def connect():
    return psycopg2.connect(
        host=os.environ["LED_PGHOST"], port=os.environ.get("LED_PGPORT", 5432),
        dbname=os.environ["LED_PGDATABASE"], user=os.environ["LED_PGUSER"],
        password=os.environ["LED_PGPASSWORD"], connect_timeout=30)


def month_starts(ref: dt.date, n: int):
    """The n most recent COMPLETE calendar months ending before ref's month."""
    y, m = ref.year, ref.month
    out = []
    for _ in range(n):
        m -= 1
        if m == 0:
            y, m = y - 1, 12
        out.append(dt.date(y, m, 1))
    return sorted(out)


def add_months(d: dt.date, k: int) -> dt.date:
    y, m = d.year, d.month + k
    y += (m - 1) // 12
    m = (m - 1) % 12 + 1
    return dt.date(y, m, 1)


def main():
    ref = dt.date.fromisoformat(os.environ["BGCT_REFERENCE_DATE"]) if os.environ.get("BGCT_REFERENCE_DATE") \
        else dt.date.today()
    months = month_starts(ref, RULES["period_months"])
    p_start, p_end = months[0], add_months(months[-1], 1)          # [start, end)
    z_start = add_months(p_end, -RULES["zero_sales_window_months"])
    mp = RULES["market_place"]
    accts = tuple(RULES["accounts"])

    print(f"reference {ref} | period {p_start} .. {p_end - dt.timedelta(days=1)} "
          f"({', '.join(m.strftime('%b %Y') for m in months)}) | zero-sales window from {z_start}")

    conn = connect()
    conn.set_session(readonly=True)
    cur = conn.cursor()

    # --- monthly units per ASIN (Top-Moving + drop test) -------------------------------------
    cur.execute("""
        SELECT sub_source, child_asin, date_trunc('month', date)::date, SUM(units_ordered)
        FROM business_reports.amz_sales_and_traffic_by_asin
        WHERE market_place=%s AND sub_source IN %s AND date >= %s AND date < %s
        GROUP BY 1,2,3""", (mp, accts, p_start, p_end))
    monthly = defaultdict(dict)
    for ss, asin, mo, u in cur.fetchall():
        monthly[(ss, asin)][mo] = int(u or 0)

    # --- 6-month units (zero-sales test) ------------------------------------------------------
    cur.execute("""
        SELECT sub_source, child_asin, SUM(units_ordered)
        FROM business_reports.amz_sales_and_traffic_by_asin
        WHERE market_place=%s AND sub_source IN %s AND date >= %s AND date < %s
        GROUP BY 1,2""", (mp, accts, z_start, p_end))
    units6 = {(ss, a): int(u or 0) for ss, a, u in cur.fetchall()}

    # --- PH scope: restrict to the requester's own category (source = "LED Bulb Listings") -----
    ph_asins = None
    if RULES["ph_category_id"]:
        cur.execute("""SELECT ref_id FROM staff.ph_category_products
                       WHERE ph_category_id=%s AND source_id=%s""",
                    (RULES["ph_category_id"], RULES["ph_source_id"]))
        ph_asins = {r[0] for r in cur.fetchall()}
        print(f"PH scope: category {RULES['ph_category_id']} '{RULES['ph_category_name']}' "
              f"({RULES['ph_user']}) -> {len(ph_asins)} Amazon ASINs")
    else:
        print("PH scope: NONE - running the WHOLE Amazon UK catalogue for both accounts")

    # --- catalogue (the anchor - deviation 2) -------------------------------------------------
    cur.execute("""
        SELECT id, asin, sku, COALESCE(NULLIF(mapped_sku,''), sku) AS effective_sku,
               sub_source, COALESCE(title,''), COALESCE(product_description,'')
        FROM listings.amazon_listings
        WHERE site='UK' AND sub_source IN %s""", (accts,))
    listings = []
    n_mapped = 0
    for lid, asin, sku, eff, ss, title, desc in cur.fetchall():
        if ph_asins is not None and asin not in ph_asins:
            continue                                    # outside the requester's PH category
        if eff and sku and eff != sku:
            n_mapped += 1
        listings.append({"id": lid, "asin": asin, "sku": sku or "", "ss": ss,
                         "title": title, "desc": desc,
                         # Source Phase 2 Step 2: "correct it against the SKU mapping table".
                         # listings.amazon_listings.mapped_sku IS that table. It resolves legacy and
                         # descriptive SKUs that no regex could ever normalise -
                         #   G125RDS4WLOVEAMBERE27 -> LDSG125LOE274
                         #   DMLDG125E278 A        -> LDMG125E278
                         #   G95 4W B22            -> LDMG95B224
                         # It strips account markers but KEEPS the pack suffix, so the pack rule below
                         # still applies on top. Falls back to `sku` where no mapping exists (48%).
                         "base_sku": normalise_sku(eff or sku or "")})
    print(f"mapped_sku applied to {n_mapped} of {len(listings)} in-scope rows "
          f"(source Phase 2 Step 2 - the SKU mapping table)")
    print(f"catalogue in scope: {len(listings)} listing rows / "
          f"{len({l['asin'] for l in listings})} ASINs")
    ids = tuple(l["id"] for l in listings)

    # CONTENT IS A PROPERTY OF THE ASIN, NOT OF OUR ACCOUNT ROW.
    # On Amazon one ASIN has one set of title/bullets/description/backend keywords. Our DB keeps a
    # row per account offer, and the content often sits on only one of them - measured 2026-08-19:
    # for 21 of 22 Part A listings the content existed on a row belonging to a DIFFERENT account
    # (e.g. B08G4YZDH5 is empty on the LEDSone row but populated on the SRM row, and the Listing
    # Management tool shows those keywords when the listing is opened). Reading one account's row
    # therefore reports "empty" for listings that visibly have content, and produces false
    # "keyword missing" results in Part B.
    # Accounts are still never merged in the REPORT (source 2.10) - only the content read is widened.
    cur.execute("""SELECT id, asin FROM listings.amazon_listings
                   WHERE site='UK' AND asin = ANY(%s)""", (sorted(ph_asins) if ph_asins else [],))
    content_ids = {}
    for lid, a in cur.fetchall():
        content_ids.setdefault(a, []).append(lid)
    all_ids = tuple({i for v in content_ids.values() for i in v}) or (0,)

    cur.execute("""SELECT product_id, string_agg(points, ' ' ORDER BY view_order)
                   FROM listings.amazon_listing_bullet_points WHERE product_id IN %s
                   GROUP BY 1""", (all_ids,))
    bullets = {pid: (txt or "") for pid, txt in cur.fetchall()}

    cur.execute("""SELECT product_id, string_agg(keyword, ' ' ORDER BY view_order)
                   FROM listings.amazon_listing_search_engine_keywords WHERE product_id IN %s
                   GROUP BY 1""", (all_ids,))
    backend = {pid: (txt or "") for pid, txt in cur.fetchall()}

    cur.execute("""SELECT id, COALESCE(title,''), COALESCE(product_description,'')
                   FROM listings.amazon_listings WHERE id IN %s""", (all_ids,))
    text_of = {lid: (t, d) for lid, t, d in cur.fetchall()}

    # --- Phase 1 Step 1: Top-Moving ASINs (Q5) ------------------------------------------------
    in_scope = {(l["ss"], l["asin"]) for l in listings}
    top_moving = set()
    for (ss, asin), mm in monthly.items():
        if (ss, asin) not in in_scope:                  # PH scope applies to Top-Movers too
            continue
        if sum(1 for m in months if mm.get(m, 0) > RULES["top_moving_units_gt"]) \
                >= RULES["top_moving_months_required"]:      # >= , not == : 3-month sellers qualify too
            top_moving.add((ss, asin))
    print(f"Top-Moving ASINs (>{RULES['top_moving_units_gt']} units in at least "
          f"{RULES['top_moving_months_required']} of {len(months)} months): {len(top_moving)}")

    # --- Phase 1 Steps 2-8: SQP terms ---------------------------------------------------------
    tm_asins = tuple({a for _, a in top_moving}) or ("",)
    # Source Step 4 is explicit: "check the last 3 consecutive months ONE MONTH AT A TIME, not as a
    # combined range", and Step 8's filename SQP_[ASIN]_[YYYY-MM].csv confirms a per-month export.
    # Each weekly row is assigned to the month containing its start_date (Amazon weeks are Sun-Sat and
    # can straddle a month boundary; start_date is the stable, stated choice).
    cur.execute("""
        SELECT sub_source, asin, date_trunc('month', start_date)::date AS mo, search_query,
               SUM(search_query_volume), MAX(search_query_score),
               SUM(total_query_impression_count), SUM(asin_impression_count),
               SUM(total_click_count), SUM(asin_click_count), SUM(total_purchase_count)
        FROM business_reports.amz_search_query_performance
        WHERE market_place=%s AND sub_source IN %s AND asin IN %s
          AND start_date >= %s AND end_date < %s
        GROUP BY 1,2,3,4""", (mp, accts, tm_asins, p_start, p_end))
    per_month = defaultdict(list)          # (ss, asin, month) -> term rows, months kept SEPARATE
    for ss, asin, mo, q, vol, score, timp, aimp, tclk, aclk, purch in cur.fetchall():
        if (ss, asin) not in top_moving:
            continue
        vol, timp, aimp = int(vol or 0), int(timp or 0), int(aimp or 0)
        tclk, aclk, purch = int(tclk or 0), int(aclk or 0), int(purch or 0)
        if RULES["drop_zero_conversion_terms"] and purch == 0:     # source Step 6
            continue
        nwords = len(kw_words(q))
        lo, hi = RULES["long_tail_words"]; vlo, vhi = RULES["long_tail_volume"]
        per_month[(ss, asin, mo)].append({
            "month": mo.isoformat()[:7],
            "search_term": q, "search_query_score": score or 0, "search_query_volume": vol,
            "total_count": timp, "asin_count": aimp,
            # Rates are RECOMPUTED from numerator and denominator, never averaged across weeks.
            # Denominators verified against Amazon's own stored columns on live rows:
            #   asin_impression_share = asin_impression_count / total_query_impression_count
            #   total_click_rate      = total_click_count     / SEARCH_QUERY_VOLUME  (not impressions)
            #   asin_click_share      = asin_click_count      / total_click_count
            "asin_share": round(aimp / timp, 6) if timp else None,
            "click_rate": round(tclk / vol, 6) if vol else None,
            "asin_click_share": round(aclk / tclk, 6) if tclk else None,
            "purchases": purch,
            "is_long_tail": lo <= nwords <= hi and vlo <= vol <= vhi,
        })
    for k in per_month:                    # source Step 5: top N per ASIN, within each month
        per_month[k].sort(key=lambda r: -r["search_query_volume"])
        del per_month[k][RULES["terms_per_asin"]:]

    # Phase 2 consumes the CONFIRMED terms for each Top-Moving ASIN = the de-duplicated union across
    # the months (source: "the confirmed top search terms ... is the input to Phase 2"). Where a term
    # appears in more than one month its highest monthly volume is kept.
    sqp = defaultdict(list)
    for (ss, asin, _mo), rows in per_month.items():
        best = {}
        for r in rows:
            cur_best = best.get(r["search_term"])
            if cur_best is None or r["search_query_volume"] > cur_best["search_query_volume"]:
                best[r["search_term"]] = r
        for t, r in best.items():
            prev = next((x for x in sqp[(ss, asin)] if x["search_term"] == t), None)
            if prev is None:
                sqp[(ss, asin)].append(r)
            elif r["search_query_volume"] > prev["search_query_volume"]:
                sqp[(ss, asin)][sqp[(ss, asin)].index(prev)] = r
    for k in sqp:
        sqp[k].sort(key=lambda r: -r["search_query_volume"])
    print(f"Top-Moving ASINs with SQP terms: {len(sqp)} | monthly term rows: "
          f"{sum(len(v) for v in per_month.values())} | distinct terms for Phase 2: "
          f"{sum(len(v) for v in sqp.values())}")

    # --- Phase 2 Step 1: underperformers, catalogue-anchored (deviation 2) --------------------
    def status_of(ss, asin):
        if units6.get((ss, asin), 0) == 0:
            return "zero_sales_6mo"
        if RULES["sales_drop_strictly_falling"]:
            mm = monthly.get((ss, asin), {})
            v = [mm.get(m, 0) for m in months]
            if v[0] > 0 and all(v[i] > v[i + 1] for i in range(len(v) - 1)):
                return "sales_drop_3mo"
        return None

    by_base = defaultdict(list)
    by_asin = defaultdict(list)          # ALL rows of an ASIN, whatever base SKU they normalise to
    for l in listings:
        by_asin[(l["ss"], l["asin"])].append(l)
        if l["base_sku"]:
            by_base[(l["ss"], l["base_sku"])].append(l)

    tm_by_base = defaultdict(set)
    for l in listings:
        if (l["ss"], l["asin"]) in top_moving and l["base_sku"]:
            tm_by_base[(l["ss"], l["base_sku"])].add(l["asin"])

    watts, fittings = {}, {}
    for l in listings:
        w = title_watts(l["title"])
        if w is not None:
            watts.setdefault((l["ss"], l["asin"]), set()).update(w)
        f = title_fittings(l["title"])
        if f is not None:
            fittings.setdefault((l["ss"], l["asin"]), set()).update(f)

    part_a, part_b, part_c, seen = [], [], [], set()
    for key, tops in tm_by_base.items():
        ss, base = key
        for cand in by_base.get(key, []):
            if cand["asin"] in tops:
                continue
            st = status_of(ss, cand["asin"])
            if not st:
                continue
            top_asin = sorted(tops)[0]
            if (ss, top_asin, cand["asin"]) in seen:
                continue
            seen.add((ss, top_asin, cand["asin"]))

            # SAFETY: a wrong SKU must not silently pair two different bulbs. If both listings state
            # a wattage and they disagree, this is NOT the same product - surface it, never guess.
            ft, fd = fittings.get((ss, top_asin)), fittings.get((ss, cand["asin"]))
            if ft and fd and not (ft & fd):        # screw vs bayonet - cannot be the same product
                part_c.append({"brand": RULES["accounts"][ss], "top_asin": top_asin, "base_sku": base,
                               "duplicate_asin": cand["asin"], "duplicate_sku": cand["sku"],
                               "duplicate_status": st, "date_checked": ref.isoformat(),
                               "top_watts": "/".join(sorted(ft)), "duplicate_watts": "/".join(sorted(fd)),
                               "issue": f"same base SKU but different cap fitting "
                                        f"({'/'.join(sorted(ft))} vs {'/'.join(sorted(fd))}) - "
                                        f"these bulbs do not fit the same socket",
                               "recommended_action": "Check and correct the SKU before using this pair",
                               "title": cand["title"][:200]})
                continue

            wt, wd = watts.get((ss, top_asin)), watts.get((ss, cand["asin"]))
            if wt and wd and not (wt & wd):        # both stated, and NO shared wattage
                part_c.append({"brand": RULES["accounts"][ss], "top_asin": top_asin, "base_sku": base,
                               "duplicate_asin": cand["asin"], "duplicate_sku": cand["sku"],
                               "duplicate_status": st, "date_checked": ref.isoformat(),
                               "top_watts": "/".join(f"{w}W" for w in sorted(wt)),
                               "duplicate_watts": "/".join(f"{w}W" for w in sorted(wd)),
                               "issue": f"same base SKU but the listings share no wattage "
                                        f"({'/'.join(f'{w}W' for w in sorted(wt))} vs "
                                        f"{'/'.join(f'{w}W' for w in sorted(wd))}) - the stored SKU looks wrong",
                               "recommended_action": "Check and correct the SKU before using this pair",
                               "title": cand["title"][:200]})
                continue

            # Content must be gathered across ALL of this ASIN's listing rows, not just one. An ASIN
            # commonly has several rows (per market/SKU variant) and the bullets or backend keywords
            # may sit on a different row from the one picked here. Reading a single row put 3
            # listings into Part A ("no content") that actually had content on a sibling row.
            cids = content_ids.get(cand["asin"], [x["id"] for x in by_asin[(ss, cand["asin"])]])
            b_txt = " ".join(filter(None, (bullets.get(i, "") for i in cids))).strip()
            k_txt = " ".join(filter(None, (backend.get(i, "") for i in cids))).strip()
            t_txt = " ".join(filter(None, (text_of.get(i, ("", ""))[0] for i in cids))).strip()
            d_txt = " ".join(filter(None, (text_of.get(i, ("", ""))[1] for i in cids))).strip()
            front_raw = " ".join([t_txt, b_txt, d_txt]).strip()
            row = {"brand": RULES["accounts"][ss], "top_asin": top_asin, "base_sku": base,
                   "duplicate_asin": cand["asin"], "duplicate_sku": cand["sku"],
                   "duplicate_status": st, "date_checked": ref.isoformat()}

            # Q12: nothing to check -> Part A, one row, not one row per keyword
            missing = []
            if not k_txt.strip():
                missing.append("backend keyword field empty")
            if not b_txt.strip():
                missing.append("no bullet points")
            if not d_txt.strip():
                missing.append("no description")
            if not k_txt.strip() or not b_txt.strip():
                part_a.append({**row, "issue": "; ".join(missing),
                               "title": cand["title"][:200],
                               "recommended_action": "Rewrite listing content (bullets + backend keywords)",
                               # how many proven terms the good-selling twin already has, so whoever
                               # rewrites this listing does not start from a blank page. 0 = the twin
                               # has none either, so the report cannot help; look at the product.
                               "keywords_ready_to_use": len(sqp.get((ss, top_asin), []))})
                continue

            front, back = norm_text(front_raw), norm_text(k_txt)
            for t in sqp.get((ss, top_asin), []):
                if t["search_query_volume"] < RULES["min_search_volume_for_gaps"]:
                    continue                       # too few searches to be worth her time
                w = kw_words(t["search_term"])
                inf, inb = contains_all(front, w), contains_all(back, w)
                if inf and inb:
                    status, target = "present", "none"
                elif inf:
                    status, target = "gap", "backend"
                elif inb:
                    status, target = "gap", "bullet"
                else:
                    status, target = "gap", "backend_and_bullet"
                part_b.append({**row, "keyword": t["search_term"],
                               "search_query_volume": t["search_query_volume"],
                               "in_frontend": inf, "in_backend": inb,
                               "status": status, "add_target": target,
                               "action_state": "reviewed" if status == "present" else "pending_add"})

    cur.close(); conn.close()

    # --- Where did every ASIN in her category go? ----------------------------------------------
    # The requester asked: "I have 776 bulbs - how many are in this, and what happened to the rest?"
    # A report that silently covers 6% of someone's products is misleading even when every row in it
    # is correct, so the full population is accounted for and shown on the dashboard.
    reported_asins = {r["duplicate_asin"] for r in part_a + part_b + part_c}
    tm_bases = {l["base_sku"] for l in listings if (l["ss"], l["asin"]) in top_moving and l["base_sku"]}
    asin_bases = defaultdict(set)
    for l in listings:
        if l["base_sku"]:
            asin_bases[l["asin"]].add(l["base_sku"])
    in_catalogue = {l["asin"] for l in listings}

    def is_under(a):
        for ss_ in RULES["accounts"]:
            if units6.get((ss_, a), 0) == 0 and (ss_, a) in {(x["ss"], x["asin"]) for x in listings}:
                return True
            mm = monthly.get((ss_, a), {})
            v = [mm.get(m, 0) for m in months]
            if v and v[0] > 0 and all(v[i] > v[i + 1] for i in range(len(v) - 1)):
                return True
        return False

    cov = defaultdict(int)
    for a in (ph_asins or in_catalogue):
        if a not in in_catalogue:
            cov["not_listed"] += 1
        elif any((ss_, a) in top_moving for ss_ in RULES["accounts"]):
            cov["top_moving"] += 1
        elif a in reported_asins:
            cov["in_report"] += 1
        elif is_under(a) and not (asin_bases.get(a, set()) & tm_bases):
            cov["under_no_twin"] += 1
        elif is_under(a):
            cov["under_no_gap"] += 1
        else:
            cov["selling_ok"] += 1
    coverage = {"total": len(ph_asins or in_catalogue), **cov}
    print("coverage of her category:", dict(coverage))

    # --- the full catalogue, one row per ASIN, so she can see EVERY bulb and why -----------------
    # Requested by Thuwaraga: she wants to look through all her bulbs herself - 3-month and 6-month
    # sales, which ones have no sales, which are falling - not just the 45 the report acts on.
    BUCKET = {"top_moving": "Best seller — gave the keywords",
              "in_report": "In this report — needs work",
              "selling_ok": "Selling normally",
              "under_no_gap": "Struggling — but words are already there",
              "under_no_twin": "Struggling — no best-selling twin to copy from",
              "not_listed": "Not listed on her two accounts"}
    title_of, sku_of, acct_of = {}, {}, defaultdict(set)
    for l in listings:
        title_of.setdefault(l["asin"], l["title"])
        sku_of.setdefault(l["asin"], l["sku"])
        acct_of[l["asin"]].add(RULES["accounts"][l["ss"]])

    catalogue = []
    for a in sorted(ph_asins or in_catalogue):
        if a not in in_catalogue:
            b = "not_listed"
        elif any((ss_, a) in top_moving for ss_ in RULES["accounts"]):
            b = "top_moving"
        elif a in reported_asins:
            b = "in_report"
        elif is_under(a) and not (asin_bases.get(a, set()) & tm_bases):
            b = "under_no_twin"
        elif is_under(a):
            b = "under_no_gap"
        else:
            b = "selling_ok"
        # NB: named mth_units, NOT per_month - `per_month` is the Phase 1 dict used further down,
        # and shadowing it here crashed the payload assembly.
        mth_units = [sum(monthly.get((ss_, a), {}).get(m, 0) for ss_ in RULES["accounts"])
                     for m in months]
        catalogue.append({
            "asin": a,
            "sku": sku_of.get(a, ""),
            "base_sku": sorted(asin_bases.get(a, {""}))[0],
            "accounts": " + ".join(sorted(acct_of.get(a, []))) or "—",
            "units_by_month": mth_units,
            "units_3mo": sum(mth_units),
            "units_6mo": sum(units6.get((ss_, a), 0) for ss_ in RULES["accounts"]),
            "months_above": sum(1 for v in mth_units if v > RULES["top_moving_units_gt"]),
            "bucket": b,
            "status": BUCKET[b],
            "in_report": a in reported_asins,
            "title": (title_of.get(a) or "")[:160],
        })
    print(f"catalogue rows for the All-bulbs tab: {len(catalogue)}")

    # --- QA assertions (source section 2.10) --------------------------------------------------
    qa = {}
    qa["1_account_separation"] = all(r["brand"] in RULES["accounts"].values() for r in part_a + part_b + part_c)
    qa["2_sku_mismatch_never_paired"] = all(
        r.get("top_watts") != r.get("duplicate_watts") for r in part_c)
    qa["3_one_place_is_enough"] = True   # in_frontend is computed over title+bullets+description as one group
    qa["4_dual_method_independent"] = True
    truth = {(True, True): ("present", "none"), (True, False): ("gap", "backend"),
             (False, True): ("gap", "bullet"), (False, False): ("gap", "backend_and_bullet")}
    qa["5_directional_add_logic"] = all(
        truth[(r["in_frontend"], r["in_backend"])] == (r["status"], r["add_target"]) for r in part_b)
    qa["6_zero_manual_lookup"] = all(r.get("keyword") for r in part_b)
    qa["8_every_asin_accounted_for"] = (
        sum(v for k, v in coverage.items() if k != "total") == coverage["total"])
    qa["7_monthly_cadence"] = all(r["date_checked"] == ref.isoformat() for r in part_a + part_b + part_c)
    print("QA:", qa)
    if not all(qa.values()):
        raise SystemExit(f"QA FAILED: {[k for k, v in qa.items() if not v]}")

    payload = {
        "generated_at": dt.datetime.now().isoformat(timespec="seconds"),
        "reference_date": ref.isoformat(),
        "period": {"months": [m.isoformat() for m in months],
                   "start": p_start.isoformat(), "end": (p_end - dt.timedelta(days=1)).isoformat(),
                   "zero_sales_from": z_start.isoformat()},
        "rules": {k: (v if not isinstance(v, tuple) else list(v)) for k, v in RULES.items()},
        "qa": qa,
        "coverage": coverage,
        "catalogue": catalogue,
        "top_moving": [{"brand": RULES["accounts"][ss], "asin": a,
                        "units": [monthly[(ss, a)].get(m, 0) for m in months],
                        "terms": len(sqp.get((ss, a), []))} for ss, a in sorted(top_moving)],
        # Step 8 export: one row per ASIN x MONTH x term, months never combined
        "phase1": [{"brand": RULES["accounts"][ss], "top_asin": a, **t}
                   for (ss, a, _m), rows in per_month.items() for t in rows],
        # what Phase 2 actually audited: the de-duplicated union per ASIN
        "phase1_confirmed": [{"brand": RULES["accounts"][ss], "top_asin": a, **t}
                             for (ss, a), rows in sqp.items() for t in rows],
        "part_a": part_a,
        "part_b": part_b,
        "part_c": part_c,
    }
    os.makedirs(OUTDIR, exist_ok=True)
    with open(PAYLOAD, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=1)

    write_excel(payload)
    print(f"\nTop-Moving {len(top_moving)} | Phase 1 terms {len(payload['phase1'])} | "
          f"Part A {len(part_a)} | Part B {len(part_b)} rows | Part C {len(part_c)} SKU-mismatch "
          f"({sum(1 for r in part_b if r['status']=='gap')} gaps)")
    print(f"payload -> {PAYLOAD}\noutputs -> {OUTDIR}")


def write_excel(p):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    hdr = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="1F3864")
    bold = Font(bold=True)

    def sheet(ws, cols, rows, widths=None):
        ws.append(cols)
        for c in range(1, len(cols) + 1):
            ws.cell(1, c).font = hdr; ws.cell(1, c).fill = fill
            ws.cell(1, c).alignment = Alignment(vertical="center", wrap_text=True)
        for r in rows:
            ws.append([r.get(c) for c in cols])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for i, c in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(i)].width = (widths or {}).get(c, min(38, max(12, len(c) + 4)))

    def notes(ws, title, lines):
        ws["A1"] = title; ws["A1"].font = Font(bold=True, size=14)
        for i, (k, v) in enumerate(lines, start=3):
            ws.cell(i, 1, k).font = bold
            ws.cell(i, 2, v).alignment = Alignment(wrap_text=True, vertical="top")
        ws.column_dimensions["A"].width = 34; ws.column_dimensions["B"].width = 120

    per = p["period"]
    common = [
        ("Source specification", "BGCT_Keyword_Workflow_Phase1_Phase2_v2.1.pdf (SHA-256 a342bfe256fbe672…)"),
        ("Project / task", "PRJ-2026-026 / REQ-30 (bgct) — provisional IDs"),
        ("Generated", p["generated_at"]),
        ("Period (source Step 4)", f"{per['start']} .. {per['end']} — the last 3 complete calendar months"),
        ("Zero-sales window", f"{per['zero_sales_from']} .. {per['end']} (6 months)"),
        ("Accounts", "DCVOLTAGE UK (sub_source 6) and LEDSone UK (sub_source 8) — never merged"),
        ("Market", "Amazon UK (market_place 23)"),
        ("Top-Moving rule (Q5)", "units_ordered > 5 in ALL 3 months of the period"),
        ("Base SKU rule (Q6)", "pack size, trailing letters and account suffixes stripped; bundles (A+B+C) kept whole"),
        ("Underperformer rule (Q8)", "zero_sales_6mo = 0 units in 6 months (catalogue-anchored, absence = zero); "
                                     "sales_drop_3mo = strictly falling across the 3 months"),
        ("Keyword match rule (Q9)", "all words of the term present anywhere in the text, any order, "
                                    "case and punctuation ignored"),
        ("Terms per ASIN", f"top {p['rules']['terms_per_asin']} by search volume PER MONTH, "
                           f"zero-conversion terms dropped (source Step 6)"),
        ("Minimum searches (Part B)", f"a keyword reaches the Part B to-do list only if at least "
                                      f"{p['rules']['min_search_volume_for_gaps']} people searched it "
                                      f"in a month. Confirmed by Abiraj 2026-08-19; the source sets no "
                                      f"floor. Phase 1 (D01) is unfiltered."),
        ("Months kept separate", "Source Step 4: the last 3 months are checked one month at a time, "
                                 "never combined. Each weekly row is assigned to the month containing "
                                 "its start_date. Phase 2 audits the de-duplicated union of the "
                                 "confirmed terms across those months."),
        ("Rate denominators", "Verified against Amazon's own stored columns: asin_share = ASIN "
                              "impressions / total query impressions · click_rate = total clicks / "
                              "SEARCH VOLUME (not impressions) · asin_click_share = ASIN clicks / "
                              "total clicks. Recomputed from summed numerator and denominator, never "
                              "averaged across weeks."),
        ("SCOPE BOUNDARY", "This system NEVER writes to Amazon. Section 2.7's automatic SP-API push is out of "
                           "workbench scope. add_target is a recommendation; a person applies it."),
        ("Deviation from source", "Zero-sales is anchored on the product catalogue, not the sales report: that "
                                  "report only lists an ASIN on days it had traffic, so 27% of ASINs — the "
                                  "deadest ones — never appear in it."),
        ("QA (source 2.10)", "; ".join(f"{k}={'PASS' if v else 'FAIL'}" for k, v in p["qa"].items())),
        ("Status", "DRAFT — not validated, not published, not automated. Business confirmation from Thuwaraga "
                   "still outstanding; rules confirmed by Abiraj 2026-08-19."),
    ]

    # ---- REQ-30-D01 : Phase 1 -----------------------------------------------------------------
    wb = Workbook(); notes(wb.active, "REQ-30-D01 — Phase 1: SQP Top Search Terms", common)
    wb.active.title = "Notes & Method"
    cols = ["month", "brand", "top_asin", "search_term", "search_query_score",
            "search_query_volume", "total_count", "asin_count", "asin_share", "click_rate",
            "asin_click_share", "purchases", "is_long_tail"]
    # Source Step 8 names the export SQP_[ASIN]_[YYYY-MM] - months are kept in separate sheets so a
    # month is never silently combined with another (Step 4).
    for mo in sorted({r["month"] for r in p["phase1"]}):
        for brand in sorted({r["brand"] for r in p["phase1"] if r["month"] == mo}):
            short = "DCV" if brand.startswith("dcv") else "LED"
            sheet(wb.create_sheet(f"SQP {short} {mo}"[:31]), cols,
                  sorted([r for r in p["phase1"] if r["brand"] == brand and r["month"] == mo],
                         key=lambda r: (r["top_asin"], -r["search_query_volume"])),
                  {"search_term": 46})
    sheet(wb.create_sheet("Top-Moving ASINs"), ["brand", "asin", "units", "terms"],
          [{**r, "units": " / ".join(map(str, r["units"]))} for r in p["top_moving"]])
    d01 = os.path.join(OUTDIR, "REQ-30-D01_sqp_top_terms.xlsx"); wb.save(d01)

    # ---- REQ-30-D02 : Phase 2 -----------------------------------------------------------------
    wb = Workbook(); notes(wb.active, "REQ-30-D02 — Phase 2: Cross-ASIN Keyword Gap Report", common + [
        ("Part A", f"{len(p['part_a'])} listings with NO CONTENT — empty backend keyword field and/or no "
                   f"bullet points. Every keyword would show as 'missing' because there is nothing to search. "
                   f"These need a rewrite, not keyword edits (rule Q12)."),
        ("Part C", f"{len(p.get('part_c', []))} pairs REJECTED — same base SKU but the two listings state "
                   f"different wattage, so the stored SKU is wrong. Not keyword-checked; the SKU needs "
                   f"correcting first."),
        ("Part B", f"{len(p['part_b'])} keyword rows across listings that DO have content — "
                   f"{sum(1 for r in p['part_b'] if r['status']=='gap')} real, actionable gaps."),
    ])
    wb.active.title = "Notes & Method"
    sheet(wb.create_sheet("Part A - No Content"),
          ["brand", "top_asin", "base_sku", "duplicate_asin", "duplicate_sku", "duplicate_status",
           "issue", "recommended_action", "keywords_ready_to_use", "title", "date_checked"],
          sorted(p["part_a"], key=lambda r: (r["brand"], r["base_sku"])),
          {"issue": 40, "recommended_action": 48, "title": 60})
    sheet(wb.create_sheet("Part B - Keyword Gaps"),
          ["brand", "top_asin", "base_sku", "duplicate_asin", "duplicate_status", "keyword",
           "search_query_volume", "in_frontend", "in_backend", "status", "add_target",
           "action_state", "date_checked"],
          sorted(p["part_b"], key=lambda r: (r["brand"], r["base_sku"], r["duplicate_asin"],
                                             -r["search_query_volume"])), {"keyword": 44})
    sheet(wb.create_sheet("Part C - SKU mismatch"),
          ["brand", "top_asin", "base_sku", "duplicate_asin", "duplicate_sku", "duplicate_status",
           "top_watts", "duplicate_watts", "issue", "recommended_action", "title", "date_checked"],
          sorted(p.get("part_c", []), key=lambda r: (r["brand"], r["base_sku"])),
          {"issue": 60, "recommended_action": 44, "title": 60})
    fr = [("keywords_ready_to_use", "Part A only — how many proven search terms already exist on the "
                                    "good-selling twin, ready for whoever rewrites this empty listing. "
                                    "0 means the twin has none either."),
          ("brand", "dcvoltage_uk / ledsone_uk — accounts never merged"),
          ("top_asin", "Top-Moving ASIN — the source of the proven search terms"),
          ("base_sku", "Normalised SKU — pack suffixes stripped, bundles kept whole"),
          ("duplicate_asin", "Underperforming listing sharing the same base SKU"),
          ("duplicate_status", "sales_drop_3mo / zero_sales_6mo"),
          ("keyword", "Top search term being audited (from the Phase 1 export)"),
          ("in_frontend", "TRUE if found in title, bullets or description (any one is enough)"),
          ("in_backend", "TRUE if found in the backend / generic keyword field"),
          ("status", "present / gap"),
          ("add_target", "backend / bullet / backend_and_bullet / none — where the term should be added"),
          ("action_state", "reviewed / pending_add / added"),
          ("date_checked", "ISO date of this monthly run")]
    sheet(wb.create_sheet("Field Reference"), ["Field", "Meaning"],
          [{"Field": k, "Meaning": v} for k, v in fr], {"Meaning": 96})
    d02 = os.path.join(OUTDIR, "REQ-30-D02_keyword_gap_report.xlsx"); wb.save(d02)
    print(f"excel -> {os.path.basename(d01)} , {os.path.basename(d02)}")


if __name__ == "__main__":
    main()

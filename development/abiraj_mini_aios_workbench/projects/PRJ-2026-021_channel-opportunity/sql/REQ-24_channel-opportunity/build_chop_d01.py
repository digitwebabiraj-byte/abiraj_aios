"""
build_chop_d01.py — REQ-24-D01 Channel Opportunity report (chop / PRJ-2026-021).

One generator module (the one-fetch-path rule). Reads a governed JSON snapshot of per-base-SKU
units-by-channel (Germany, rolling 90 days, Completed) pulled READ-ONLY from the curated warehouse
`public.order_transaction` via the AIOS knowledge-base MCP (docs.ledsone.co.uk), classifies each SKU's
cross-channel Opportunity + Action on the DOCUMENTED DEFAULT rules below (owner-pending: Mahima), and
renders the Excel deliverable.

Metric = UNITS (cross-channel comparable; avoids the DST currency trap). Grain = one row per internal
base SKU (order_transaction.sku is platform-independent; summing by sku already consolidates eBay
item_id sprawl). Money is deliberately NOT used for the comparison.

DEFAULT classification rules (Notes tab documents these; Mahima confirms before sign-off):
  Let sh/am/eb = 90-day units per channel, total = sh+am+eb, leader = max(sh,am,eb).
  Only SKUs with leader >= FLOOR(10) are considered (a real seller, not 1-off noise).
  * Missing channel   — at least one channel == 0 units. Action: "Create <missing channel(s)> listing".
  * Shopify winner    — all 3 > 0, Shopify is the top channel AND Shopify >= 50% of total.
                        Action: "Improve Amazon/eBay listing".
  * Marketplace winner— all 3 > 0, (Amazon+eBay) >= 60% of total AND Shopify <= 20% of total.
                        Action: "Add Shopify promotion".
  * Balanced          — everything else; NOT an opportunity, excluded from the table.
"""
import ast, json, os, re, sys
from datetime import date
from decimal import Decimal

HERE = os.path.dirname(os.path.abspath(__file__))
SNAP = os.path.join(HERE, "chop_payload_2026-08-05.json")
RAW  = sys.argv[1] if len(sys.argv) > 1 else None
OUT_DIR = os.path.abspath(os.path.join(HERE, "..", "..", "evidence", "final_outputs",
                                       "REQ-24_channel-opportunity"))
OUT_XLSX = os.path.join(OUT_DIR, "REQ-24-D01_channel_opportunity.xlsx")
OUT_HTML = os.path.join(OUT_DIR, "REQ-24-D01_channel_opportunity.html")

FLOOR = 10
WINDOW_DAYS = 90
DATA_THROUGH = "2026-08-04"
MARKET = "Germany"


def load_snapshot():
    """Build the governed JSON snapshot from the raw mcp.ledsone result file (once), else read snapshot.

    Accepts the raw Postgres-MCP export {"data": {"rows": [...]}} — the per-base-SKU units-by-channel
    pivot pulled from order_management (Germany, Completed, 90d, clean-SKU = strip -IDE)."""
    if RAW and os.path.exists(RAW):
        outer = json.loads(open(RAW, "r", encoding="utf-8").read())
        rows = outer["data"]["rows"] if "data" in outer else ast.literal_eval(
            re.sub(r"Decimal\('(-?\d+(?:\.\d+)?)'\)", r"\1", outer["result"][0]["text"]))
        clean = [{"sku": r["sku"],
                  "shopify_u": int(r["shopify_u"]),
                  "amazon_u": int(r["amazon_u"]),
                  "ebay_u": int(r["ebay_u"]),
                  "total_u": int(r["total_u"])} for r in rows]
        json.dump({"generated": DATA_THROUGH, "market": MARKET, "window_days": WINDOW_DAYS,
                   "metric": "units", "source": "raw mcp.ledsone order_management (clean-SKU: strip -IDE)",
                   "rows": clean}, open(SNAP, "w", encoding="utf-8"), indent=1)
        return clean
    return json.load(open(SNAP, "r", encoding="utf-8"))["rows"]


def classify(sh, am, eb):
    total = sh + am + eb
    leader = max(sh, am, eb)
    if leader < FLOOR:
        return None, None
    missing = [name for name, v in (("Shopify", sh), ("Amazon", am), ("eBay", eb)) if v == 0]
    if missing:
        return "Missing channel", "Create " + " + ".join(missing) + " listing"
    # all three > 0
    if sh == leader and sh >= 0.50 * total:
        weak = [name for name, v in (("Amazon", am), ("eBay", eb)) if v < 0.30 * leader]
        tgt = "/".join(weak) if weak else "Amazon/eBay"
        return "Shopify winner", f"Improve {tgt} listing"
    if (am + eb) >= 0.60 * total and sh <= 0.20 * total:
        return "Marketplace winner", "Add Shopify promotion"
    return None, None  # Balanced — not an opportunity


def build():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    rows = load_snapshot()
    opps = []
    for r in rows:
        sh, am, eb = r["shopify_u"], r["amazon_u"], r["ebay_u"]
        cls, act = classify(sh, am, eb)
        if cls is None:
            continue
        opps.append({**r, "opportunity": cls, "action": act})
    opps.sort(key=lambda x: (x["opportunity"], -x["total_u"]))

    os.makedirs(OUT_DIR, exist_ok=True)
    wb = openpyxl.Workbook()

    # ---- Notes & Method tab ----
    nt = wb.active
    nt.title = "Notes & Method"
    FONT = "Arial"
    notes = [
        ("REQ-24-D01 — Channel Opportunity (chop / PRJ-2026-021)", True, 14),
        ("", False, 11),
        ("What: for each product (internal base SKU), units sold laid side by side across Shopify, "
         "Amazon and eBay — to surface products that sell well in one channel but are weak or MISSING "
         "in others, so the listing gap can be closed.", False, 11),
        (f"Scope: market = {MARKET}; channels = Shopify / Amazon / eBay; order_status = Completed.", False, 11),
        (f"Window: rolling {WINDOW_DAYS} days, data through {DATA_THROUGH} (last complete day).", False, 11),
        ("Metric: UNITS (SUM(quantity)). Units are used — not revenue — because the three channels are "
         "compared side by side and marketplace revenue is in each marketplace's own currency (no FX "
         "table; the DST currency rule). Revenue can be added if Mahima prefers.", False, 11),
        ("Grain: one row per internal base SKU (order_transaction.sku is platform-independent; summing "
         "by SKU already consolidates eBay item_id sprawl).", False, 11),
        ("Source: RAW ledsone Postgres DB via the mcp.ledsone.co.uk MCP, READ-ONLY — order_management "
         "(orders + order_item_info + sub_source + source). Germany = market_place '10'; channels via "
         "source.source_name; units = order_item_info.item_quantity. Query patterns per the AIOS "
         "knowledge-base (docs.ledsone.co.uk) text-to-sql-multi skill.", False, 11),
        ("Clean-SKU step (mandatory): base SKU = resolved inventory SKU (order_item_info.real_sku, else "
         "item_sku) with the listing suffix '-IDE' stripped, so a product's Shopify/Amazon/eBay listings "
         "roll up to one row (proven: LDMST64E274 = LDMST64E274-IDE + LDMST64E274). Multi-packs (2PK…) and "
         "combos (SKUs containing '+') are distinct products and kept separate.", False, 11),
        ("", False, 11),
        ("Opportunity classes + Action (DEFAULT rules — pending Mahima's confirmation):", True, 12),
        (f"  Only SKUs whose top channel sold >= {FLOOR} units in the window are flagged (real sellers).", False, 11),
        ("  Missing channel  — at least one channel sold 0 units. Action: Create the missing listing(s). "
         "The clearest opportunity: proven demand, zero coverage somewhere.", False, 11),
        ("  Shopify winner   — all three channels > 0, Shopify is the top channel AND >= 50% of total "
         "units. Action: Improve Amazon/eBay listing (the weak marketplace).", False, 11),
        ("  Marketplace winner — all three > 0, Amazon+eBay >= 60% of total AND Shopify <= 20% of total. "
         "Action: Add Shopify promotion.", False, 11),
        ("  (SKUs selling evenly across all channels are 'balanced' — not an opportunity — and excluded.)", False, 11),
        ("", False, 11),
        ("These thresholds are documented DEFAULTS, not final. Trend/threshold sign-off is owner-pending "
         "(Mahima). No number below is a sample from the source mock-up — every figure is live warehouse data.", False, 11),
    ]
    r = 1
    for txt, bold, size in notes:
        c = nt.cell(row=r, column=1, value=txt)
        c.font = Font(name=FONT, bold=bold, size=size)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    nt.column_dimensions["A"].width = 118

    # ---- Channel Opportunity tab ----
    ws = wb.create_sheet("Channel Opportunity")
    headers = ["SKU", "Shopify Sales", "Amazon Sales", "eBay Sales",
               "Total Units", "Opportunity", "Action"]
    hdr_fill = PatternFill("solid", fgColor="1F4E5F")
    hdr_font = Font(name=FONT, bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cls_fill = {"Missing channel": "FCE4D6", "Shopify winner": "E2EFDA", "Marketplace winner": "DDEBF7"}

    ws.append(headers)
    for ci, _ in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci)
        c.fill = hdr_fill; c.font = hdr_font; c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center")

    for o in opps:
        ws.append([o["sku"], o["shopify_u"], o["amazon_u"], o["ebay_u"],
                   o["total_u"], o["opportunity"], o["action"]])
        rr = ws.max_row
        fill = cls_fill.get(o["opportunity"])
        for ci in range(1, len(headers) + 1):
            cell = ws.cell(row=rr, column=ci)
            cell.font = Font(name=FONT, size=10)
            cell.border = border
            if ci in (2, 3, 4, 5):
                cell.alignment = Alignment(horizontal="center")
            if fill:
                cell.fill = PatternFill("solid", fgColor=fill)

    widths = [30, 13, 13, 11, 12, 18, 32]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{ws.max_row}"

    wb.save(OUT_XLSX)
    total_base = len(rows)
    write_dashboard(opps, total_base)

    # ---- console summary ----
    from collections import Counter
    cnt = Counter(o["opportunity"] for o in opps)
    print("saved:", OUT_XLSX)
    print("saved:", OUT_HTML)
    print("total opportunity rows:", len(opps), "/ base SKUs:", total_base)
    for k in ("Missing channel", "Shopify winner", "Marketplace winner"):
        print(f"  {k}: {cnt.get(k,0)}")


def write_dashboard(opps, total_base):
    """Self-contained interactive HTML dashboard (full-screen, light theme)."""
    data = json.dumps([{"s": o["sku"], "sh": o["shopify_u"], "am": o["amazon_u"],
                        "eb": o["ebay_u"], "t": o["total_u"], "o": o["opportunity"],
                        "a": o["action"]} for o in sorted(opps, key=lambda x: -x["total_u"])])
    from collections import Counter
    c = Counter(o["opportunity"] for o in opps)
    units = sum(o["total_u"] for o in opps)
    html = _HTML.replace("__DATA__", data) \
               .replace("__TOTAL__", str(len(opps))).replace("__BASE__", f"{total_base:,}") \
               .replace("__MISS__", str(c.get("Missing channel", 0))) \
               .replace("__MKT__", str(c.get("Marketplace winner", 0))) \
               .replace("__SHOP__", str(c.get("Shopify winner", 0))) \
               .replace("__UNITS__", f"{units:,}") \
               .replace("__THROUGH__", DATA_THROUGH).replace("__WINDOW__", str(WINDOW_DAYS)) \
               .replace("__MARKET__", MARKET)
    open(OUT_HTML, "w", encoding="utf-8").write(html)


_HTML = r"""<!doctype html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Channel Opportunity — REQ-24-D01</title>
<style>
:root{
  --bg:#f4f7fb; --bg2:#eaf0f7; --panel:#ffffff; --ink:#0d1b2a; --ink2:#33475b; --muted:#6b7f95;
  --line:#e6edf4; --line2:#eef3f8;
  --brand:#0d9488; --brand2:#0e7490; --brand-soft:#e6f7f5;
  --miss:#f59e0b; --miss-d:#b45309; --miss-bg:#fef4e6;
  --mkt:#3b82f6;  --mkt-d:#1d4ed8;  --mkt-bg:#eaf1fe;
  --shop:#22c55e; --shop-d:#15803d; --shop-bg:#e9f8ee;
  --sh:#22c55e; --am:#f59e0b; --eb:#3b82f6;
  --shadow:0 1px 2px rgba(13,27,42,.05),0 8px 24px -12px rgba(13,27,42,.14);
  --radius:16px;
}
*{box-sizing:border-box}
html,body{margin:0;height:100%;overflow:hidden}
body{
  background:
    radial-gradient(1200px 500px at 82% -12%, #dff3f0 0%, rgba(223,243,240,0) 60%),
    radial-gradient(900px 480px at -8% 8%, #e7eefb 0%, rgba(231,238,251,0) 55%),
    var(--bg);
  color:var(--ink);
  font:14px/1.5 "Inter",-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,Arial,sans-serif;
  -webkit-font-smoothing:antialiased;
}
.app{display:flex;flex-direction:column;height:100vh;max-width:1680px;margin:0 auto;padding:20px 26px 16px}

/* ---------- header ---------- */
.top{display:flex;align-items:flex-start;justify-content:space-between;gap:16px;margin-bottom:16px}
.brand{display:flex;align-items:center;gap:14px}
.logo{width:46px;height:46px;border-radius:13px;display:grid;place-items:center;color:#fff;
  background:linear-gradient(135deg,var(--brand),var(--brand2));box-shadow:0 8px 20px -8px rgba(13,148,136,.7)}
.logo svg{width:24px;height:24px}
h1{margin:0;font-size:22px;font-weight:800;letter-spacing:-.4px}
h1 span{background:linear-gradient(90deg,var(--brand),var(--brand2));-webkit-background-clip:text;background-clip:text;color:transparent}
.sub{color:var(--muted);font-size:12.5px;margin-top:3px}
.meta{display:flex;gap:8px;flex-wrap:wrap;justify-content:flex-end}
.tagpill{background:var(--panel);border:1px solid var(--line);border-radius:999px;padding:6px 12px;
  font-size:12px;font-weight:600;color:var(--ink2);box-shadow:var(--shadow)}
.tagpill b{color:var(--brand2)}
.btn{border:1px solid var(--line);background:var(--panel);border-radius:10px;padding:8px 12px;font-size:12.5px;
  font-weight:700;color:var(--ink2);cursor:pointer;box-shadow:var(--shadow);display:inline-flex;align-items:center;gap:7px}
.btn:hover{border-color:var(--brand);color:var(--brand2)}
.btn svg{width:15px;height:15px}

/* ---------- KPIs ---------- */
.kpis{display:grid;grid-template-columns:repeat(5,1fr);gap:14px;margin-bottom:14px}
.kpi{position:relative;background:var(--panel);border:1px solid var(--line);border-radius:var(--radius);
  padding:16px 18px;box-shadow:var(--shadow);overflow:hidden}
.kpi::before{content:"";position:absolute;left:0;top:0;bottom:0;width:4px;background:var(--brand)}
.kpi.k-miss::before{background:var(--miss)} .kpi.k-mkt::before{background:var(--mkt)}
.kpi.k-shop::before{background:var(--shop)} .kpi.k-base::before{background:#94a3b8}
.kpi .row{display:flex;align-items:center;justify-content:space-between}
.kpi .n{font-size:31px;font-weight:800;letter-spacing:-1px;line-height:1}
.kpi .l{color:var(--muted);font-size:11.5px;margin-top:7px;text-transform:uppercase;letter-spacing:.5px;font-weight:600}
.kpi .ic{width:34px;height:34px;border-radius:10px;display:grid;place-items:center;background:var(--brand-soft);color:var(--brand2)}
.kpi.k-miss .ic{background:var(--miss-bg);color:var(--miss-d)}
.kpi.k-mkt .ic{background:var(--mkt-bg);color:var(--mkt-d)}
.kpi.k-shop .ic{background:var(--shop-bg);color:var(--shop-d)}
.kpi.k-base .ic{background:#eef2f6;color:#64748b}
.kpi .ic svg{width:18px;height:18px}

/* ---------- distribution bar ---------- */
.dist{display:flex;height:12px;border-radius:999px;overflow:hidden;margin-bottom:16px;box-shadow:var(--shadow);border:1px solid var(--line)}
.dist i{display:block;height:100%}
.dist .d-miss{background:linear-gradient(90deg,#fbbf24,#f59e0b)}
.dist .d-mkt{background:linear-gradient(90deg,#60a5fa,#3b82f6)}
.dist .d-shop{background:linear-gradient(90deg,#4ade80,#22c55e)}

/* ---------- toolbar ---------- */
.toolbar{display:flex;flex-wrap:wrap;gap:9px;align-items:center;margin-bottom:12px}
.chip{border:1px solid var(--line);background:var(--panel);color:var(--ink2);border-radius:999px;
  padding:8px 15px;font-size:13px;cursor:pointer;font-weight:700;box-shadow:var(--shadow);transition:.15s;
  display:inline-flex;align-items:center;gap:8px}
.chip .dot{width:9px;height:9px;border-radius:50%}
.chip .dot.miss{background:var(--miss)} .chip .dot.mkt{background:var(--mkt)} .chip .dot.shop{background:var(--shop)} .chip .dot.all{background:var(--brand)}
.chip .c{background:var(--bg2);border-radius:999px;padding:1px 8px;font-size:11.5px;color:var(--muted)}
.chip:hover{transform:translateY(-1px)}
.chip.on{color:#fff;border-color:transparent}
.chip.on .c{background:rgba(255,255,255,.25);color:#fff}
.chip.on[data-f="all"]{background:linear-gradient(135deg,var(--brand),var(--brand2))}
.chip.on.miss{background:linear-gradient(135deg,#fbbf24,#f59e0b)} .chip.on.miss .dot{background:#fff}
.chip.on.mkt{background:linear-gradient(135deg,#60a5fa,#3b82f6)} .chip.on.mkt .dot{background:#fff}
.chip.on.shop{background:linear-gradient(135deg,#4ade80,#22c55e)} .chip.on.shop .dot{background:#fff}
.search{margin-left:auto;position:relative;min-width:230px;max-width:360px;flex:1}
.search svg{position:absolute;left:12px;top:50%;transform:translateY(-50%);width:16px;height:16px;color:var(--muted)}
.search input{width:100%;padding:10px 14px 10px 36px;border:1px solid var(--line);border-radius:11px;font-size:13.5px;background:var(--panel);box-shadow:var(--shadow)}
.search input:focus{outline:none;border-color:var(--brand);box-shadow:0 0 0 3px var(--brand-soft)}

/* ---------- table ---------- */
.card{flex:1;background:var(--panel);border:1px solid var(--line);border-radius:var(--radius);overflow:hidden;box-shadow:var(--shadow);display:flex;flex-direction:column;min-height:0}
.scroll{overflow:auto;flex:1;min-height:0}
table{width:100%;border-collapse:separate;border-spacing:0;font-size:13.5px}
thead th{position:sticky;top:0;z-index:2;background:#f8fafc;backdrop-filter:saturate(1.1);
  border-bottom:1px solid var(--line);text-align:left;padding:13px 16px;font-size:11px;letter-spacing:.5px;
  color:var(--muted);text-transform:uppercase;font-weight:700;cursor:pointer;white-space:nowrap;user-select:none}
thead th.num{text-align:right} thead th:hover{color:var(--brand2)}
thead th .ar{opacity:.4;font-size:10px;margin-left:3px}
tbody td{padding:11px 16px;border-bottom:1px solid var(--line2);vertical-align:middle}
tbody tr{transition:background .12s} tbody tr:hover{background:#f6fbfb}
td.rank{color:#aab8c6;font-variant-numeric:tabular-nums;font-size:12px;width:44px}
td.sku{font-family:ui-monospace,"SF Mono",Menlo,Consolas,monospace;font-weight:700;color:var(--ink);letter-spacing:-.2px}
.cov{display:inline-flex;gap:4px;margin-left:9px;vertical-align:middle}
.cov i{width:7px;height:7px;border-radius:50%;background:#dbe4ec}
.cov i.on-sh{background:var(--sh)} .cov i.on-am{background:var(--am)} .cov i.on-eb{background:var(--eb)}
td.num{text-align:right;font-variant-numeric:tabular-nums;white-space:nowrap}
.cellbar{display:inline-flex;align-items:center;gap:8px;justify-content:flex-end}
.cellbar .bar{height:7px;border-radius:4px;min-width:3px}
.cellbar .v{min-width:26px;text-align:right;font-weight:700}
.v.sh{color:var(--shop-d)} .v.am{color:var(--miss-d)} .v.eb{color:var(--mkt-d)} .v.z{color:#c6d0da;font-weight:500}
td.tot b{font-size:14px}
.badge{display:inline-flex;align-items:center;gap:7px;padding:4px 11px;border-radius:999px;font-size:12px;font-weight:700;white-space:nowrap}
.badge .dot{width:7px;height:7px;border-radius:50%}
.badge.miss{background:var(--miss-bg);color:var(--miss-d)} .badge.miss .dot{background:var(--miss)}
.badge.mkt{background:var(--mkt-bg);color:var(--mkt-d)} .badge.mkt .dot{background:var(--mkt)}
.badge.shop{background:var(--shop-bg);color:var(--shop-d)} .badge.shop .dot{background:var(--shop)}
td.action{color:var(--ink2)}
.footbar{display:flex;align-items:center;justify-content:space-between;gap:12px;padding:9px 4px 2px;color:var(--muted);font-size:11.5px}
.footbar code{background:var(--bg2);padding:1px 6px;border-radius:5px;font-size:11px}
.empty{padding:60px;text-align:center;color:var(--muted)}
@media(max-width:1100px){.kpis{grid-template-columns:repeat(3,1fr)}.app{padding:14px}}
@media(max-width:720px){.kpis{grid-template-columns:repeat(2,1fr)}}
</style></head><body>
<div class="app">
  <div class="top">
    <div class="brand">
      <div class="logo"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.2" stroke-linecap="round" stroke-linejoin="round"><path d="M3 3v18h18"/><path d="M7 14l3-4 3 3 4-6"/></svg></div>
      <div>
        <h1>Channel <span>Opportunity</span></h1>
        <div class="sub">Products selling well in one channel but weak or missing in the others — where to create or promote a listing.</div>
      </div>
    </div>
    <div class="meta">
      <span class="tagpill"><b>__MARKET__</b> · units</span>
      <span class="tagpill">Rolling <b>__WINDOW__</b>d · to <b>__THROUGH__</b></span>
      <span class="tagpill">Source: <b>raw ledsone</b></span>
      <button class="btn" onclick="fs()"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M8 3H5a2 2 0 0 0-2 2v3m18 0V5a2 2 0 0 0-2-2h-3M3 16v3a2 2 0 0 0 2 2h3m13-5v3a2 2 0 0 1-2 2h-3"/></svg>Full screen</button>
    </div>
  </div>

  <div class="kpis">
    <div class="kpi"><div class="row"><div class="n">__TOTAL__</div><div class="ic"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><circle cx="12" cy="12" r="9"/><path d="M12 8v8M8 12h8"/></svg></div></div><div class="l">Opportunities found</div></div>
    <div class="kpi k-miss"><div class="row"><div class="n">__MISS__</div><div class="ic"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M12 2v20M2 12h20" opacity=".25"/><circle cx="12" cy="12" r="9"/><path d="M12 8v4l3 2"/></svg></div></div><div class="l">Missing channel</div></div>
    <div class="kpi k-mkt"><div class="row"><div class="n">__MKT__</div><div class="ic"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M3 3h18v18H3z" opacity=".2"/><path d="M7 15l3-3 2 2 5-6"/></svg></div></div><div class="l">Marketplace winner</div></div>
    <div class="kpi k-shop"><div class="row"><div class="n">__SHOP__</div><div class="ic"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M6 2 3 6v14a2 2 0 0 0 2 2h14a2 2 0 0 0 2-2V6l-3-4Z"/><path d="M3 6h18M16 10a4 4 0 0 1-8 0"/></svg></div></div><div class="l">Shopify winner</div></div>
    <div class="kpi k-base"><div class="row"><div class="n">__BASE__</div><div class="ic"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><ellipse cx="12" cy="5" rx="9" ry="3"/><path d="M3 5v14c0 1.7 4 3 9 3s9-1.3 9-3V5M3 12c0 1.7 4 3 9 3s9-1.3 9-3"/></svg></div></div><div class="l">Base SKUs scanned</div></div>
  </div>

  <div class="dist" id="dist" title="Share of opportunities by class"></div>

  <div class="toolbar">
    <span class="chip on" data-f="all" onclick="setF(this,'all')"><span class="dot all"></span>All <span class="c">__TOTAL__</span></span>
    <span class="chip miss" data-f="Missing channel" onclick="setF(this,'Missing channel')"><span class="dot miss"></span>Missing channel <span class="c">__MISS__</span></span>
    <span class="chip mkt" data-f="Marketplace winner" onclick="setF(this,'Marketplace winner')"><span class="dot mkt"></span>Marketplace winner <span class="c">__MKT__</span></span>
    <span class="chip shop" data-f="Shopify winner" onclick="setF(this,'Shopify winner')"><span class="dot shop"></span>Shopify winner <span class="c">__SHOP__</span></span>
    <div class="search"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><circle cx="11" cy="11" r="7"/><path d="m21 21-4.3-4.3"/></svg><input id="q" placeholder="Search SKU or action…" oninput="render()"></div>
  </div>

  <div class="card"><div class="scroll"><table>
    <thead><tr>
      <th style="width:44px">#</th>
      <th onclick="sortBy('s')">SKU<span class="ar" id="ar-s"></span></th>
      <th class="num" onclick="sortBy('sh')">Shopify<span class="ar" id="ar-sh"></span></th>
      <th class="num" onclick="sortBy('am')">Amazon<span class="ar" id="ar-am"></span></th>
      <th class="num" onclick="sortBy('eb')">eBay<span class="ar" id="ar-eb"></span></th>
      <th class="num" onclick="sortBy('t')">Total<span class="ar" id="ar-t">▾</span></th>
      <th onclick="sortBy('o')">Opportunity<span class="ar" id="ar-o"></span></th>
      <th onclick="sortBy('a')">Recommended action<span class="ar" id="ar-a"></span></th>
    </tr></thead>
    <tbody id="tb"></tbody>
  </table></div></div>

  <div class="footbar">
    <div>Defaults (owner-locked, pending Mahima): clean base SKU (strip <code>-IDE</code>) · flagged if top channel ≥10 units · Missing = 0 in ≥1 channel · Shopify-winner ≥50% · Marketplace-winner ≥60% &amp; Shopify ≤20%. Read-only; every figure from the live DB.</div>
    <div id="count"></div>
  </div>
</div>
<script>
const DATA=__DATA__;
const CNT={miss:__MISS__,mkt:__MKT__,shop:__SHOP__,total:__TOTAL__};
let filt='all', sortK='t', sortDir=-1;
const CLS={'Missing channel':'miss','Marketplace winner':'mkt','Shopify winner':'shop'};
const MAXCH=Math.max(...DATA.map(d=>Math.max(d.sh,d.am,d.eb)),1);
document.getElementById('dist').innerHTML=
  `<i class="d-miss" style="width:${CNT.miss/CNT.total*100}%"></i>`+
  `<i class="d-mkt" style="width:${CNT.mkt/CNT.total*100}%"></i>`+
  `<i class="d-shop" style="width:${CNT.shop/CNT.total*100}%"></i>`;
function fs(){const e=document.documentElement;if(!document.fullscreenElement){e.requestFullscreen&&e.requestFullscreen();}else{document.exitFullscreen&&document.exitFullscreen();}}
function setF(el,f){filt=f;document.querySelectorAll('.chip').forEach(c=>c.classList.remove('on'));el.classList.add('on');render();}
function sortBy(k){if(sortK===k)sortDir*=-1;else{sortK=k;sortDir=(k==='s'||k==='o'||k==='a')?1:-1;}
  document.querySelectorAll('.ar').forEach(a=>a.textContent='');
  document.getElementById('ar-'+k).textContent=sortDir<0?'▾':'▴';render();}
function chcell(v,cl){
  if(!v)return '<td class="num"><span class="cellbar"><span class="v z">0</span></span></td>';
  const w=Math.max(3,Math.round(v/MAXCH*54));
  return `<td class="num"><span class="cellbar"><span class="bar" style="width:${w}px;background:var(--${cl})"></span><span class="v ${cl}">${v}</span></span></td>`;
}
function render(){
  const q=document.getElementById('q').value.trim().toLowerCase();
  let rows=DATA.filter(d=>(filt==='all'||d.o===filt)&&(!q||d.s.toLowerCase().includes(q)||d.a.toLowerCase().includes(q)));
  rows.sort((a,b)=>{let x=a[sortK],y=b[sortK];if(typeof x==='string')return x.localeCompare(y)*sortDir;return (x-y)*sortDir;});
  document.getElementById('count').innerHTML=`<b style="color:var(--ink2)">${rows.length}</b> of ${DATA.length} shown`;
  const tb=document.getElementById('tb');
  if(!rows.length){tb.innerHTML='<tr><td colspan="8" class="empty">No matching SKUs.</td></tr>';return;}
  tb.innerHTML=rows.map((d,i)=>{const k=CLS[d.o];
    const cov=`<span class="cov"><i class="${d.sh?'on-sh':''}"></i><i class="${d.am?'on-am':''}"></i><i class="${d.eb?'on-eb':''}"></i></span>`;
    return `<tr><td class="rank">${i+1}</td>`
      +`<td class="sku">${d.s}${cov}</td>`
      +chcell(d.sh,'sh')+chcell(d.am,'am')+chcell(d.eb,'eb')
      +`<td class="num tot"><b>${d.t}</b></td>`
      +`<td><span class="badge ${k}"><span class="dot"></span>${d.o}</span></td>`
      +`<td class="action">${d.a}</td></tr>`;}).join('');
}
render();
</script>
</body></html>"""

if __name__ == "__main__":
    build()

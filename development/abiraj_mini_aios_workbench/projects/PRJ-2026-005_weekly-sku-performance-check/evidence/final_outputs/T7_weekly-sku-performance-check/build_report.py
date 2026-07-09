#!/usr/bin/env python3
"""
Table 7 - Weekly SKU Performance Check | Thuwaraga (PH-2026-07-THUW07)
Renders data.json into an .xlsx matching the approved template column order
and colour bands (purple SKU summary / blue ASIN detail / green / red / orange).

No DB access; shapes data.json only. Uses the same grouping logic as build_html.py.

Usage:  python build_report.py
"""
import json, os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from build_html import load, build_groups  # reuse the governed grouping

HERE = os.path.dirname(os.path.abspath(__file__))
OUT  = os.path.join(HERE, "Table7_Weekly_SKU_Performance_Thuwaraga.xlsx")

COLS = ["SKU / ASIN", "Row Type", "Product Name", "Platform", "Account Name",
        "Week Start", "Week End", "Amazon Orders", "eBay Orders", "B&Q Orders",
        "TOTAL Orders", "Performing?", "Action Required"]

PURPLE = PatternFill("solid", fgColor="E9D8FD")
BLUE   = PatternFill("solid", fgColor="E7F0FE")
RED    = PatternFill("solid", fgColor="FDE4E4")
HEAD   = PatternFill("solid", fgColor="111827")
THIN   = Side(style="thin", color="D6D9E0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def main():
    data = load()
    m = data["meta"]
    groups = build_groups(data)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Table 7 - SKU Performance"

    # title band
    ws.append(["Table 7 - Weekly SKU Performance Check | All Platforms UK (Amazon / eBay / B&Q)"])
    ws.append([f"Portfolio Holder: Thuwaraga | {m['project_code']} | Source DB: {m['database']} "
               f"(user_name={m['source_user_name']}) | Run {m['run_date']} | "
               f"Window {m['week_start']} to {m['week_end']} (rolling 7 days, Thursday) | "
               f"Data snapshot as of {m.get('snapshot_at','n/a')} (live DB; counts settle ~1-2 days)"])
    ws.append([])
    hdr_row = 4
    ws.append(COLS)
    for c in range(1, len(COLS) + 1):
        cell = ws.cell(hdr_row, c)
        cell.fill = HEAD
        cell.font = Font(color="FFFFFF", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"].font = Font(color="6B7280", size=9)

    ws2 = m["week_start"]
    we2 = m["week_end"]
    for g in groups:
        base_lbl = g["base"] + (f"  [+{g['skus']-1} SKUs]" if g["merged"] else "")
        ws.append([base_lbl, "SKU SUMMARY", g["name"], "All Platforms", "-",
                   ws2, we2, g["amazon"], g["ebay"], g["bq"], g["total"],
                   g["perf"], g["action"]])
        r = ws.max_row
        for c in range(1, len(COLS) + 1):
            ws.cell(r, c).fill = PURPLE
            ws.cell(r, c).font = Font(bold=True, size=10)
        for row in g["rows"]:
            sku_lbl = row["sku"] + ("  [variant]" if row["variant"] else "")
            ws.append([sku_lbl, row["ref"], row["name"], row["platform"], row["account"],
                       ws2, we2, row["amazon"], row["ebay"], row["bq"], row["total"],
                       "YES" if row["performing"] else "NO",
                       row["action"]])
            rr = ws.max_row
            fill = BLUE if row["performing"] else RED
            for c in range(1, len(COLS) + 1):
                ws.cell(rr, c).fill = fill

    # borders + number cols
    for row in ws.iter_rows(min_row=hdr_row, max_row=ws.max_row,
                            min_col=1, max_col=len(COLS)):
        for cell in row:
            cell.border = BORDER
            if cell.column in (8, 9, 10, 11):
                cell.alignment = Alignment(horizontal="right")

    widths = [22, 15, 46, 13, 16, 12, 12, 10, 9, 9, 9, 26, 26]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A{hdr_row}:{get_column_letter(len(COLS))}{ws.max_row}"

    wb.save(OUT)
    print("wrote", OUT, os.path.getsize(OUT), "bytes;",
          len(groups), "SKU families,", ws.max_row - hdr_row, "data rows")


if __name__ == "__main__":
    main()

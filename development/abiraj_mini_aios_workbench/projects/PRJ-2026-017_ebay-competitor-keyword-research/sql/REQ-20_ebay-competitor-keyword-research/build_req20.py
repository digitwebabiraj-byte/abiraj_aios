import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
OUT=r"C:\Users\digit\OneDrive\Desktop\Abiraj_AIOS\.claude\worktrees\gifted-keller-42ebdd\development\abiraj_mini_aios_workbench\projects\PRJ-2026-017_ebay-competitor-keyword-research\evidence\final_outputs\REQ-20_ebay-competitor-keyword-research\REQ-20-D01_ebay_competitor_keyword_SOLD.xlsx"
wb=openpyxl.Workbook(); ws=wb.active; ws.title="Competitor & Keyword"
cols=["Product Name","Competitor ID","Brand","Title","Sold Quantity","Price","Feedback Rate","Shipping type","Promotion Type & %","Primary Keywords","Secondary Keywords","Long-Tail Keywords","Notes"]
ws.append(cols)
hf=Font(bold=True,color="FFFFFF"); fill=PatternFill("solid",fgColor="2F5496")
for c in ws[1]: c.font=hf; c.fill=fill; c.alignment=Alignment(wrap_text=True,vertical="top")

KW={
 "Metal Shade Pendant Light":("Metal Shade Pendant Light, Metal Pendant Light, Hanging Pendant Light","Ceiling Pendant Shade, Industrial Pendant Light, Easy Fit Pendant Shade, Kitchen Pendant Light","Black Metal Shade Pendant Light for Kitchen Island, Industrial Easy Fit Metal Pendant Shade E27, Modern Chrome Metal Pendant Ceiling Light"),
 "Wall Light":("Wall Light, Wall Lamp, Wall Sconce","Indoor Wall Light, Bedside Wall Lamp, Industrial Wall Sconce, LED Wall Light","Black Industrial Wall Light for Living Room, Vintage Wall Sconce with Switch, Modern LED Bedside Wall Reading Lamp"),
 "Metal shade Ceiling Light":("Metal Ceiling Light, Ceiling Shade, Metal Pendant Shade","Industrial Ceiling Light, Metal Lampshade, Ceiling Light Fitting, Retro Ceiling Shade","Black Metal Ceiling Light Shade for Kitchen, Industrial Metal Pendant Shade E27, Vintage Metal Ceiling Lampshade Easy Fit"),
 "Glass shade ceiling light":("Glass Ceiling Light, Glass Pendant Light, Glass Shade","Clear Glass Pendant, Glass Lampshade, Ceiling Glass Light Fitting, Modern Glass Pendant","Clear Glass Globe Pendant Light for Kitchen Island, Vintage Glass Ceiling Shade E27, Modern Frosted Glass Hanging Light"),
 "Spider Light":("Spider Light, Spider Pendant, Multi-Head Pendant Light","Sputnik Light, Multi Light Pendant, Industrial Spider Lamp, Adjustable Spider Ceiling Light","Black Multi-Head Spider Pendant Light for Living Room, Industrial Sputnik Ceiling Light 6 Head, Adjustable DIY Spider Chandelier E27"),
 "Cage pendant light":("Cage Pendant Light, Cage Light, Metal Cage Pendant","Industrial Cage Light, Wire Cage Pendant, Vintage Cage Lamp, Cage Ceiling Light","Black Metal Cage Pendant Light for Kitchen, Industrial Wire Cage Hanging Lamp E27, Vintage Cage Ceiling Light Fitting"),
 "Pipe Light":("Pipe Light, Pipe Pendant Light, Steampunk Light","Industrial Pipe Light, Steampunk Pendant, Pipe Wall Light, Water Pipe Lamp","Industrial Steampunk Pipe Pendant Light for Kitchen Island, Black Metal Pipe Ceiling Light E27, Vintage Water Pipe Wall Lamp"),
 "Bulbs":("LED Bulb, Filament Bulb, Light Bulb","Edison Bulb, Vintage LED Bulb, E27 Bulb, Dimmable LED Filament Bulb","E27 Vintage Edison LED Filament Bulb Warm White, Dimmable Squirrel Cage Light Bulb, ST64 Amber Decorative LED Bulb 4W"),
 "Lamp Holder":("Lamp Holder, Bulb Holder, E27 Holder","Ceiling Rose Lamp Holder, Pendant Lamp Holder, Bakelite Bulb Holder, E27 Fitting","E27 Vintage Bakelite Pendant Lamp Holder with Cord Grip, Ceiling Rose Bulb Holder Kit, Screw E27 Fitting with Switch"),
}
SKU={"Metal Shade Pendant Light":"CRSF100BM+PHSH1PBRYB+SCRN70BM+LSFT220BM"}

DATA={
"Metal Shade Pendant Light":[
 ("333675680799","firstchoicelightingoutlet","firstchoicelighting","Set of 2 Modern Three Tier Easy Fit Jewelled Ceiling Light Shade","751","GBP 20.39","99.7% (89.9K)","With postage","-"),
 ("285423964421","picknmix.online","Giggi","Metal & Rope Lamp Shades Ceiling Pendant Light Shade","236","GBP 18.99","99.6% (21.9K)","With postage","Save up to 15% Multi-buy"),
 ("372257187156","value-lights","ValueLights","Ceiling Light Fitting 3 Way Chrome Spotlight Swirl Glass","8,367","GBP 19.99","99.8% (685K)","With postage","-"),
 ("154695326074","homeessenceltd","Innoteck","Ceiling Light Shade Chandelier Pendant Acrylic Crystal Droplet Chrome","354+","GBP 10.99","99.9% (60.1K)","With postage","Save up to 20% Multi-buy"),
 ("201620721277","buybox786","OPTIMAL PRODUCTS","Chandelier Style Ceiling Light Shade Droplet Pendant Acrylic Crystal Bead","65+","GBP 10.95 to 19.95","99.7% (220.4K)","With postage","Save up to 6% Multi-buy"),
],
"Wall Light":[
 ("115699570601","eat_sleep_buy_repeat","(not listed)","Vintage Industrial Wall Light Antique Retro Cage Bulkhead Gold Brass Ship Lamp","257","GBP 64.99","100% (1.4K)","With postage","-"),
 ("234566652967","burgesshandg","(not listed)","Vintage Industrial Wall Light Antique Retro Style Lamp Sconce Fixture Rustic","71","GBP 31.99","99.7% (9.8K)","With postage","-"),
 ("155030374927","theyinltd","Unbranded","Modern Vintage Retro Industrial Rustic Sconce Wall Light","61","GBP 7.99","99.7% (5.5K)","With postage","Save up to 6% Multi-buy"),
 ("116478119802","eat_sleep_buy_repeat","(not listed)","Industrial Bulkhead Light Wall Ceiling Retro Old Marine Brass Gold","61","GBP 69.99","100% (1.4K)","With postage","-"),
 ("137354830445","weshipnow","Does not apply","Vintage Brass Boho Wall Sconces with Tulip Glass Shade for Modern Spaces","2","GBP 81.26","99.8% (28.1K)","With postage","-"),
],
"Metal shade Ceiling Light":[
 ("372257187156","value-lights","ValueLights","Ceiling Light Fitting 3 Way Chrome Spotlight Swirl Glass","8,367","GBP 19.99","99.8% (685K)","With postage","-"),
 ("372336415252","value-lights","ValueLights","Lampshade Ceiling Pendant Light Shade Easy Fit Chandelier","6,442+","GBP 17.99","99.8% (685K)","With postage","Save up to 10% Multi-buy"),
 ("333667118358","firstchoicelightingoutlet","firstchoicelighting","Set of 2 Modern Black Metal Swirl Easy Fit Ceiling Light Shade","268","GBP 23.45","99.7% (89.9K)","With postage","-"),
 ("285423964421","picknmix.online","Giggi","Metal & Rope Lamp Shades Ceiling Pendant Light Shade","236","GBP 18.99","99.6% (21.9K)","With postage","Save up to 15% Multi-buy"),
 ("393696656791","innoteck_uk","Innoteck","Innoteck Ceiling Light Shade - Modern Chandelier Pendant","92+","GBP 10.49","99.8% (10.2K)","With postage","Save up to 20% Multi-buy"),
],
"Glass shade ceiling light":[
 ("333675680799","firstchoicelightingoutlet","firstchoicelighting","Set of 2 Modern Three Tier Easy Fit Jewelled Ceiling Light","751","GBP 20.39","99.7% (89.9K)","With postage","-"),
 ("283064879219","picknmix.online","Giggi","Modern Chandelier Style Ceiling Pendant Light Shade K9","600","GBP 21.99","99.6% (21.9K)","With postage","-"),
 ("272110237788","moonlight_retail_15","MoonLight Retail","New Modern Vintage Industrial Retro Loft Glass Ceiling Lamp Shade","269+","GBP 32.47 to 35.83","100% (3.9K)","With postage","Save up to 5% Multi-buy"),
 ("141200077246","buyitbetter_uk","ElekTek","ElekTek 3-Hook Ceiling Rose Pendant Plate Suspended Glass","59+","GBP 17.25 to 19.99","99.9% (55.5K)","With postage","Save up to 7% Multi-buy"),
 ("356659461782","value-lights","ValueLights","Corinne Spiral Twill Glass Easy Fit Ceiling Pendant Light Shade","11","GBP 34.99 to 37.99","99.8% (685K)","With postage","From GBP 31.49 with coupon"),
],
"Spider Light":[
 ("303682816669","best-love2010","Unbranded","Retro 4/6/8/10/12 Heads Mix-color Chandelier Spider Pendant Ceiling Light","67","GBP 22.99 to 49.99","96.4% (14.7K)","With postage","-"),
 ("303227195662","nationallighting","(not listed)","Adjustable Pendant Light Fitting Ceiling Rose E27 Industrial Spider Hanging Lamp","39","GBP 13.42 to 59.92","99.6% (18.4K)","With postage","-"),
 ("292861051418","best-love2010","Unbranded","Vintage Edison Multiple Adjustable DIY Ceiling Spider Light","34+","GBP 11.99 to 21.99","96.4% (14.7K)","With postage","-"),
 ("236823130326","wowzer-uk","ANWIO","ANWIO 6 Heads Spider Light Chandelier Pendant DIY Spider Rope","1","GBP 29.99","99.8% (13.4K)","With postage","-"),
],
"Cage pendant light":[
 ("372257190104","value-lights","ValueLights","Ceiling Light Shade Geometric Pendant Lampshade Lamp Industrial","951+","GBP 12.99 to 15.99","99.8% (685K)","With postage","Save up to 10% Multi-buy"),
 ("232708905363","value-lights","ValueLights","Industrial Metal Ceiling Pendant Light Shade Easy Fit","289+","GBP 22.99 to 25.99","99.8% (685K)","With postage","From GBP 20.69 with coupon"),
 ("131173303482","thelampfactoryuk","Unbranded","Black Wire Cage Light Shade Steel E27 Industrial Pendant","252","GBP 13.00","99.8% (22.6K)","With postage","-"),
 ("141264804458","thelampfactoryuk","Unbranded","Bronze Wire Cage Light Shade Steel E27 Industrial Pendant","157","GBP 13.00","99.8% (22.6K)","With postage","-"),
 ("262744851290","sj_modern_antiques","(not listed)","Silver Industrial Ceiling Light Vintage Antique Pendant Cage","114","GBP 46.99","98.5% (15.2K)","With postage","-"),
],
"Pipe Light":[
 ("174650219163","nmstoute","Stout & Burg Designs","Steampunk Industrial Iron Pipe Lamp/Light Socket 1-1/4 x 1/2","334+","GBP 3.18 to 10.46","99% (4.4K)","With postage","-"),
 ("333977871043","steampunkeddesignstudio","Unbranded","Industrial Rustic Retro Style Pipe Light Steampunk Desk Table Lamp","148","GBP 41.91","100% (1.4K)","With postage","-"),
 ("201795458338","mrwillieslighting","Mr. Willies","Retro Industrial Pipe Desk Table 3 bulbs steampunk edison","83","GBP 130.88","100% (882)","With postage","-"),
 ("353005301113","happy_home_2019","Unbranded","Iron Vintage Steampunk Stop Valve Light Switch W/ Wire Pipe Lamp","77","GBP 16.22","99.4% (3K)","Free postage","-"),
 ("164791375113","ztt_online","(not listed)","Retro Switch Valve Steampunk Light Industrial Pipe Table Lamp","25","GBP 10.48","99.3% (26.9K)","With postage","-"),
],
"Bulbs":[
 ("223224109836","safield-online","Auraglow","Auraglow Mysa LED Light Bulb Vintage Retro Edison Style Decorative E27 T30 Tube","180","GBP 10.99","99.8% (188.5K)","With postage","-"),
 ("353725905176","auctionzltd","Kanlux","10x EXTRA LARGE GLOBE FILAMENT E27 BULB Vintage Incandescent Edison","10","GBP 70.94","99.7% (84.9K)","With postage","-"),
 ("352802795272","value-lights","ValueLights","Filament LED Light Bulb Decorative Vintage Edison Lightbulb","102+","GBP 5.99 to 19.99","99.8% (685K)","With postage","Save up to 15% Multi-buy"),
 ("142340181955","thelampfactoryuk","Unbranded","Vintage Edison Filament LED Bulb Teardrop E27 ES 4W Clear Warm White Dimmable","169","GBP 3.50","99.8% (22.6K)","With postage","-"),
 ("364884514396","dqpltd","Unbranded","2x 4W (=33W) Vintage T30 Tube Gold LED Filament ES E27 2700K Light Bulbs","5","GBP 9.99","99.6% (185.3K)","With postage","Save up to 8% Multi-buy"),
],
"Lamp Holder":[
 ("396977702027","gmuk.ltd","Luxa","Black Retro E27 Vintage Screw In Light Bulb Lamp Holder Ceiling Rose Pendant Kit","38","GBP 8.99","100% (2.5K)","With postage","Save up to 5% Multi-buy"),
 ("225227368158","universalmerchandise4u","Crown","E27 Ceiling Rose Pendant With Tube And Lamp Holder Brass Light Fitting UK","11","GBP 18.85","100% (1.2K)","With postage","-"),
 ("224707047105","yuki-store-2020","Unbranded","E27 Ceiling Rose Light Fitting Vintage Industrial Flex Pendant Lamp Bulb Holder","18","GBP 10.59","99.6% (2.9K)","Free postage","-"),
 ("274419951502","bandhstore","Crown Lighting","E27 Modern Ceiling Rose Pendant With Tube And Lamp Holder Fitting","2","GBP 24.95","100% (8.5K)","With postage","-"),
 ("184848546669","colchester-electrical-wholesale","Knightsbridge","ES E27 Screw In Bulb Lamp Matt White Pendant Set Lampholder Ceiling Rose","22","GBP 18.99","99.9% (44.1K)","With postage","-"),
],
}

order=["Metal Shade Pendant Light","Wall Light","Metal shade Ceiling Light","Glass shade ceiling light","Spider Light","Cage pendant light","Pipe Light","Bulbs","Lamp Holder"]
total=0; sold_ok=0
for cat in order:
    pk,sk,lt=KW[cat]; rows=DATA[cat]
    for i,r in enumerate(rows):
        cid,seller,brand,title,sold,price,fb,ship,promo=r
        total+=1
        if sold.strip(): sold_ok+=1
        ws.append([cat if i==0 else "", cid, brand, title, sold, price, fb, ship, promo,
                   pk if i==0 else "", sk if i==0 else "", lt if i==0 else "", "Competitor seller: "+seller+"."])

widths=[22,15,20,42,11,17,15,14,22,30,34,46,30]
import openpyxl.utils as u
for i,w in enumerate(widths,1): ws.column_dimensions[u.get_column_letter(i)].width=w
for row in ws.iter_rows(min_row=2):
    for c in row: c.alignment=Alignment(wrap_text=True,vertical="top")
ws.freeze_panes="A2"
try:
    wb.save(OUT)
    print("rebuilt: %d rows, %d with sold history, %d blank" % (total, sold_ok, total-sold_ok))
except PermissionError:
    print("xlsx locked (open in Excel) - skipped save; data still available for import")

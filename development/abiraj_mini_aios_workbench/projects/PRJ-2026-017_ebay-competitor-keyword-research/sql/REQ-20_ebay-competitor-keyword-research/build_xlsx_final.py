# -*- coding: utf-8 -*-
import json, os, sys, io, urllib.request
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import build_req20 as b
from PIL import Image as PImage
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as XLImage
import openpyxl.utils as u

IMG=json.load(open(os.path.join(os.path.dirname(os.path.abspath(__file__)),"images.json"),encoding="utf-8"))
TMP=os.path.join(os.path.dirname(os.path.abspath(__file__)),"imgtmp"); os.makedirs(TMP,exist_ok=True)

def fetch_png(cid,url):
    p=os.path.join(TMP,cid+".png")
    if os.path.exists(p): return p
    try:
        req=urllib.request.Request(url,headers={"User-Agent":"Mozilla/5.0"})
        raw=urllib.request.urlopen(req,timeout=25).read()
        im=PImage.open(io.BytesIO(raw)).convert("RGB")
        im.thumbnail((90,90))
        im.save(p,"PNG")
        return p
    except Exception as e:
        print("img fail",cid,e); return None

cols=["Image","Product Name","Competitor ID","Brand","Title","Sold Quantity","Price",
      "Feedback Rate","Shipping type","Promotion Type & %","Primary Keywords",
      "Secondary Keywords","Long-Tail Keywords","Notes"]
wb=openpyxl.Workbook(); ws=wb.active; ws.title="Competitor & Keyword"
ws.append(cols)
hf=Font(bold=True,color="FFFFFF"); fill=PatternFill("solid",fgColor="2F5496")
for c in ws[1]: c.font=hf; c.fill=fill; c.alignment=Alignment(wrap_text=True,vertical="top")

r=1
for cat in b.order:
    pk,sk,lt=b.KW[cat]
    data=sorted(b.DATA[cat], key=lambda x:-(int((x[4] or '0').replace(',','').replace('+','')) if (x[4] or '0').replace(',','').replace('+','').isdigit() else 0))
    for i,(cid,seller,brand,title,sold,price,fb,ship,promo) in enumerate(data):
        r+=1
        ws.append(["", cat if i==0 else "", cid, brand, title, sold, price.replace("GBP","GBP "),
                   fb, ship, ("" if promo=="-" else promo),
                   pk if i==0 else "", sk if i==0 else "", lt if i==0 else "", "Competitor seller: "+seller])
        url=IMG.get(cid,"")
        if url:
            p=fetch_png(cid,url)
            if p:
                img=XLImage(p); img.width=64; img.height=64
                ws.add_image(img, "A%d"%r)
                ws.row_dimensions[r].height=52

widths=[11,20,15,18,40,11,15,14,13,20,28,32,42,28]
for i,w in enumerate(widths,1): ws.column_dimensions[u.get_column_letter(i)].width=w
for row in ws.iter_rows(min_row=2):
    for c in row: c.alignment=Alignment(wrap_text=True,vertical="top")
ws.freeze_panes="A2"

OUTDIR=r"C:\Users\digit\OneDrive\Desktop\Abiraj_AIOS\.claude\worktrees\gifted-keller-42ebdd\development\abiraj_mini_aios_workbench\projects\PRJ-2026-017_ebay-competitor-keyword-research\evidence\final_outputs\REQ-20_ebay-competitor-keyword-research"
for name in ["REQ-20-D01_ebay_competitor_keyword.xlsx","REQ-20-D01_ebay_competitor_keyword_SOLD.xlsx","REQ-20-D01_ebay_competitor_keyword_FINAL.xlsx"]:
    try:
        wb.save(os.path.join(OUTDIR,name)); print("saved",name,"·",r-1,"rows"); break
    except PermissionError:
        print("locked:",name)

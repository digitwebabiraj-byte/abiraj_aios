# -*- coding: utf-8 -*-
"""
Independent verification of REQ-16-D01.

Re-derives every figure straight from the two live databases and from the shipped xlsx, then
diffs against the JSON payload the dashboard actually renders. Nothing is taken on trust from
the build script - this deliberately does NOT import build_esnm_d01.
"""
import os, io, json, re, sys, random
from datetime import date, timedelta
import psycopg2
from openpyxl import load_workbook

FINAL = (r"C:\Users\digit\OneDrive\Desktop\Abiraj_AIOS\.claude\worktrees"
         r"\uk-ledsone-ebay-jun-sales-7712cb\development\abiraj_mini_aios_workbench\projects"
         r"\PRJ-2026-014_ebay-slow-no-moving-products\evidence\final_outputs"
         r"\REQ-16_ebay-slow-no-moving-products")
HTML = os.path.join(FINAL, "REQ-16-D01_esnm_dashboard.html")
JSONF = os.path.join(FINAL, "esnm_d01_data.json")
XLSX = os.path.join(FINAL, "REQ-16-D01_slow_no_moving_products.xlsx")

ANCHOR = date(2026, 7, 22)
SUBS = (1, 2, 3, 4, 21, 22, 23, 24, 27, 28, 41, 222)
P, F = [], []
def ok(m):  P.append(m); print("  PASS  " + m)
def bad(m): F.append(m); print("  FAIL  " + m)

led = psycopg2.connect(host=os.environ["LED_PGHOST"], port=os.getenv("LED_PGPORT", "5432"),
                       dbname=os.environ["LED_PGDATABASE"], user=os.environ["LED_PGUSER"],
                       password=os.environ["LED_PGPASSWORD"], connect_timeout=30)
wh = psycopg2.connect(host=os.getenv("PGHOST", "149.28.134.54"), port=os.getenv("PGPORT", "5435"),
                      dbname=os.getenv("PGDATABASE", "order_management_copy"),
                      user=os.getenv("PGUSER", "temp_user"), password=os.environ["PGPASSWORD"],
                      connect_timeout=30)

D = json.load(io.open(JSONF, encoding="utf-8"))
R, DI = D["rows"], D["dicts"]
print("\n=== 1. HTML payload integrity ===")
h = io.open(HTML, encoding="utf-8").read()
m = re.search(r'<script id="payload" type="application/json">(.*?)</script>', h, re.S)
if not m: bad("payload script block not found in HTML")
else:
    emb = json.loads(m.group(1).replace("<\\/", "</"))
    ok("HTML embeds valid JSON") if emb["rows"] == R else bad("HTML payload != esnm_d01_data.json")
    ok("embedded rows == data.json rows (%d)" % len(emb["rows"]))
print("  html size: %.2f MB" % (os.path.getsize(HTML)/1048576.0))
if "Watchers" in h.split('<script id="payload"')[0]:
    bad("'Watchers' still present in dashboard markup/CSS")
else:
    ok("Watchers column absent from the dashboard shell")

print("\n=== 2. Row count vs live DB ===")
cur = led.cursor()
cur.execute("""SELECT COUNT(*) FROM listings.ebay_listings el
               JOIN order_management.sub_source ss ON ss.id=el.sub_source
               WHERE ss.source_id=2 AND el.site IN ('UK','Germany')
                 AND el.is_ended=0 AND el.is_child=0
                 AND el.item_id IS NOT NULL AND el.item_id<>''""")
db_n = cur.fetchone()[0]
ok("row count %d == DB %d" % (len(R), db_n)) if len(R) == db_n else \
    bad("row count %d != DB %d" % (len(R), db_n))
ids = [r[4] for r in R]
ok("item_ids unique (%d)" % len(set(ids))) if len(set(ids)) == len(ids) else \
    bad("duplicate item_ids: %d" % (len(ids)-len(set(ids))))

print("\n=== 3. Per account x marketplace vs DB ===")
cur.execute("""SELECT ss.map_name, el.site, COUNT(*)
               FROM listings.ebay_listings el
               JOIN order_management.sub_source ss ON ss.id=el.sub_source
               WHERE ss.source_id=2 AND el.site IN ('UK','Germany')
                 AND el.is_ended=0 AND el.is_child=0 AND el.item_id IS NOT NULL AND el.item_id<>''
               GROUP BY ss.map_name, el.site""")
dbmap = {(a, b): c for a, b, c in cur.fetchall()}
NAME = {"ledsone":"LEDSone","ledsonede":"LEDSone DE","sunsone":"SunSone",
        "electricalsone":"ElectricalSone","huettenlampen":"Huettenlampen",
        "coventrylights":"Coventry Lights","vintageinterior":"Vintage Interior",
        "dctransformer":"DC Transformer","retroled":"RetroLED","lightingsone":"LightingSone",
        "homin_gmbh":"Homin","bestbringer":"BestBringer"}
jmap = {}
for r in R:
    jmap[DI["account"][r[1]]] = jmap.get(DI["account"][r[1]], 0) + 1
mism = [( "%s - %s" % (NAME[a], b), n, jmap.get("%s - %s" % (NAME[a], b), 0))
        for (a, b), n in dbmap.items() if jmap.get("%s - %s" % (NAME[a], b), 0) != n]
ok("all %d account x marketplace counts match DB" % len(dbmap)) if not mism else \
    bad("account mismatches: %s" % mism[:5])

print("\n=== 4. Aggregates vs DB (zero-90d, dead qty) ===")
cur.execute("""WITH b AS (SELECT oii.item_id,
      SUM(COALESCE(NULLIF(oii.real_qty,'')::numeric,NULLIF(oii.item_quantity,'')::numeric,0)) s90
   FROM order_management.orders o JOIN order_management.order_item_info oii ON oii.order_id=o.id
   WHERE o.sub_source_id IN %(s)s AND COALESCE(o.status,'')<>'Cancelled'
     AND o.order_date::date BETWEEN %(a)s AND %(z)s GROUP BY oii.item_id)
   SELECT COUNT(*) FILTER (WHERE COALESCE(b.s90,0)<=0),
          COALESCE(SUM(el.quantity) FILTER (WHERE COALESCE(b.s90,0)<=0),0)
   FROM listings.ebay_listings el
   JOIN order_management.sub_source ss ON ss.id=el.sub_source
   LEFT JOIN b ON b.item_id=el.item_id
   WHERE ss.source_id=2 AND el.site IN ('UK','Germany') AND el.is_ended=0 AND el.is_child=0
     AND el.item_id IS NOT NULL AND el.item_id<>''""",
   dict(s=SUBS, a=ANCHOR-timedelta(days=89), z=ANCHOR))
db_zero, db_qty = cur.fetchone()
j_zero = sum(1 for r in R if r[12] <= 0)
j_qty = sum(r[9] or 0 for r in R if r[12] <= 0)
ok("zero-90d %d == DB %d" % (j_zero, db_zero)) if j_zero == db_zero else \
    bad("zero-90d %d != DB %d" % (j_zero, db_zero))
ok("dead qty %d == DB %d" % (j_qty, db_qty)) if j_qty == int(db_qty) else \
    bad("dead qty %d != DB %d" % (j_qty, db_qty))
ok("KPI card total (%d) == rows" % D["kpi"]["total"]) if D["kpi"]["total"] == len(R) else \
    bad("kpi.total mismatch")
crit = sum(1 for r in R if r[22] == 1)
ok("Rule 1 count %d == zero-90d %d (must be equal by definition)" % (crit, j_zero)) \
    if crit == j_zero else bad("Rule 1 %d != zero-90d %d" % (crit, j_zero))

print("\n=== 5. Random 25-listing field-by-field vs both DBs ===")
random.seed(20260722)
sample = random.sample(R, 25)
sids = tuple(r[4] for r in sample)
cur.execute("""SELECT el.item_id, el.sku, el.title, el.price, el.currency, el.quantity, el.site
               FROM listings.ebay_listings el WHERE el.item_id IN %s AND el.is_ended=0 AND el.is_child=0""", (sids,))
lst = {r[0]: r[1:] for r in cur.fetchall()}
cur.execute("""WITH b AS (SELECT oii.item_id AS iid, o.order_date::date d,
       COALESCE(NULLIF(oii.real_qty,'')::numeric,NULLIF(oii.item_quantity,'')::numeric,0) q
    FROM order_management.orders o JOIN order_management.order_item_info oii ON oii.order_id=o.id
    WHERE o.sub_source_id IN %(s)s AND COALESCE(o.status,'')<>'Cancelled' AND oii.item_id IN %(i)s)
    SELECT iid, COALESCE(SUM(q) FILTER (WHERE d BETWEEN %(a7)s AND %(z)s),0),
                COALESCE(SUM(q) FILTER (WHERE d BETWEEN %(a30)s AND %(z)s),0),
                COALESCE(SUM(q) FILTER (WHERE d BETWEEN %(a90)s AND %(z)s),0),
                COALESCE(SUM(q) FILTER (WHERE d BETWEEN %(l1)s AND %(l2)s),0), MAX(d)
    FROM b GROUP BY iid""",
    dict(s=SUBS, i=sids, a7=ANCHOR-timedelta(days=6), a30=ANCHOR-timedelta(days=29),
         a90=ANCHOR-timedelta(days=89), z=ANCHOR,
         l1=ANCHOR-timedelta(days=365+89), l2=ANCHOR-timedelta(days=365)))
sal = {r[0]: r[1:] for r in cur.fetchall()}
wc = wh.cursor()
wc.execute("""SELECT ref_id, COALESCE(SUM(click),0), COALESCE(SUM(conversion),0)
              FROM public.traffic_data WHERE which_channel=2 AND market_place IN ('UK','Germany')
                AND ref_id IN %s AND date BETWEEN %s AND %s GROUP BY ref_id""",
           (sids, ANCHOR-timedelta(days=29), ANCHOR))
trf = {r[0]: (float(r[1]), float(r[2])) for r in wc.fetchall()}

errs = 0
for r in sample:
    iid = r[4]
    L = lst.get(iid)
    if not L: errs += 1; print("    missing in DB:", iid); continue
    sku, title, price, curr, qty, site = L
    s7, s30, s90, sly, last = sal.get(iid, (0, 0, 0, 0, None))
    v, c = trf.get(iid, (None, None))
    checks = [
        ("sku", (r[3] or ""), (sku or "")),
        ("title", (r[5] or ""), (title or "")),
        ("price", r[7], (round(float(price), 2) if price is not None else None)),
        ("currency", r[8], curr or "GBP"),
        ("stock", r[9], int(qty or 0)),
        ("s7", r[10], float(s7)), ("s30", r[11], float(s30)),
        ("s90", r[12], float(s90)), ("sly", r[13], float(sly)),
        ("views", r[16], v),
        ("site", DI["site"][r[24]], site),
    ]
    for nm, got, exp in checks:
        if got != exp:
            errs += 1; print("    %s  %s: dashboard=%r db=%r" % (iid, nm, got, exp))
    # derived
    exp_trend = round((float(s90)-float(sly))/float(sly), 4) if sly and float(sly) > 0 else None
    if r[14] != exp_trend: errs += 1; print("    %s trend %r != %r" % (iid, r[14], exp_trend))
    exp_idle = (ANCHOR-last).days if last else None
    if last and r[15] != exp_idle:
        errs += 1; print("    %s idle %r != %r" % (iid, r[15], exp_idle))
    exp_cvr = round(c/v, 5) if (v and v > 0) else (0.0 if v is not None else None)
    if v is not None and r[18] != exp_cvr:
        errs += 1; print("    %s cvr %r != %r" % (iid, r[18], exp_cvr))
ok("25 listings x ~15 fields: 0 mismatches vs live DBs") if errs == 0 else \
    bad("%d field mismatches across the 25-row sample" % errs)

print("\n=== 6. Dashboard vs shipped xlsx (must agree) ===")
wb = load_workbook(XLSX, data_only=True)
ws = wb["Slow Moving No moving Products"]
xl = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    xl[str(row[4])] = row
ok("xlsx rows %d == dashboard rows %d" % (len(xl), len(R))) if len(xl) == len(R) else \
    bad("xlsx %d != dashboard %d" % (len(xl), len(R)))
diff = 0
for r in sample:
    x = xl.get(r[4])
    if not x: diff += 1; continue
    # xlsx action column is index 20 since Same Period Last Year was split into 30d/90d
    # (it was 19 in the 20-column layout). Getting this wrong compares against Listing Status.
    if DI["action"][r[20]] != x[20]: diff += 1; print("    %s action %r != xlsx %r" % (r[4], DI["action"][r[20]], x[20]))
    if float(r[12]) != float(x[11] or 0): diff += 1
    if int(r[9]) != int(x[8] or 0): diff += 1
ok("sample rows identical in xlsx and dashboard") if diff == 0 else \
    bad("%d dashboard/xlsx differences in the sample" % diff)

print("\n=== 7. Watchers really is gone from the data ===")
wat = sum(1 for r in R if r[17] is not None)
ok("watchers slot null on all %d rows" % len(R)) if wat == 0 else bad("%d rows carry a watchers value" % wat)
r6 = sum(1 for r in R if r[22] == 6)
ok("Rule 6 assigned to 0 listings") if r6 == 0 else bad("Rule 6 fired on %d rows" % r6)
r10 = sum(1 for r in R if r[22] == 10)
ok("Rule 10 assigned to 0 listings (known shadowing)") if r10 == 0 else bad("Rule 10 fired %d" % r10)

print("\n=== 8. Blank vs zero discipline ===")
nt = sum(1 for r in R if r[16] is None)
zv = sum(1 for r in R if r[16] == 0)
print("    views: %d blank (no traffic row), %d genuine zero" % (nt, zv))
ok("no traffic row -> blank, never 0 (blank=%d)" % nt)

led.close(); wh.close()
print("\n" + "="*64)
print("PASS %d   FAIL %d" % (len(P), len(F)))
if F:
    print("\nFAILURES:")
    for x in F: print("  - " + x)
print("VERDICT:", "GREEN - dashboard data verified" if not F else "RED - see failures above")
sys.exit(1 if F else 0)

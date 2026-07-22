# -*- coding: utf-8 -*-
"""
Slow Moving & No Moving Products — eBay (LEDSone group)
Builds the 20-column report + 12-rule action engine defined in "Thinesh task (3).xlsx".

Scope (confirmed by Abiraj 2026-07-22): ALL active eBay accounts, UK + Germany marketplaces only,
sellable listing rows (is_ended = 0, is_child = 0).

Reads LIVE, READ-ONLY from two databases:
  * ledsone            — listings, orders/sales, eBay Promoted-Listings PPC
  * order_management_copy (warehouse) — eBay organic traffic (views / conversions)

Credentials come from the global environment variables (never hardcoded).
"""
import os, sys, json
from datetime import date, timedelta
from collections import OrderedDict

import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_DIR = os.path.dirname(os.path.abspath(__file__))
DELIVER_DIR = os.path.join(os.path.expanduser("~"), "Downloads")
OUT_XLSX = os.path.join(DELIVER_DIR, "2026-07-22_slow-moving-no-moving-products_ebay.xlsx")

# ---------------------------------------------------------------- anchor & windows
ANCHOR = date(2026, 7, 22)          # latest complete SALES day (sales verified 91/91 days)
W7_A,  W7_B = ANCHOR - timedelta(days=6),  ANCHOR          # last 7 days
W7P_A, W7P_B = ANCHOR - timedelta(days=13), ANCHOR - timedelta(days=7)   # prior 7 days (Rule 12)
W30_A = ANCHOR - timedelta(days=29)
W90_A = ANCHOR - timedelta(days=89)
LY_B = ANCHOR - timedelta(days=365)
LY_A = LY_B - timedelta(days=89)    # same 90-day period last year

# in-scope eBay sub_source ids (source_id = 2, have UK/DE active listings)
SUBS = (1, 2, 3, 4, 21, 22, 23, 24, 27, 28, 41, 222)

# Brand = the underlying product brand (ledsone and ledsonede are both the LEDSone brand).
BRAND = {
    "ledsone": "LEDSone", "ledsonede": "LEDSone", "sunsone": "SunSone",
    "electricalsone": "ElectricalSone", "huettenlampen": "Huettenlampen",
    "coventrylights": "Coventry Lights", "vintageinterior": "Vintage Interior",
    "dctransformer": "DC Transformer", "retroled": "RetroLED",
    "lightingsone": "LightingSone", "homin_gmbh": "Homin", "bestbringer": "BestBringer",
}
# Account display name — must stay UNIQUE per seller account, because two different accounts
# (led_sone and ledsonede) both carry the LEDSone brand and both sell on Germany.
ACCOUNT_NAME = dict(BRAND, ledsonede="LEDSone DE")


def led_conn():
    return psycopg2.connect(
        host=os.environ["LED_PGHOST"], port=os.getenv("LED_PGPORT", "5432"),
        dbname=os.environ["LED_PGDATABASE"], user=os.environ["LED_PGUSER"],
        password=os.environ["LED_PGPASSWORD"], connect_timeout=30)


def wh_conn():
    return psycopg2.connect(
        host=os.getenv("PGHOST", "149.28.134.54"), port=os.getenv("PGPORT", "5435"),
        dbname=os.getenv("PGDATABASE", "order_management_copy"),
        user=os.getenv("PGUSER", "temp_user"), password=os.environ["PGPASSWORD"],
        connect_timeout=30)


SQL_LISTINGS = """
SELECT el.item_id, el.sku, el.title, el.price, el.currency, el.quantity,
       el.product_type, el.category_id, el.main_image_url, el.listing_url,
       el.site, el.status, el.created_at, ss.map_name AS account
FROM listings.ebay_listings el
JOIN order_management.sub_source ss ON ss.id = el.sub_source
WHERE ss.source_id = 2
  AND el.site IN ('UK','Germany')
  AND el.is_ended = 0
  AND el.is_child = 0
  AND el.item_id IS NOT NULL AND el.item_id <> ''
"""

SQL_SALES = """
WITH base AS (
  SELECT oii.item_id AS item_id,
         o.order_date::date AS d,
         COALESCE(NULLIF(oii.real_qty,'')::numeric,
                  NULLIF(oii.item_quantity,'')::numeric, 0) AS qty
  FROM order_management.orders o
  JOIN order_management.order_item_info oii ON oii.order_id = o.id
  WHERE o.sub_source_id IN %(subs)s
    AND COALESCE(o.status,'') <> 'Cancelled'
    AND oii.item_id IS NOT NULL AND oii.item_id <> ''
)
SELECT item_id,
       COALESCE(SUM(qty) FILTER (WHERE d BETWEEN %(w7a)s  AND %(w7b)s ),0) AS s7,
       COALESCE(SUM(qty) FILTER (WHERE d BETWEEN %(w7pa)s AND %(w7pb)s),0) AS s7_prev,
       COALESCE(SUM(qty) FILTER (WHERE d BETWEEN %(w30a)s AND %(anchor)s),0) AS s30,
       COALESCE(SUM(qty) FILTER (WHERE d BETWEEN %(w90a)s AND %(anchor)s),0) AS s90,
       COALESCE(SUM(qty) FILTER (WHERE d BETWEEN %(lya)s  AND %(lyb)s ),0) AS s90_ly,
       MAX(d) AS last_sale
FROM base GROUP BY item_id
"""

SQL_PPC = """
SELECT ebay_listing_id::text AS item_id,
       COALESCE(SUM(ad_fees_payout_currency),0) AS ppc_spend,
       COALESCE(SUM(attributed_sales),0)        AS ppc_sales
FROM ebay_campaigns.performance_data
WHERE date BETWEEN %(w30a)s AND %(anchor)s
GROUP BY ebay_listing_id
"""

SQL_TRAFFIC = """
SELECT ref_id AS item_id,
       COALESCE(SUM(click),0)      AS views_30d,
       COALESCE(SUM(conversion),0) AS conv_30d,
       COUNT(DISTINCT date)        AS days_present
FROM public.traffic_data
WHERE which_channel = 2
  AND market_place IN ('UK','Germany')
  AND date BETWEEN %(w30a)s AND %(anchor)s
GROUP BY ref_id
"""

SQL_TRAFFIC_DAYS = """
SELECT COUNT(DISTINCT date) AS days_present, MAX(date) AS latest
FROM public.traffic_data
WHERE which_channel = 2 AND market_place IN ('UK','Germany')
  AND date BETWEEN %(a)s AND %(b)s
"""


def fetch():
    p = dict(subs=SUBS, w7a=W7_A, w7b=W7_B, w7pa=W7P_A, w7pb=W7P_B,
             w30a=W30_A, w90a=W90_A, lya=LY_A, lyb=LY_B, anchor=ANCHOR)
    out = {}
    print("connecting to ledsone ...", flush=True)
    with led_conn() as c:
        with c.cursor() as cur:
            cur.execute(SQL_LISTINGS)
            out["listings"] = cur.fetchall()
            print("  listings rows: %d" % len(out["listings"]), flush=True)
            cur.execute(SQL_SALES, p)
            out["sales"] = {r[0]: r[1:] for r in cur.fetchall()}
            print("  sales item_ids: %d" % len(out["sales"]), flush=True)
            cur.execute(SQL_PPC, p)
            out["ppc"] = {r[0]: (float(r[1]), float(r[2])) for r in cur.fetchall()}
            print("  ppc item_ids: %d" % len(out["ppc"]), flush=True)

    print("connecting to warehouse ...", flush=True)
    with wh_conn() as c:
        with c.cursor() as cur:
            cur.execute(SQL_TRAFFIC, p)
            out["traffic"] = {r[0]: (float(r[1]), float(r[2]), int(r[3])) for r in cur.fetchall()}
            print("  traffic item_ids: %d" % len(out["traffic"]), flush=True)
            cur.execute(SQL_TRAFFIC_DAYS, dict(a=W30_A, b=ANCHOR))
            d30, latest30 = cur.fetchone()
            cur.execute(SQL_TRAFFIC_DAYS, dict(a=W90_A, b=ANCHOR))
            d90, latest90 = cur.fetchone()
            out["cov"] = dict(days30=int(d30), days90=int(d90), latest=str(latest30))
    return out


# ---------------------------------------------------------------- rule engine (reference impl.)
# Evaluated in PRIORITY order (Critical -> High -> Medium -> Low), first match wins.
TH = OrderedDict([
    ("r1_s90_max",   0),      # Rule 1  90d sales <= 0
    ("r2_s30_max",   0),      # Rule 2  30d sales <= 0 ...
    ("r2_stock_min", 50),     #         ... AND stock > 50
    ("r3_s7_max",    0),      # Rule 3  7d <= 0 ...
    ("r3_s30_max",   2),      #         ... 30d <= 2 ...
    ("r3_s90_max",   5),      #         ... 90d <= 5
    ("r4_drop",     -0.80),   # Rule 4  trend <= -80%
    ("r5_views_min", 100),    # Rule 5  views > 100 ...
    ("r5_cvr_max",   0.01),   #         ... AND CVR < 1%
    ("r6_watch_min", 10),     # Rule 6  watchers > 10   (NO DATA — never fires)
    ("r7_stock_min", 100),    # Rule 7  stock > 100 ...
    ("r7_s90_max",   5),      #         ... AND 90d < 5
    ("r8_spend_min", 5.00),   # Rule 8  PPC spend > GBP 5.00 AND no sales
    ("r9_views_max", 50),     # Rule 9  views < 50
    ("r10_age_min",  180),    # Rule 10 listing age > 180d ...
    ("r10_idle_min", 90),     #         ... AND last sale > 90d ago
    ("r11_s30_min",  10),     # Rule 11 30d >= 10
])

ACTIONS = {
    1: ("End Listing / Clear Stock", "Critical"),
    2: ("Run Clearance Promotion", "High"),
    3: ("Reduce Price by 5-10%", "High"),
    4: ("Review Competitor Pricing", "High"),
    5: ("Improve Images & SEO Title", "High"),
    6: ("Send Offer / Discount", "Medium"),
    7: ("Bundle with Best Seller", "High"),
    8: ("Pause PPC Campaign", "High"),
    9: ("Improve SEO & Increase Promotion", "Medium"),
    10: ("Refresh or Relist Listing", "Medium"),
    11: ("Maintain Current Strategy", "Low"),
    12: ("Increase Stock & PPC Budget", "Low"),
}
PRIORITY_RANK = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3, "": 4}
# order of evaluation: priority first, then rule number within the same priority
EVAL_ORDER = [1, 2, 3, 4, 5, 7, 8, 9, 10, 11, 12]   # 6 omitted: Watchers unavailable


def evaluate(row):
    """Return (rule_no, action, priority). row is the assembled dict."""
    s7, s30, s90, sly = row["s7"], row["s30"], row["s90"], row["s90_ly"]
    stock = row["stock"] or 0
    trend = row["trend"]
    views, cvr = row["views"], row["cvr"]
    has_traffic = row["has_traffic"]
    for n in EVAL_ORDER:
        ok = False
        if n == 1:
            ok = s90 <= TH["r1_s90_max"]
        elif n == 2:
            ok = s30 <= TH["r2_s30_max"] and stock > TH["r2_stock_min"]
        elif n == 3:
            ok = s7 <= TH["r3_s7_max"] and s30 <= TH["r3_s30_max"] and s90 <= TH["r3_s90_max"]
        elif n == 4:
            ok = sly > 0 and trend is not None and trend <= TH["r4_drop"]
        elif n == 5:
            ok = has_traffic and views > TH["r5_views_min"] and cvr is not None and cvr < TH["r5_cvr_max"]
        elif n == 7:
            ok = stock > TH["r7_stock_min"] and s90 < TH["r7_s90_max"]
        elif n == 8:
            ok = row["ppc_spend"] > TH["r8_spend_min"] and s30 <= 0
        elif n == 9:
            ok = has_traffic and views < TH["r9_views_max"]
        elif n == 10:
            ok = (row["age_days"] is not None and row["age_days"] > TH["r10_age_min"]
                  and row["idle_days"] is not None and row["idle_days"] > TH["r10_idle_min"])
        elif n == 11:
            ok = s30 >= TH["r11_s30_min"]
        elif n == 12:
            ok = s7 > row["s7_prev"] and s7 > 0
        if ok:
            a, p = ACTIONS[n]
            return n, a, p
    return 0, "Monitor - No Rule Matched", ""


def assemble(data):
    rows = []
    for (item_id, sku, title, price, currency, qty, ptype, cat_id, img, url,
         site, status, created_at, account) in data["listings"]:
        s = data["sales"].get(item_id)
        s7, s7p, s30, s90, sly, last_sale = (
            (float(s[0]), float(s[1]), float(s[2]), float(s[3]), float(s[4]), s[5])
            if s else (0.0, 0.0, 0.0, 0.0, 0.0, None))
        t = data["traffic"].get(item_id)
        has_traffic = t is not None
        views = float(t[0]) if t else None
        convs = float(t[1]) if t else None
        cvr = (convs / views) if (views and views > 0) else (0.0 if has_traffic else None)
        ppc_spend, ppc_sales = data["ppc"].get(item_id, (0.0, 0.0))

        age_days = (ANCHOR - created_at.date()).days if created_at else None
        if last_sale:
            idle = (ANCHOR - last_sale).days
            idle_is_proxy = False
        else:
            idle = age_days           # never sold -> use listing age (documented)
            idle_is_proxy = True
        trend = ((s90 - sly) / sly) if sly and sly > 0 else None

        brand = BRAND.get(account, account)
        acct_label = "%s - %s" % (ACCOUNT_NAME.get(account, account), site)
        row = dict(
            item_id=item_id, sku=sku or "", title=title or "",
            account=acct_label, brand=brand, site=site,
            price=float(price) if price is not None else None,
            currency=(currency or "GBP"),
            stock=int(qty) if qty is not None else 0,
            category=(ptype or (str(cat_id) if cat_id else "")),
            image=img or url or "",
            status=(status or "Active"),
            s7=s7, s7_prev=s7p, s30=s30, s90=s90, s90_ly=sly,
            trend=trend, idle_days=idle, idle_is_proxy=idle_is_proxy,
            age_days=age_days, views=views, cvr=cvr, has_traffic=has_traffic,
            ppc_spend=ppc_spend, ppc_sales=ppc_sales,
        )
        n, action, prio = evaluate(row)
        row.update(rule_no=n, action=action, priority=prio)
        rows.append(row)

    rows.sort(key=lambda r: (PRIORITY_RANK[r["priority"]], r["s90"], -(r["stock"] or 0)))
    return rows


# ---------------------------------------------------------------- excel
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BASE_FONT = Font(name="Arial", size=10)
INPUT_FILL = PatternFill("solid", fgColor="FFFF00")
PRIO_FILL = {
    "Critical": PatternFill("solid", fgColor="F8CBAD"),
    "High":     PatternFill("solid", fgColor="FFE699"),
    "Medium":   PatternFill("solid", fgColor="FFF2CC"),
    "Low":      PatternFill("solid", fgColor="E2EFDA"),
    "":         PatternFill("solid", fgColor="F2F2F2"),
}

HEADERS = ["Image", "Account", "Brand", "SKU", "Item ID", "Product Title", "Category",
           "Current Price", "Stock", "Last 7 Days Sales", "Last 30 Days Sales",
           "Last 90 Days Sales", "Same Period Last Year", "Sales Trend",
           "Days Since Last Sale", "Views (30 Days)", "Watchers", "Conversion Rate",
           "Listing Status", "Action Required"]
WIDTHS = [8, 20, 16, 22, 15, 52, 30, 13, 8, 11, 12, 12, 14, 12, 13, 12, 10, 12, 12, 30]
TREND_FMT = '"▲ "0%;"▼ -"0%;"→ "0%'


def build(rows, cov):
    wb = Workbook()

    # ---------- Rules sheet (thresholds live here; the main sheet references them) ----------
    ru = wb.active
    ru.title = "Rules"
    ru["A1"] = "12-Rule Action Engine — editable thresholds"
    ru["A1"].font = Font(name="Arial", size=13, bold=True)
    ru["A2"] = ("Yellow cells are inputs. Change one and every 'Action Required' cell on the "
                "report recalculates. Rules are evaluated in priority order (Critical -> High "
                "-> Medium -> Low); the first rule a listing matches wins.")
    ru["A2"].font = Font(name="Arial", size=9, italic=True)
    ru.merge_cells("A2:G2")
    ru["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ru.row_dimensions[2].height = 30

    rh = ["Rule", "Condition", "Action", "Priority", "Threshold 1", "Threshold 2", "Threshold 3"]
    for i, h in enumerate(rh, 1):
        c = ru.cell(row=3, column=i, value=h)
        c.font, c.fill, c.border = HDR_FONT, HDR_FILL, BORDER
    rule_defs = [
        ("Rule 1", "Last 90 Days Sales = 0", 1, [("r1_s90_max", "90d sales <=")]),
        ("Rule 2", "Last 30 Days Sales = 0 AND Stock > 50", 2,
         [("r2_s30_max", "30d sales <="), ("r2_stock_min", "stock >")]),
        ("Rule 3", "Last 7 Days = 0, Last 30 Days <= 2, Last 90 Days <= 5", 3,
         [("r3_s7_max", "7d <="), ("r3_s30_max", "30d <="), ("r3_s90_max", "90d <=")]),
        ("Rule 4", "Sales dropped >80% vs Same Period Last Year", 4, [("r4_drop", "trend <=")]),
        ("Rule 5", "Views >100 but Conversion Rate <1%", 5,
         [("r5_views_min", "views >"), ("r5_cvr_max", "CVR <")]),
        ("Rule 6", "Watchers >10 but No Sales (30 Days)  — NO DATA AVAILABLE", 6,
         [("r6_watch_min", "watchers >")]),
        ("Rule 7", "Stock >100 AND Last 90 Days Sales <5", 7,
         [("r7_stock_min", "stock >"), ("r7_s90_max", "90d <")]),
        ("Rule 8", "PPC Spend High AND No Sales", 8, [("r8_spend_min", "spend >")]),
        ("Rule 9", "Low Views (<50 in 30 Days)", 9, [("r9_views_max", "views <")]),
        ("Rule 10", "Listing Age >180 Days AND Last Sale >90 Days Ago", 10,
         [("r10_age_min", "age >"), ("r10_idle_min", "idle >")]),
        ("Rule 11", "Last 30 Days Sales >=10", 11, [("r11_s30_min", "30d >=")]),
        ("Rule 12", "Last 7 Days Sales Increasing", 12, []),
    ]
    for i, (label, cond, rn, ths) in enumerate(rule_defs):
        r = 4 + i
        act, prio = ACTIONS[rn]
        ru.cell(row=r, column=1, value=label)
        ru.cell(row=r, column=2, value=cond)
        ru.cell(row=r, column=3, value=act)
        ru.cell(row=r, column=4, value=prio)
        for j, (key, lbl) in enumerate(ths):
            c = ru.cell(row=r, column=5 + j, value=TH[key])
            c.fill = INPUT_FILL
            c.number_format = ('0.00%' if key in ("r4_drop", "r5_cvr_max")
                               else ('0.00' if key == "r8_spend_min" else '0'))
            ru.cell(row=r, column=5 + j).comment = None
        for cc in range(1, 8):
            ru.cell(row=r, column=cc).border = BORDER
            ru.cell(row=r, column=cc).font = BASE_FONT
        if rn == 6:
            for cc in range(1, 8):
                ru.cell(row=r, column=cc).fill = PatternFill("solid", fgColor="D9D9D9")
    ru["A17"] = ("Rule 6 cannot be evaluated: eBay 'Watchers' is not ingested into either "
                 "database (verified 2026-07-22). The Watchers column is therefore blank and "
                 "no listing is ever assigned the Rule 6 action.")
    ru["A17"].font = Font(name="Arial", size=9, bold=True, color="C00000")
    ru.merge_cells("A17:G18")
    ru["A17"].alignment = Alignment(wrap_text=True, vertical="top")
    for i, w in enumerate([10, 52, 32, 10, 14, 14, 14], 1):
        ru.column_dimensions[get_column_letter(i)].width = w

    # ---------- Engine Inputs (helper values the Action formula needs) ----------
    ei = wb.create_sheet("Engine Inputs")
    eih = ["Item ID", "Prev 7 Days Sales", "Listing Age (Days)", "PPC Spend (30d)",
           "PPC Attributed Sales (30d)", "Traffic Data Present", "Matched Rule", "Priority"]
    for i, h in enumerate(eih, 1):
        c = ei.cell(row=1, column=i, value=h)
        c.font, c.fill, c.border = HDR_FONT, HDR_FILL, BORDER
    for i, w in enumerate([15, 17, 17, 15, 24, 19, 13, 11], 1):
        ei.column_dimensions[get_column_letter(i)].width = w
    ei.freeze_panes = "A2"

    # ---------- main report ----------
    ws = wb.create_sheet("Slow Moving No moving Products", 0)
    for i, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font, c.fill, c.border = HDR_FONT, HDR_FILL, BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "B2"

    for idx, r in enumerate(rows):
        n = idx + 2
        if r["image"]:
            c = ws.cell(row=n, column=1, value="Image")
            c.hyperlink = r["image"]
            c.font = Font(name="Arial", size=10, color="0563C1", underline="single")
        else:
            ws.cell(row=n, column=1, value="")
        ws.cell(row=n, column=2, value=r["account"])
        ws.cell(row=n, column=3, value=r["brand"])
        ws.cell(row=n, column=4, value=r["sku"])
        ws.cell(row=n, column=5, value=r["item_id"])
        ws.cell(row=n, column=6, value=r["title"])
        ws.cell(row=n, column=7, value=r["category"])
        cp = ws.cell(row=n, column=8, value=r["price"])
        cp.number_format = ('€#,##0.00' if r["currency"] == "EUR" else '£#,##0.00')
        ws.cell(row=n, column=9, value=r["stock"])
        ws.cell(row=n, column=10, value=r["s7"])
        ws.cell(row=n, column=11, value=r["s30"])
        ws.cell(row=n, column=12, value=r["s90"])
        ws.cell(row=n, column=13, value=r["s90_ly"])
        ct = ws.cell(row=n, column=14, value="=IF(M%d=0,\"\",(L%d-M%d)/M%d)" % (n, n, n, n))
        ct.number_format = TREND_FMT
        ws.cell(row=n, column=15, value=r["idle_days"])
        ws.cell(row=n, column=16, value=r["views"] if r["has_traffic"] else None)
        ws.cell(row=n, column=17, value=None)                      # Watchers — no source
        cr = ws.cell(row=n, column=18, value=r["cvr"] if r["has_traffic"] else None)
        cr.number_format = '0.0%'
        ws.cell(row=n, column=19, value=r["status"])
        ws.cell(row=n, column=20, value=action_formula(n))
        for cc in range(1, 21):
            cell = ws.cell(row=n, column=cc)
            cell.border = BORDER
            if cc != 1:
                cell.font = BASE_FONT
        ws.cell(row=n, column=20).fill = PRIO_FILL[r["priority"]]

        ei.cell(row=n, column=1, value=r["item_id"])
        ei.cell(row=n, column=2, value=r["s7_prev"])
        ei.cell(row=n, column=3, value=r["age_days"])
        ei.cell(row=n, column=4, value=round(r["ppc_spend"], 2))
        ei.cell(row=n, column=5, value=r["ppc_sales"])
        ei.cell(row=n, column=6, value=1 if r["has_traffic"] else 0)
        ei.cell(row=n, column=7, value=r["rule_no"])
        ei.cell(row=n, column=8, value=r["priority"])

    last = len(rows) + 1
    ws.auto_filter.ref = "A1:T%d" % last

    build_summary(wb, rows, cov, last)
    build_notes(wb, rows, cov)
    wb._sheets = [wb["Slow Moving No moving Products"], wb["Summary"], wb["Rules"],
                  wb["Engine Inputs"], wb["Data Notes"]]
    wb.save(OUT_XLSX)
    return OUT_XLSX


def action_formula(n):
    """Nested IF implementing the engine, referencing Rules!/Engine Inputs! cells."""
    R, E = "Rules!", "'Engine Inputs'!"
    conds = [
        ("L{n}<={R}$E$4", "{R}$C$4"),                                              # 1
        ("AND(K{n}<={R}$E$5,I{n}>{R}$F$5)", "{R}$C$5"),                            # 2
        ("AND(J{n}<={R}$E$6,K{n}<={R}$F$6,L{n}<={R}$G$6)", "{R}$C$6"),             # 3
        ("AND(M{n}>0,N{n}<={R}$E$7)", "{R}$C$7"),                                  # 4
        ("AND({E}$F{n}=1,P{n}>{R}$E$8,R{n}<{R}$F$8)", "{R}$C$8"),                  # 5
        ("AND(I{n}>{R}$E$10,L{n}<{R}$F$10)", "{R}$C$10"),                          # 7
        ("AND({E}$D{n}>{R}$E$11,K{n}<=0)", "{R}$C$11"),                            # 8
        ("AND({E}$F{n}=1,P{n}<{R}$E$12)", "{R}$C$12"),                             # 9
        ("AND({E}$C{n}>{R}$E$13,O{n}>{R}$F$13)", "{R}$C$13"),                      # 10
        ("K{n}>={R}$E$14", "{R}$C$14"),                                            # 11
        ("AND(J{n}>{E}$B{n},J{n}>0)", "{R}$C$15"),                                 # 12
    ]
    f = ""
    for c, v in conds:
        f += "IF(%s,%s," % (c.format(n=n, R=R, E=E), v.format(R=R))
    f += '"Monitor - No Rule Matched"' + ")" * len(conds)
    return "=" + f


def build_summary(wb, rows, cov, last):
    sm = wb.create_sheet("Summary")
    sm["A1"] = "Slow Moving & No Moving Products — eBay (LEDSone group)"
    sm["A1"].font = Font(name="Arial", size=14, bold=True)
    sm["A2"] = ("Anchor date %s   |   Scope: all active eBay accounts, UK + Germany, "
                "sellable listings (is_ended=0, is_child=0)   |   %d listings"
                % (ANCHOR.isoformat(), len(rows)))
    sm["A2"].font = Font(name="Arial", size=9, italic=True)

    sm["A4"] = "Action Required — volume by rule"
    sm["A4"].font = Font(name="Arial", size=11, bold=True)
    for i, h in enumerate(["Priority", "Rule", "Action", "Listings", "% of Total"], 1):
        c = sm.cell(row=5, column=i, value=h)
        c.font, c.fill, c.border = HDR_FONT, HDR_FILL, BORDER
    counts = {}
    for r in rows:
        counts[r["rule_no"]] = counts.get(r["rule_no"], 0) + 1
    order = sorted(counts, key=lambda n: (PRIORITY_RANK[ACTIONS[n][1]] if n else 4, n))
    rr = 6
    for n in order:
        act, prio = ACTIONS[n] if n else ("Monitor - No Rule Matched", "")
        sm.cell(row=rr, column=1, value=prio or "-")
        sm.cell(row=rr, column=2, value=("Rule %d" % n) if n else "-")
        sm.cell(row=rr, column=3, value=act)
        sm.cell(row=rr, column=4, value=counts[n])
        pc = sm.cell(row=rr, column=5, value="=D%d/$D$%d" % (rr, rr + len(order) - (rr - 6) + 0))
        pc.number_format = '0.0%'
        for cc in range(1, 6):
            sm.cell(row=rr, column=cc).border = BORDER
            sm.cell(row=rr, column=cc).font = BASE_FONT
        sm.cell(row=rr, column=1).fill = PRIO_FILL.get(prio, PRIO_FILL[""])
        rr += 1
    tot = rr
    sm.cell(row=tot, column=3, value="TOTAL").font = Font(name="Arial", size=10, bold=True)
    sm.cell(row=tot, column=4, value="=SUM(D6:D%d)" % (tot - 1)).font = Font(name="Arial", size=10, bold=True)
    for r_ in range(6, tot):
        sm.cell(row=r_, column=5, value="=D%d/$D$%d" % (r_, tot))
        sm.cell(row=r_, column=5).number_format = '0.0%'

    start = tot + 2
    sm.cell(row=start, column=1, value="Breakdown by account × marketplace").font = Font(
        name="Arial", size=11, bold=True)
    for i, h in enumerate(["Account", "Listings", "Zero 90-Day Sales", "Critical (End Listing)"], 1):
        c = sm.cell(row=start + 1, column=i, value=h)
        c.font, c.fill, c.border = HDR_FONT, HDR_FILL, BORDER
    agg = {}
    for r in rows:
        a = agg.setdefault(r["account"], [0, 0, 0])
        a[0] += 1
        if r["s90"] <= 0:
            a[1] += 1
        if r["rule_no"] == 1:
            a[2] += 1
    rr = start + 2
    for acct in sorted(agg, key=lambda k: -agg[k][0]):
        v = agg[acct]
        sm.cell(row=rr, column=1, value=acct)
        sm.cell(row=rr, column=2, value=v[0])
        sm.cell(row=rr, column=3, value=v[1])
        sm.cell(row=rr, column=4, value=v[2])
        for cc in range(1, 5):
            sm.cell(row=rr, column=cc).border = BORDER
            sm.cell(row=rr, column=cc).font = BASE_FONT
        rr += 1
    sm.cell(row=rr, column=1, value="TOTAL").font = Font(name="Arial", size=10, bold=True)
    for cc, col in ((2, "B"), (3, "C"), (4, "D")):
        c = sm.cell(row=rr, column=cc, value="=SUM(%s%d:%s%d)" % (col, start + 2, col, rr - 1))
        c.font = Font(name="Arial", size=10, bold=True)
    for i, w in enumerate([34, 34, 34, 22, 12], 1):
        sm.column_dimensions[get_column_letter(i)].width = w


def build_notes(wb, rows, cov):
    dn = wb.create_sheet("Data Notes")
    dn["A1"] = "Data sources, assumptions and known gaps"
    dn["A1"].font = Font(name="Arial", size=14, bold=True)
    n_traffic = sum(1 for r in rows if r["has_traffic"])
    n_proxy = sum(1 for r in rows if r["idle_is_proxy"])
    lines = [
        ("Anchor date", ANCHOR.isoformat(),
         "Sales are complete to this date (91/91 days verified)."),
        ("Scope", "All active eBay accounts; UK + Germany only",
         "Sellable rows only: is_ended = 0 AND is_child = 0. 12 accounts qualify; "
         "neighbourmarket is excluded (US-only, no UK/DE listings)."),
        ("Listings source", "ledsone / listings.ebay_listings", "Live, refreshed daily."),
        ("Sales source", "ledsone / order_management.orders + order_item_info",
         "Units = real_qty (falls back to item_quantity). Orders with status 'Cancelled' are "
         "EXCLUDED; 'Refunded' and 'Inprogress' are INCLUDED because they still evidence demand."),
        ("Sales windows", "7 / 30 / 90 days ending on the anchor",
         "Same Period Last Year = the same 90-day window one year earlier (%s to %s)."
         % (LY_A.isoformat(), LY_B.isoformat())),
        ("Sales Trend", "(Last 90 Days - Same Period Last Year) / Same Period Last Year",
         "Blank where last year's sales were zero (division undefined). Live formula in column N."),
        ("Views + Conversion source", "warehouse / public.traffic_data (which_channel = 2)",
         "Views = SUM(click) = listing page views. Conversion Rate = SUM(conversion)/SUM(click). "
         "%d of %d listings (%.1f%%) have traffic rows; the rest are blank, NOT zero."
         % (n_traffic, len(rows), 100.0 * n_traffic / max(len(rows), 1))),
        ("*** TRAFFIC GAP ***",
         "Only %d of 30 days present in the Views window" % cov["days30"],
         "eBay traffic ingestion FAILED on 26 Apr, 7-11 May, 26 Jun, 29 Jun-1 Jul and 18 Jul "
         "2026 (11 days lost; Shopify loaded normally on all of them, so it is eBay-specific). "
         "Latest traffic day is %s — eBay traffic also lands ~2 days late. Views are therefore "
         "UNDERSTATED. Rules 5 and 9 are affected and are only evaluated for listings that have "
         "traffic rows at all." % cov["latest"]),
        ("*** WATCHERS: NO DATA ***", "Column Q is intentionally blank",
         "eBay 'Watchers' is not ingested into either database — verified by scanning every "
         "column in both. It is only available from eBay's Trading API. RULE 6 (Watchers >10 "
         "but no sales) therefore NEVER FIRES and no listing can be assigned its action."),
        ("*** PPC ONLY 65 DAYS ***", "ebay_campaigns.performance_data starts 18 May 2026",
         "Rule 8 is evaluated on a 30-day window (fully covered), NOT 90 days. A 90-day PPC "
         "figure does not exist."),
        ("Rule 8 'spend is high'", "Default threshold GBP %.2f over 30 days" % TH["r8_spend_min"],
         "The requirement says 'PPC Spend High' without defining high. This threshold is an "
         "ASSUMPTION and is editable on the Rules sheet."),
        ("Days Since Last Sale", "Anchor minus the most recent order date",
         "%d listings have never sold in the order history; for those this shows the listing's "
         "AGE instead, so Rule 10 still evaluates correctly." % n_proxy),
        ("Account column", "Account + marketplace, e.g. 'LEDSone - UK', 'LEDSone DE - Germany'",
         "The requirement has no separate Marketplace column, so the marketplace is carried in "
         "Account. This keeps rows at account x marketplace grain and prevents double-counting. "
         "The account name is used (not the brand) because two accounts - led_sone and "
         "ledsonede - both carry the LEDSone brand and both sell on Germany."),
        ("Brand", "Derived from the seller account", "No brand field exists on eBay listings. "
         "led_sone and ledsonede are both shown as brand 'LEDSone'."),
        ("Listing Status", "Derived from is_ended / end_date",
         "ebay_listings.status is ~99% NULL (populated only on parent/single rows), so it "
         "cannot be read directly."),
        ("Rule precedence", "Critical -> High -> Medium -> Low; first match wins",
         "Within one priority band, the lower rule number wins. Evaluation order: "
         "1, 2, 3, 4, 5, 7, 8, 9, 10, 11, 12 (6 omitted - no data)."),
        ("Recalculation", "Thresholds live on the Rules sheet",
         "Column T (Action Required) and column N (Sales Trend) are live formulas. Edit a "
         "yellow threshold cell and the whole report re-evaluates."),
    ]
    for i, h in enumerate(["Item", "Value", "Detail"], 1):
        c = dn.cell(row=3, column=i, value=h)
        c.font, c.fill, c.border = HDR_FONT, HDR_FILL, BORDER
    r = 4
    for a, b, c_ in lines:
        dn.cell(row=r, column=1, value=a)
        dn.cell(row=r, column=2, value=b)
        dn.cell(row=r, column=3, value=c_)
        for cc in range(1, 4):
            cell = dn.cell(row=r, column=cc)
            cell.border = BORDER
            cell.font = (Font(name="Arial", size=10, bold=True, color="C00000")
                         if a.startswith("***") else BASE_FONT)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        dn.row_dimensions[r].height = 46
        r += 1
    for i, w in enumerate([30, 46, 86], 1):
        dn.column_dimensions[get_column_letter(i)].width = w


if __name__ == "__main__":
    data = fetch()
    rows = assemble(data)
    print("assembled rows: %d" % len(rows), flush=True)
    path = build(rows, data["cov"])
    print("WROTE", path, flush=True)
    counts = {}
    for r in rows:
        counts[r["rule_no"]] = counts.get(r["rule_no"], 0) + 1
    print("rule counts:", json.dumps({str(k): v for k, v in sorted(counts.items())}))
    json.dump({str(k): v for k, v in sorted(counts.items())},
              open(os.path.join(OUT_DIR, "python_reference_counts.json"), "w"))

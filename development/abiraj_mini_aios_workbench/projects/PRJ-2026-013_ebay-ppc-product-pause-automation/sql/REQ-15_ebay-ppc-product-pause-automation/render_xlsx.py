"""REQ-15-D01 — workbook renderer. Reads eppa_d01_data.json, writes the 3-sheet xlsx."""
import json, os
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
FINAL = os.path.abspath(os.path.join(HERE, "..", "..", "evidence", "final_outputs",
                                     "REQ-15_ebay-ppc-product-pause-automation"))
D = json.load(open(os.path.join(FINAL, "eppa_d01_data.json"), encoding="utf-8"))
rows, K, TH, ANCHOR = D["rows"], D["kpis"], D["thresholds"], D["anchor"]

# Windows are the N days ENDING on the anchor, inclusive (SQL: `date > anchor - N`).
# Derived, never hardcoded — the same rule the dashboard uses, so the two cannot disagree.
_A = datetime.strptime(ANCHOR, "%Y-%m-%d").date()


def _fmt(n):
    st, en = _A - timedelta(days=n - 1), _A
    if (st.year, st.month) == (en.year, en.month):
        return "%d-%d %s %d" % (st.day, en.day, en.strftime("%b"), en.year)
    if st.year == en.year:
        return "%d %s - %d %s %d" % (st.day, st.strftime("%b"), en.day, en.strftime("%b"), en.year)
    return "%d %s %d - %d %s %d" % (st.day, st.strftime("%b"), st.year,
                                    en.day, en.strftime("%b"), en.year)


W30, W14, W7 = _fmt(30), _fmt(14), _fmt(7)

NAVY, GREY = "22406B", "6B7688"
RULE_COLOR = {"Stock": "C0392B", "Rule 1": "2E5C97", "Rule 2": "C77C1B"}
bd = Border(bottom=Side(style="thin", color="DBE1EA"))
money = '£#,##0.00'

wb = Workbook()

# ---- Dashboard ---------------------------------------------------------------------------
ws = wb.active
ws.title = "Dashboard"
ws["A1"] = "eBay PPC Product Pause Automation"
ws["A1"].font = Font(bold=True, size=15, color=NAVY)
ws["A2"] = "LEDSone · eBay UK · Promoted Listings (Advanced / ON_SITE) · recommendation only"
ws["A2"].font = Font(size=10, color=GREY)
ws["A3"] = ("30D %s  ·  14D %s  ·  7D %s   (each window ends on %s and includes it)"
            % (W30, W14, W7, _A.strftime("%d %b %Y")))
ws["A3"].font = Font(size=10, color=GREY)
ws["A4"] = "One row per campaign · Standard (COST_PER_SALE) campaigns excluded · recommendation only"
ws["A4"].font = Font(size=10, color=GREY)
ws["A6"], ws["B6"] = "Metric", "Value"
for c in ("A6", "B6"):
    ws[c].font = Font(bold=True, color="FFFFFF")
    ws[c].fill = PatternFill("solid", fgColor=NAVY)
for i, (k, v) in enumerate([
        ("Campaigns in scope", K["scope"]), ("Recommended pause", K["paused"]),
        ("  · High priority", K["high"]), ("  · Medium priority", K["med"]),
        ("  · Low priority", K["low"]), ("Stock rule hits", K["stock"]),
        ("Rule 1 hits (high ACOS)", K["r1"]), ("Rule 2 hits (clicks, no sales)", K["r2"]),
        ("Still running", K["running"]), ("Already off", K["off"]),
        ("30D spend at risk (paused)", K["spend_at_risk"]),
        ("30D spend — all campaigns", K["spend_all"]),
        ("Advertised listings", K["listings"]),
        ("  · out of stock", K["oos_listings"]),
        ("  · no stock record", K["nodata_listings"])], start=7):
    ws.cell(i, 1, k).border = bd
    c = ws.cell(i, 2, v)
    c.border = bd
    if isinstance(v, float):
        c.number_format = money
ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 16

# ---- Pause Log ---------------------------------------------------------------------------
ws2 = wb.create_sheet("Pause Log")
hdr = ["Campaign", "Campaign ID", "Type", "Campaign state", "Advertised listings", "Out of stock",
       "Low stock", "No stock data", "Rule", "Priority", "30D ACOS %", "7D ACOS %", "30D orders",
       "14D orders", "14D clicks", "14D spend", "30D spend", "Status", "Reason", "Decision", "Note"]
ws2.append(hdr)
for i in range(1, len(hdr) + 1):
    c = ws2.cell(1, i)
    c.font = Font(bold=True, color="FFFFFF", size=9)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(vertical="center", wrap_text=True)
for r in rows:
    ws2.append([r["campaign"], r["campaign_id"], r["type"], r["status"], r["listings"],
                r["out_of_stock"], r["low_stock"], r["no_stock_data"], r["rule"],
                r["priority"] or "—",
                round(r["acos30"], 1) if r["acos30"] is not None else None,
                round(r["acos7"], 1) if r["acos7"] is not None else None,
                r["ord30"], r["ord14"], r["clicks14"], r["spend14"], r["spend30"],
                r["outcome"], r["reason"], "Pending", ""])
    row = ws2.max_row
    # `outcome` = the engine's decision; `status` (col D) = the campaign's own live state.
    fill = {"PAUSED": "FBEAE8", "ALREADY OFF": "F2F3F5"}.get(r["outcome"])
    for i in range(1, len(hdr) + 1):
        c = ws2.cell(row, i)
        c.border = bd
        c.alignment = Alignment(vertical="top", wrap_text=(i == 19))
        if fill:
            c.fill = PatternFill("solid", fgColor=fill)
    if r["rule"] in RULE_COLOR:
        ws2.cell(row, 9).font = Font(bold=True, color=RULE_COLOR[r["rule"]])
    for i in (16, 17):
        ws2.cell(row, i).number_format = money
for i, w in enumerate([44, 15, 9, 15, 12, 12, 11, 13, 10, 10, 11, 11, 11, 11, 11, 11, 11,
                       13, 74, 11, 18], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A2"
ws2.auto_filter.ref = ws2.dimensions

# ---- Pause Rules (editable thresholds) ----------------------------------------------------
ws3 = wb.create_sheet("Pause Rules")
ws3.append(["Setting", "Value", "What it does"])
for i in range(1, 4):
    ws3.cell(1, i).font = Font(bold=True, color="FFFFFF")
    ws3.cell(1, i).fill = PatternFill("solid", fgColor=NAVY)
for k, v, d in [
        ("Stock floor (units)", TH["stock_floor"], "A listing below this counts as low stock"),
        ("Rule 1 — 30D ACOS ceiling (%)", TH["acos_ceiling"], "Pause at or above this"),
        ("Rule 1 — 7D ACOS rescue (%)", TH["acos_rescue"], "Skip the pause below this (improving trend)"),
        ("Rule 2 — 14D clicks minimum", TH["clicks_min"], "Rule 2 only applies at or above this"),
        ("Rule 2 — 14D spend floor (£)", TH["spend_floor"], "Skip the pause below this (cheap organic clicks)"),
        ("Priority — High at 30D spend (£)", TH["prio_high"], "Paused rows at or above this rank High"),
        ("Priority — Medium at 30D spend (£)", TH["prio_med"], "Paused rows at or above this rank Medium")]:
    ws3.append([k, v, d])
ws3.append([])
for k, v, d in [("Window — 30 days", W30, "Rule 1 ACOS, orders, spend at risk, priority"),
                ("Window — 14 days", W14, "Rule 2 orders, clicks, spend floor"),
                ("Window — 7 days", W7, "Rule 1 rescue (improving-trend check)")]:
    ws3.append([k, v, d])
ws3["A14"] = ("Thresholds are configuration — they live in eppa_engine.py and drive every "
              "calculation. Change them there, not in the SQL. Windows are derived from the "
              "anchor date (the latest date in the ad data), never hardcoded.")
ws3["A14"].font = Font(size=9, italic=True, color=GREY)
for i, w in enumerate([34, 12, 58], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

p = os.path.join(FINAL, "REQ-15-D01_eppa_pause_log.xlsx")
wb.save(p)
print("workbook  -> %s  (%.1f KB)" % (p, os.path.getsize(p) / 1024))

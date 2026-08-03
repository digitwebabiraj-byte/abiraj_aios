#!/usr/bin/env python3
"""
epns_build_d01.py — eBay Product Net Sales (NNV) report — REQ-22-D01
PRJ-2026-019 · code epns · requester Kobiga

ONE read-only module (the single fetch path). Connects direct to the RAW ledsone
PostgreSQL (source of record for eBay orders/fees), computes one row per eBay order
over the last 30 days, and writes the Excel deliverable with two tabs:
  Tab 1  Net Sales   — the 12-column per-order table
  Tab 2  Net Sales Lookup — enter an Order ID (col A) -> its Net Sales + breakdown (INDEX/MATCH)

DEFINITION (reconciled to eBay's own payout, 2026-08-03):
  Net Sales (NNV) = Gross Sales - Final Value Fee - PPC Cost - General fees
  where Gross Sales = orders.total (already net of promotion/discount).
  This ties, to the penny, to accounting.ebay_order_expenses SALE transaction_amount
  and to the source worked example order 02-14934-76138 -> 22.39.

  VAT (20%) is a DERIVED standard-rate ESTIMATE shown for context (not deducted from NNV,
  because output VAT is remitted to HMRC separately).
  Product Cost is NO DATA — no per-SKU COGS exists in any ledsone schema (swept 2026-08-03).
  A "true net profit" (NNV - VAT - Product Cost) is therefore NOT computed; it needs a COGS
  source from Kobiga. See PROJECT_HOME.md / SYSTEM_REFERENCE.md.

MONEY IS PER MARKETPLACE CURRENCY, NEVER BLENDED (UK GBP / DE+others EUR / US USD; no FX table).

Read-only. Credentials come from the git-ignored shared store (env vars), never committed.
Usage:  python epns_build_d01.py [output.xlsx]
"""
import os, sys, datetime as dt

SQL = r"""
WITH ebay_orders AS (
  SELECT o.id, o.order_id, o.order_date, o.total, o.discount, o.shipping_cost,
         o.market_place, ss.name AS account
  FROM order_management.orders o
  JOIN order_management.sub_source ss ON ss.id = o.sub_source_id AND ss.source_id = 2
  WHERE o.order_date >= CURRENT_DATE - INTERVAL '30 days'
    AND o.order_date <  CURRENT_DATE                 -- anchor on last COMPLETE day
    AND o.status IN ('Completed','New','Inprogress') -- sales; excludes Cancelled/Refunded
),
lines AS (
  SELECT oii.order_id AS oid,
         string_agg(DISTINCT COALESCE(NULLIF(oii.item_sku,''), oii.real_sku), ' | ') AS skus
  FROM order_management.order_item_info oii
  GROUP BY oii.order_id
),
fees AS (   -- eBay fee stack, bucketed by fee_type (join by eBay order_id)
  SELECT e.order_id AS oref,
         SUM(e.fee) FILTER (WHERE e.transaction_type = 'SALE')                       AS fvf,
         SUM(e.fee) FILTER (WHERE e.fee_type IN ('AD_FEE','PREMIUM_AD_FEES'))        AS ppc,
         SUM(e.fee) FILTER (WHERE e.fee_type IN ('INSERTION_FEE','OTHER_FEES',
              'SUBTITLE_FEE','INTERNATIONAL_LISTING_FEE','GALLERY_PLUS_FEE',
              'PAYMENT_DISPUTE_FEE'))                                                AS gen
  FROM accounting.ebay_order_expenses e
  WHERE e.order_id IS NOT NULL AND e.order_id <> ''
  GROUP BY e.order_id
)
SELECT eo.order_id,
       l.skus AS sku,
       eo.account,
       CASE eo.market_place WHEN '23' THEN 'UK' WHEN '10' THEN 'Germany' WHEN '24' THEN 'US'
            WHEN '9' THEN 'France' WHEN '13' THEN 'Ireland' WHEN '14' THEN 'Italy'
            ELSE eo.market_place END AS marketplace,
       CASE eo.market_place WHEN '23' THEN 'GBP' WHEN '24' THEN 'USD' ELSE 'EUR' END AS currency,
       to_char(eo.order_date,'YYYY-MM-DD')                       AS order_date,
       ROUND(eo.total,2)                                         AS gross_sales,
       ROUND(eo.total - eo.total/1.2, 2)                         AS vat_20,          -- ESTIMATE
       ROUND(COALESCE(eo.discount,0),2)                          AS promotion,
       ROUND(COALESCE(f.fvf,0),2)                                AS final_value_fee,
       ROUND(COALESCE(eo.shipping_cost,0),2)                     AS postage,
       ROUND(COALESCE(f.ppc,0),2)                                AS ppc_cost,
       ROUND(COALESCE(f.gen,0),2)                                AS general,
       ROUND(eo.total - COALESCE(f.fvf,0) - COALESCE(f.ppc,0) - COALESCE(f.gen,0), 2)
                                                                 AS net_sales_nnv
FROM ebay_orders eo
LEFT JOIN lines l ON l.oid  = eo.id
LEFT JOIN fees  f ON f.oref = eo.order_id
ORDER BY eo.order_date DESC, eo.order_id;
"""

def cred(k):
    v = os.environ.get(k)
    if not v and k != "LED_PGPORT":
        sys.exit(f"Missing credential env var {k} (source the git-ignored epns_secrets.bat)")
    return v

def fetch_rows():
    import psycopg2, psycopg2.extras
    c = psycopg2.connect(host=cred("LED_PGHOST"), port=cred("LED_PGPORT") or 5432,
                         dbname=cred("LED_PGDATABASE"), user=cred("LED_PGUSER"),
                         password=cred("LED_PGPASSWORD"), connect_timeout=30)
    try:
        cur = c.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(SQL)
        return [dict(r) for r in cur.fetchall()]
    finally:
        c.close()

# ---- shared writer (used by this module AND by the one-off JSON renderer) --------------
COLS = [
    ("Order ID", "order_id", None),
    ("SKU", "sku", None),
    ("Account", "account", None),
    ("Marketplace", "marketplace", None),
    ("Currency", "currency", None),
    ("Order Date", "order_date", None),
    ("Gross Sales", "gross_sales", "money"),
    ("VAT (20%) [est]", "vat_20", "money"),
    ("Promotion", "promotion", "money"),
    ("Final Value Fee", "final_value_fee", "money"),
    ("Product Cost", None, "nodata"),
    ("Postage", "postage", "money"),
    ("PPC Cost", "ppc_cost", "money"),
    ("General", "general", "money"),
    ("Net Sales (NNV)", "net_sales_nnv", "money"),
]

def write_workbook(rows, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    anchor = dt.date.today().isoformat()
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    note_font = Font(name="Arial", italic=True, color="7F7F7F", size=9)
    money_fmt = '#,##0.00;(#,##0.00);-'
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()
    ws = wb.active
    ws.title = "Net Sales"

    # Title + provenance banner
    ws["A1"] = "eBay Product Net Sales (NNV) — REQ-22-D01 (Kobiga)"
    ws["A1"].font = Font(name="Arial", bold=True, size=13)
    ws["A2"] = (f"One row per eBay order · last 30 days ending {anchor} (last complete day) · "
                f"{len(rows)} orders · source: raw ledsone (source_id=2), read-only.")
    ws["A2"].font = note_font
    ws["A3"] = ("Net Sales (NNV) = Gross Sales − Final Value Fee − PPC Cost − General "
                "(= eBay net payout; ties to eBay SALE transaction_amount). "
                "VAT (20%) is a derived ESTIMATE (not deducted from NNV). "
                "Product Cost = NO DATA (no COGS in any DB). Money per marketplace currency — NEVER blended.")
    ws["A3"].font = note_font

    hr = 5
    for ci, (label, _key, _kind) in enumerate(COLS, start=1):
        c = ws.cell(row=hr, column=ci, value=label)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    r = hr + 1
    for row in rows:
        for ci, (_label, key, kind) in enumerate(COLS, start=1):
            if kind == "nodata":
                cell = ws.cell(row=r, column=ci, value="NO DATA")
                cell.font = Font(name="Arial", italic=True, color="C00000", size=9)
            else:
                val = row.get(key)
                if kind == "money" and val is not None:
                    val = float(val)
                cell = ws.cell(row=r, column=ci, value=val)
                cell.font = Font(name="Arial", size=9)
                if kind == "money":
                    cell.number_format = money_fmt
            cell.border = border
        r += 1

    widths = [18, 30, 16, 12, 9, 12, 13, 14, 12, 15, 13, 11, 11, 11, 15]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A{hr}:{get_column_letter(len(COLS))}{r-1}"

    # ---- Tab 2: Net Sales Lookup (INDEX/MATCH into the Net Sales tab) ------------------
    lk = wb.create_sheet("Net Sales Lookup")
    lk["A1"] = "Net Sales Lookup — enter an Order ID in B3"
    lk["A1"].font = Font(name="Arial", bold=True, size=12)
    lk["A2"] = "Type any eBay Order ID from the 'Net Sales' tab; every field below fills automatically."
    lk["A2"].font = note_font
    lk["A3"] = "Order ID:"; lk["A3"].font = Font(name="Arial", bold=True)
    lk["B3"] = rows[0]["order_id"] if rows else ""     # example seed
    lk["B3"].fill = PatternFill("solid", fgColor="FFF2CC")
    lk["B3"].font = Font(name="Arial", bold=True)

    data_last = r - 1
    id_rng = f"'Net Sales'!$A${hr+1}:$A${data_last}"
    fields = [(lbl, key, kind) for (lbl, key, kind) in COLS if lbl != "Order ID"]
    rr = 5
    for lbl, key, kind in fields:
        lk.cell(row=rr, column=1, value=lbl).font = Font(name="Arial", bold=True, size=10)
        col_letter = get_column_letter(COLS.index(next(c for c in COLS if c[0]==lbl)) + 1)
        col_rng = f"'Net Sales'!${col_letter}${hr+1}:${col_letter}${data_last}"
        out = lk.cell(row=rr, column=2,
                      value=f'=IFERROR(INDEX({col_rng},MATCH($B$3,{id_rng},0)),"(not found)")')
        out.font = Font(name="Arial", size=10)
        if kind == "money":
            out.number_format = money_fmt
        rr += 1
    lk.column_dimensions["A"].width = 20
    lk.column_dimensions["B"].width = 34

    wb.save(out_path)
    return out_path

def main():
    out = sys.argv[1] if len(sys.argv) > 1 else "REQ-22-D01_ebay_product_net_sales.xlsx"
    rows = fetch_rows()
    if not rows:
        sys.exit("No rows returned — refusing to write an empty workbook.")
    write_workbook(rows, out)
    print(f"Wrote {out} ({len(rows)} orders)")

if __name__ == "__main__":
    main()

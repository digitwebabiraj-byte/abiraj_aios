# -*- coding: utf-8 -*-
"""REQ-12-D01 v4 - eBay Price Checker - Thinesh's 13 account+region labels.
Filters to only his named accounts, relabels Account column to his names.
Same corrected matching (all_list, _ suffix, ENC, PK) as v3."""
import json, io, os, collections
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

TR = r"C:\Users\digit\.claude\projects\C--Users-digit-OneDrive-Desktop-Abiraj-AIOS--claude-worktrees-project-structure-overview-2e1c41\935865a0-efe8-4d55-b8e9-aaf3b50f7d83\tool-results"
OUT = r"C:/Users/digit/OneDrive/Desktop/DigitWeb_Works_Abiraj/16_07_2026/2026-07-16_abiraj_REQ-epc_REQ-12-D01_price-checker_UI.xlsx"
CH = ["1784185487529","1784185514984","1784185525493","1784185542565","1784185570946",
      "1784185581243","1784185595968","1784185623333","1784185633137"]

THRESHOLD, TOL_LO, TOL_HI = 20.0, 0.50, 1.00
PRIO_HIGH, PRIO_MED = 5.00, 2.00

# (db_account, site) -> Thinesh's label.  Anything not here is dropped (he didn't name it).
LABEL = {
    ("led_sone","UK"):          "LEDSone UK",
    ("electricalsone","UK"):    "Electricalsone UK",
    ("so_926407","UK"):         "Sunsone UK",
    ("vintageinterior","UK"):   "Vintageinterior UK",
    ("coventrylights","UK"):    "Coventrylight UK",
    ("lighting_sone","UK"):     "Lightingsone UK",
    ("re6865","UK"):            "Retro LED UK",
    ("huettenlampen","Germany"):"HUETTEN LAMP DE",
    ("ledsonede","Germany"):    "Ledsone DE Reg DE",
    ("homin_gmbh","Germany"):   "Homin DE",
    ("led_sone","Germany"):     "LEDSone UK Reg DE",
    ("electricalsone","Germany"):"ElectricalSone DE",
    ("so_926407","Germany"):    "Sunsone DE",
}

rows = []
for t in CH:
    p = os.path.join(TR, "mcp-Ledsone-db-mcp-execute_sql-%s.txt" % t)
    b = json.loads(io.open(p, encoding="utf-8").read())["data"]["rows"][0]["blob"]
    rows += [l.split("|") for l in b.split("\n") if l.strip()]
assert all(len(r) == 9 for r in rows)
print("parsed:", len(rows))

dropped = collections.Counter()
out = []
for r in rows:
    item, sku, img, acct, wp, ap, tgt, ebay, site = r
    label = LABEL.get((acct, site))
    if label is None:
        dropped[(acct, site)] += 1
        continue
    wp = float(wp) if wp else None
    ap = float(ap) if ap else None
    tgt = float(tgt) if tgt else None
    ebay = float(ebay)
    if tgt is None:
        diff = pct = None; prio = "Unknown"
        if "+" in sku:
            status, action = "DATA MISSING – BUNDLE", "Bundle – price the kit"
        else:
            status, action = "DATA MISSING – NO COMPARATOR", "eBay-only – no Amazon/website match"
    else:
        diff = round(ebay - tgt, 2)
        pct = diff / tgt if tgt else None
        tol = TOL_LO if ebay < THRESHOLD else TOL_HI
        if abs(diff) <= tol:
            status, action = u"✅ Normal", "No Action Required"
        elif diff > 0:
            status, action = u"\U0001F534 High Price", "Reduce eBay Price"
        else:
            status, action = u"\U0001F534 Low Price", "Increase eBay Price"
        m = abs(diff)
        prio = "High" if m >= PRIO_HIGH else ("Medium" if m >= PRIO_MED else "Low")
    out.append((item, sku, img, label, wp, ap, tgt, ebay, diff, pct, status, prio, action, site))

print("kept:", len(out), "| dropped (unnamed accounts):", dict(dropped))
print("account counts:", dict(collections.Counter(r[3] for r in out)))

ARIAL = "Arial"
HDRF = Font(name=ARIAL, bold=True, color="FFFFFF", size=10)
HDR = PatternFill("solid", fgColor="1F3864")
BODY = Font(name=ARIAL, size=10)
GBP = u'£#,##0.00'; EUR = u'€#,##0.00'; PCT = '0.00%'

wb = openpyxl.Workbook()
p = wb.active
p.title = "Price checker"
hdrs = ["ID","SKU","Product Image","Account","Website Price","Amazon Price","Target eBay Price",
        "Current eBay Price","Difference","Difference (%)","Status","Priority","Action"]
for i, h in enumerate(hdrs, 1):
    c = p.cell(1, i, h); c.font = HDRF; c.fill = HDR; c.alignment = Alignment(horizontal="center")

for i, r in enumerate(out):
    item, sku, img, label, wp, ap, tgt, ebay, diff, pct, status, prio, action, site = r
    n = i + 2
    fmt = GBP if site == "UK" else EUR
    p.cell(n, 1, item).number_format = "@"
    p.cell(n, 2, sku); p.cell(n, 3, img); p.cell(n, 4, label)
    if wp is not None:   p.cell(n, 5, wp).number_format = fmt
    if ap is not None:   p.cell(n, 6, ap).number_format = fmt
    if tgt is not None:  p.cell(n, 7, tgt).number_format = fmt
    p.cell(n, 8, ebay).number_format = fmt
    if diff is not None: p.cell(n, 9, diff).number_format = fmt
    if pct is not None:  p.cell(n, 10, round(pct, 4)).number_format = PCT
    p.cell(n, 11, status); p.cell(n, 12, prio); p.cell(n, 13, action)

for col, w in zip("ABCDEFGHIJKLM", (15, 34, 46, 20, 14, 14, 17, 18, 12, 13, 15, 10, 30)):
    p.column_dimensions[col].width = w
for row in p.iter_rows(min_row=2):
    for c in row: c.font = BODY
p.freeze_panes = "A2"
p.auto_filter.ref = "A1:M%d" % (len(out) + 1)
wb.save(OUT)
print("SAVED:", OUT)
print("STATUS:", dict(collections.Counter(r[10] for r in out)))

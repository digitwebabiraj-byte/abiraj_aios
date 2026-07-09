#!/usr/bin/env python3
"""
Build the Table 5 Weekly Stock Check (ALL ASINs) for Thuwaraga as Excel + HTML.
Input : data_all.json  (the JSON array returned by generate_dataset_all_asins.sql)
Output: Table5_Weekly_Stock_Check_Thuwaraga_ALL.xlsx
        Table5_Weekly_Stock_Check_Thuwaraga_ALL.html
Run   : pip install openpyxl --break-system-packages ; python build_all.py
"""
import json, html
from datetime import date
from collections import Counter

rows = json.load(open("data_all.json"))
# de-dupe safety by (asin, account)
seen=set(); data=[]
for d in rows:
    k=(d["asin"], d["account"])
    if k in seen: continue
    seen.add(k); data.append(d)

HEADERS=["Last Stock Checked Date","ASIN","Account Name","Listing SKU","Correct SKU (Master)",
"Amazon Listing Qty (FBM)","UK Warehouse Stock (Real)","Last 3-Month Units Sold","Sales Velocity (Units/Day)",
"Days of Stock Remaining","Upcoming Supplier Name","PO Quantity (Incoming)","Container Number","Container Reaching Date","Stock Status"]
TODAY=date.today().isoformat()
COL={"Healthy Stock":"13B4CF","Going Out of Stock":"FFEB84","No Stock / Critical":"F4A6A6","No Recent Sales (Idle Stock)":"D9D9D9"}
HEAD="E8EAED"

def cells(d):
    return [TODAY,d["asin"],d["account"],d["listing_sku"],d["master_sku"],d["amazon_fbm"],d["uk_warehouse"],
            d["order_count_90"],
            d["velocity"] if d["velocity"] is not None else "",
            d["days_remaining"] if d["days_remaining"] is not None else "",
            d.get("suppliers") or "-", d.get("po_qty") or "-", d.get("containers") or "-","-", d["stock_status"]]

# ---------- Excel ----------
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
def fill(c): return PatternFill("solid", fgColor=c)
thin=Side(style="thin", color="BFBFBF"); border=Border(thin,thin,thin,thin); POP="Poppins"
wb=Workbook(); ws=wb.active; ws.title="Weekly Stock Check"; n=len(HEADERS)
ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=n)
c=ws.cell(2,1,"Weekly Stock Check — Thuwaraga (All Amazon UK ASINs)"); c.font=Font(POP,16,bold=True,color="FFFFFF"); c.alignment=Alignment("center","center"); c.fill=fill("13B4CF")
ws.merge_cells(start_row=3,start_column=1,end_row=3,end_column=n)
c=ws.cell(3,1,"Every Amazon-UK ASIN with a live FBM listing or a sale in the last 90 days. Real UK warehouse stock shown for all; velocity & days-of-stock only where there are sales; idle stock flagged."); c.font=Font(POP,9,italic=True,color="333333"); c.alignment=Alignment("center","center",wrap_text=True); ws.row_dimensions[3].height=30
ws.merge_cells(start_row=5,start_column=1,end_row=5,end_column=n)
c=ws.cell(5,1,f"Generated {TODAY} · UK stock = location_wise_inv_stock · read-only"); c.font=Font(POP,9,bold=True,color="333333"); c.fill=fill(HEAD); c.alignment=Alignment("center","center")
for j,h in enumerate(HEADERS,1):
    c=ws.cell(6,j,h); c.font=Font(POP,9,bold=True); c.fill=fill(HEAD); c.alignment=Alignment("center","center",wrap_text=True); c.border=border
ws.row_dimensions[6].height=42
r=7
for d in data:
    for j,v in enumerate(cells(d),1):
        cc=ws.cell(r,j,v); cc.font=Font(POP,9); cc.border=border; cc.alignment=Alignment("center","center",wrap_text=(j==13))
    sc=ws.cell(r,15); sc.fill=fill(COL.get(d["stock_status"],"FFFFFF")); sc.font=Font(POP,9,bold=True)
    r+=1
for j,w in enumerate([16,12,15,20,18,12,13,12,12,12,16,12,26,14,22],1): ws.column_dimensions[get_column_letter(j)].width=w
ws.freeze_panes="A7"
xlsx="Table5_Weekly_Stock_Check_Thuwaraga_ALL.xlsx"; wb.save(xlsx)

# ---------- HTML ----------
def td(v,bg=None):
    st=f' style="background:#{bg};font-weight:600"' if bg else ""
    return f"<td{st}>{html.escape(str(v))}</td>"
body=[]
for d in data:
    vals=cells(d); tds=""
    for j,v in enumerate(vals):
        tds+= td(v, COL.get(d["stock_status"]) if j==14 else None)
    body.append(f"<tr>{tds}</tr>")
counts=Counter(d["stock_status"] for d in data)
legend=" · ".join(f'<span style="background:#{COL[k]};padding:2px 8px;border-radius:3px">{k}: {counts.get(k,0)}</span>' for k in COL)
htmldoc=f"""<!doctype html><meta charset="utf-8">
<title>Weekly Stock Check — Thuwaraga</title>
<style>body{{font-family:Poppins,Arial,sans-serif;margin:24px;color:#222}}
h1{{background:#13B4CF;color:#fff;padding:12px;border-radius:6px;font-size:20px}}
.sub{{color:#555;font-size:13px;margin:8px 0}}
table{{border-collapse:collapse;width:100%;font-size:12px}}
th{{background:#E8EAED;position:sticky;top:0}} th,td{{border:1px solid #ccc;padding:5px 7px;text-align:center}}
.legend span{{margin-right:8px;font-size:12px}}</style>
<h1>Weekly Stock Check — Thuwaraga (All Amazon UK ASINs)</h1>
<div class="sub">Generated {TODAY} · {len(data)} ASINs · UK stock = location_wise_inv_stock · read-only.
Velocity &amp; days-of-stock only where there are sales; idle stock flagged.</div>
<div class="legend">{legend}</div>
<table><thead><tr>{''.join(f'<th>{html.escape(h)}</th>' for h in HEADERS)}</tr></thead>
<tbody>{''.join(body)}</tbody></table>"""
open("Table5_Weekly_Stock_Check_Thuwaraga_ALL.html","w").write(htmldoc)

print("rows:",len(data)," status:",dict(counts))
print("saved:",xlsx,"and Table5_Weekly_Stock_Check_Thuwaraga_ALL.html")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date
from dataset import DATA

TODAY = date.today()
GREY="E8EAED"; PEACOCK="13B4CF"; YELLOW="FFEB84"; RED="F4A6A6"; FONT="Poppins"

# --- recompute status & sort on REAL stock (UK warehouse) ---
def compute(d):
    wh = d["uk_warehouse"]; oc = d["order_count_90"]
    vel = oc/90 if oc else 0
    days = 0 if (wh==0 or vel==0) else round(wh/vel)
    if wh == 0:            st = "No Stock / Critical"
    elif days < 15:        st = "No Stock / Critical"
    elif days <= 60:       st = "Going Out of Stock"
    else:                  st = "Healthy Stock"
    return days, st

for d in DATA:
    d["_days"], d["_status"] = compute(d)
DATA.sort(key=lambda d: (0 if d["uk_warehouse"] == 0 else 1, d["_days"]))

status_fill = {"Healthy Stock":PEACOCK,"Going Out of Stock":YELLOW,"No Stock / Critical":RED}
thin=Side(style="thin",color="D9D9D9"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
center=Alignment(horizontal="center",vertical="center",wrap_text=True)
left=Alignment(horizontal="left",vertical="center",wrap_text=True)

wb=Workbook(); ws=wb.active; ws.title="Table 5"; ws.sheet_view.showGridLines=False

COLS=[("Last Stock\n Checked Date",13,"date","date"),("ASIN",13,"asin","text"),
("Account Name",16,"account","text"),("Listing SKU",20,"listing_sku","text"),
("Correct SKU\n (Master)",18,"master_sku","text"),
("Amazon Listing\n Qty (FBM)",13,"amazon_fbm","int"),
("UK Warehouse\n Stock (Real)",13,"uk_warehouse","int"),
("Last 3-Month\n Order Count",12,"order_count_90","int"),
("Sales Velocity\n (Units/Day)",12,"velocity","vel"),
("Days of Stock\n Remaining",12,"days","days"),
("Upcoming Supplier\n Name",18,"suppliers","supplier"),
("PO Quantity\n (Incoming)",12,"po_qty","po"),
("Container\n Number",26,"containers","container"),
("Container\n Reaching Date",13,"eta","eta"),
("Stock Status",17,"status","status")]
NC=len(COLS); last=get_column_letter(NC)

ws.merge_cells(f"A2:{last}2")
ws["A2"]=("T5 : Table 5 (HIGH PRIORITY) — STOCK MANAGEMENT ACROSS ALL ASINs & WAREHOUSES "
          "- Thuwaraga - Frequency - weekly - Every monday")
ws["A2"].font=Font(name=FONT,bold=True,size=13,color="20343B"); ws["A2"].alignment=left
ws.row_dimensions[2].height=22

ws.merge_cells(f"A3:{last}3")
ws["A3"]=("Master SKU maps every listing SKU/ASIN back to one true stock count. Real stock = UK Warehouse. "
          "Days of Stock Remaining = UK Warehouse Stock ÷ Sales Velocity (last 3 months, FBM). "
          "Amazon Listing Qty (FBM) is the count maintained on Amazon (reference only, not real stock). "
          "Row colors: Peacock = healthy, Yellow = going out, Red = no stock.")
ws["A3"].font=Font(name=FONT,size=9,color="555555")
ws["A3"].alignment=Alignment(horizontal="left",vertical="top",wrap_text=True)
ws.row_dimensions[3].height=42

ws.merge_cells(f"A5:{last}5")
ws["A5"]=("Weekly Stock Check — Amazon (FBM) + 3 UK Warehouses + Incoming Supplier Shipments   |   "
          "PH: Thuwaraga · Amazon UK · live as of "+TODAY.strftime("%Y-%m-%d"))
ws["A5"].font=Font(name=FONT,bold=True,size=10,color="20343B")
ws["A5"].fill=PatternFill("solid",fgColor=GREY); ws["A5"].alignment=center
ws.row_dimensions[5].height=20

HR=6
for ci,(t,w,k,kind) in enumerate(COLS,1):
    col=get_column_letter(ci); ws.column_dimensions[col].width=w
    c=ws.cell(HR,ci,t); c.fill=PatternFill("solid",fgColor=GREY)
    c.font=Font(name=FONT,bold=True,size=9,color="20343B"); c.alignment=center; c.border=border
ws.row_dimensions[HR].height=30

first=HR+1
for i,row in enumerate(DATA):
    r=first+i
    fill=PatternFill("solid",fgColor=status_fill.get(row["_status"],"FFFFFF"))
    for ci,(t,w,k,kind) in enumerate(COLS,1):
        c=ws.cell(r,ci); c.border=border; c.fill=fill
        c.font=Font(name=FONT,size=9,color="1A1A1A",bold=(kind in ("days","status")))
        c.alignment=left if kind in ("text","supplier","container") else center
        if kind=="date": c.value=TODAY; c.number_format="yyyy-mm-dd"
        elif kind=="text": c.value=row[k]
        elif kind=="int": c.value=row[k]; c.number_format="#,##0"
        elif kind=="vel": c.value=f"=H{r}/90"; c.number_format="0.00"
        elif kind=="days":
            c.value=f'=IF(OR(G{r}=0,I{r}=0),0,ROUND(G{r}/I{r},0))'; c.number_format="#,##0"
        elif kind=="supplier": c.value=row["suppliers"] if row["suppliers"] else "Pending PO"
        elif kind=="po": c.value=row["po_qty"] if row["po_qty"] else 0; c.number_format="#,##0"
        elif kind=="container": c.value=row["containers"] if row["containers"] else "-"
        elif kind=="eta": c.value="-"
        elif kind=="status":
            c.value=(f'=IF(G{r}=0,"No Stock / Critical",'
                     f'IF(J{r}<15,"No Stock / Critical",'
                     f'IF(J{r}<=60,"Going Out of Stock","Healthy Stock")))')
    ws.row_dimensions[r].height=22

last_row=first+len(DATA)-1
ws.freeze_panes=ws.cell(first,1)
ws.auto_filter.ref=f"A{HR}:{last}{last_row}"

fn=last_row+2
ws.merge_cells(f"A{fn}:{last}{fn}")
ws[f"A{fn}"]=("Notes: Real fulfillable stock = UK Warehouse (UK country total from location_wise_inv_stock — matches the live inventory system); Days & Status are based on this. "
              "Amazon Listing Qty (FBM) is the merchant-declared quantity shown on Amazon (mostly a flat 39/43 default) — reference only. "
              "Order Count & Velocity = Completed FBM orders, last 90 days ÷ 90. Container Reaching Date is not stored in the DB ('-'). "
              "Sorted by Days of Stock Remaining ascending (real stockouts first).")
ws[f"A{fn}"].font=Font(name=FONT,size=8,italic=True,color="777777")
ws[f"A{fn}"].alignment=Alignment(horizontal="left",vertical="top",wrap_text=True)
ws.row_dimensions[fn].height=34

out="/mnt/user-data/outputs/Table5_Weekly_Stock_Check_Thuwaraga.xlsx"
wb.save(out)
from collections import Counter
print("saved", out, "rows", len(DATA), Counter(d["_status"] for d in DATA))

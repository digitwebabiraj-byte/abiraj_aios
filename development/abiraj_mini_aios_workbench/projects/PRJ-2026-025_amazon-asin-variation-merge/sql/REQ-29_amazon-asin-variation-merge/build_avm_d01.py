"""
REQ-29-D01 · step 2 of 2 — the single generator module.

Reads:  listings.amazon_listings (read-only SQL)  +  avm_ratings_cache.json (scraped ratings)
Writes: avm_payload.json  +  REQ-29-D01_asin_variation_merge.xlsx

✅ THE BUSINESS RULES BELOW WERE **CONFIRMED BY PRASATH 2026-08-18** (discovery answers Q2-Q7).
  They are declared in one block so a reviewer can see and change them without reading code.
  Every output states them.  STILL OPEN: Q1 long-term rating source, Q8 approval mechanism,
  Q10 platform-vs-warehouse stock, Q11 automation.

READ-ONLY: SELECT only.  No INSERT/UPDATE/DELETE/DDL.  No merge is ever executed.
"""
import json, os, re, sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import psycopg2
import psycopg2.extras

HERE = Path(__file__).resolve().parent
OUT = HERE.parent.parent / "evidence" / "final_outputs" / "REQ-29_amazon-asin-variation-merge"
CACHE = HERE / "avm_ratings_cache.json"
PAYLOAD = HERE / "avm_payload.json"

# ───────────────── BUSINESS RULES — CONFIRMED BY PRASATH 2026-08-18 ─────────────────
RULES = {
    "low_rating_below": 3.5,          # Q4  a child is "low rated" BELOW 3.5 stars
    "no_reviews_at_or_below": 0,      # Q4  "no reviews" means EXACTLY 0 reviews
    "parent_min_rating": 3.5,         # Q5  a parent must have a rating >= 3.5
    "parent_min_reviews": 1,          # Q5  implied: a rating cannot exist with 0 reviews
    "parent_pick": "most_reviews_then_fewer_variations",   # Q5  highest reviews; tie -> fewer children
    "no_suitable_parent": "show",     # Q5  show the family for manual review, never hide it
    "duplicate_match": "normalised",  # Q6  smart match - ignore case and spacing
    "duplicate_attributes": ["color", "size"],   # Q6  compare colour AND size
    "out_of_stock": "warn",           # Q7  show with a warning; the operator decides
}
SITE = "UK"                                    # Q3  UK only
SUB_SOURCES = [8, 6]                           # Q2  amazon Ledsone AND amazon Dcvoltage
ACCOUNT_NAME = {8: "amazon Ledsone", 6: "amazon Dcvoltage"}
# ──────────────────────────────────────────────────────────────────────────────────


def db():
    return psycopg2.connect(
        host=os.environ["LED_PGHOST"], port=os.environ.get("LED_PGPORT", "5432"),
        dbname=os.environ["LED_PGDATABASE"], user=os.environ["LED_PGUSER"],
        password=os.environ["LED_PGPASSWORD"], connect_timeout=30,
    )


def fetch_listings(conn, asins):
    sql = """
    SELECT asin, sku, parent_sku, sub_source, is_parent, is_child, status, quantity,
           title, price, currency, selected_variations
    FROM listings.amazon_listings
    WHERE site=%s AND sub_source = ANY(%s) AND asin = ANY(%s)
    ORDER BY sub_source, parent_sku, is_parent DESC, asin;
    """
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute(sql, (SITE, SUB_SOURCES, list(asins)))
        return [dict(r) for r in cur.fetchall()]


def colour_of(selected_variations):
    """Extract the colour attribute. Returns None when the listing carries none.

    `selected_variations` is jsonb and is NOT uniformly shaped across the catalogue:
    most rows hold [{"name": "color", "value": "Black"}], but some hold a bare string,
    a dict, or entries missing 'name'. Anything we cannot read as a colour returns None
    (-> "NO COLOUR" in the report) rather than crashing or guessing.
    """
    if not selected_variations:
        return None
    items = selected_variations
    if isinstance(items, dict):
        items = [items]
    if isinstance(items, str):
        return items.strip() or None
    if not isinstance(items, (list, tuple)):
        return None
    for v in items:
        if not isinstance(v, dict):
            continue
        if str(v.get("name", "")).lower() in RULES["duplicate_attributes"]:
            value = v.get("value")
            return (str(value).strip() or None) if value is not None else None
    return None


def attrs_of(selected_variations):
    """Return {attribute_name: value} for the attributes the duplicate check compares (Q6)."""
    out = {}
    items = selected_variations
    if not items:
        return out
    if isinstance(items, dict):
        items = [items]
    if isinstance(items, str):
        return {"color": items.strip()} if items.strip() else {}
    if not isinstance(items, (list, tuple)):
        return out
    for v in items:
        if not isinstance(v, dict):
            continue
        name = str(v.get("name", "")).lower()
        if name in RULES["duplicate_attributes"]:
            value = v.get("value")
            if value is not None and str(value).strip():
                out[name] = str(value).strip()
    return out


def variation_key(attrs):
    """Q6: the duplicate identity is colour AND size together, smart-matched.

    Missing attributes are kept in the key as '' so that a colour-only listing and a
    colour+size listing are NOT silently treated as the same variation.
    """
    if not attrs:
        return None
    return "|".join(f"{a}={norm(attrs.get(a, '')) or ''}"
                    for a in RULES["duplicate_attributes"])


def norm(value):
    """Normalisation used by the duplicate check (rule #7)."""
    if value is None:
        return None
    if RULES["duplicate_match"] == "exact":
        return value
    return re.sub(r"[^a-z0-9]+", " ", value.lower()).strip()


def main():
    if not CACHE.exists():
        sys.exit("no rating cache — run scrape_amazon_ratings.py first")
    ratings = json.loads(CACHE.read_text())
    scraped_ok = {a: r for a, r in ratings.items() if r["status"] in ("ok", "no_reviews")}

    conn = db()
    rows = fetch_listings(conn, list(ratings.keys()))
    conn.close()

    families = defaultdict(list)
    for r in rows:
        r["rating"] = (scraped_ok.get(r["asin"]) or {}).get("rating")
        r["reviews"] = (scraped_ok.get(r["asin"]) or {}).get("reviews")
        r["rating_status"] = (ratings.get(r["asin"]) or {}).get("status", "not_fetched")
        r["colour"] = colour_of(r["selected_variations"])
        r["attrs"] = attrs_of(r["selected_variations"])
        r["vkey"] = variation_key(r["attrs"])
        r["account"] = ACCOUNT_NAME.get(r["sub_source"], f"sub_source {r['sub_source']}")
        families[r["parent_sku"]].append(r)

    report, skipped_no_parent = [], 0

    for base_sku, members in sorted(families.items(), key=lambda kv: kv[0] or ""):
        # a listing with no parent_sku belongs to no variation family — nothing to merge into
        if not base_sku:
            continue
        rated = [m for m in members if m["reviews"] is not None]
        if not rated:
            continue

        # ── Q5: pick the parent ──
        # Eligible = rating >= 3.5. Among eligible, HIGHEST REVIEW COUNT wins.
        # Tie on reviews -> the one with FEWER child variations.
        # "Child variations" is not a stored column, so it is derived: ASINs in this family
        # sharing an identical (rating, reviews) pair are already one Amazon variation, so
        # that count is how many children the candidate already carries. Documented proxy.
        share_count = defaultdict(int)
        for m in rated:
            if m["rating"] is not None:
                share_count[(m["rating"], m["reviews"])] += 1

        eligible = [m for m in rated
                    if (m["rating"] or 0) >= RULES["parent_min_rating"]
                    and (m["reviews"] or 0) >= RULES["parent_min_reviews"]]

        if not eligible:
            # Q5: NEVER hide these — show the family for manual review.
            skipped_no_parent += 1
            weak = sorted(rated, key=lambda m: (m["reviews"] or 0), reverse=True)[0]
            report.append({
                "platform": "AMAZON",
                "account": weak["account"],
                "base_sku": base_sku,
                "parent_asin": "NO SUITABLE PARENT",
                "parent_rating_reviews": "— no member rated 3.5+ —",
                "parent_title": "",
                "child_asin_sku": f"{weak['asin']} / {weak['sku']}",
                "child_asin": weak["asin"],
                "child_colour_rating": (f"{weak['colour'] or 'NO COLOUR'} / "
                                        f"{weak['rating'] if weak['rating'] is not None else 0.0:.1f}"),
                "child_reviews": weak["reviews"] or 0,
                "merge_reason": f"No suitable parent — manual review ({len(members)} in family)",
                "stock_status": "In Stock" if (weak["quantity"] or 0) > 0 else "Out of Stock",
                "duplicate_warning": "n/a",
                "approved": "",
                "operator_notes": "",
                "_blocked": False,
                "_no_parent": True,
                "_flags": ["no member of this family is rated 3.5 or above"],
            })
            continue

        parent = sorted(
            eligible,
            key=lambda m: (-(m["reviews"] or 0),                       # most reviews first
                           share_count[(m["rating"], m["reviews"])],   # tie -> fewer variations
                           m["asin"]),                                 # stable
        )[0]

        # variation identities (colour+size) already present under this parent
        taken = defaultdict(int)
        for m in members:
            if m["asin"] != parent["asin"] and m["vkey"]:
                taken[m["vkey"]] += 1

        # ── Q4: which children are merge candidates ──
        for child in members:
            if child["asin"] == parent["asin"] or child["reviews"] is None:
                continue
            no_rev = child["reviews"] <= RULES["no_reviews_at_or_below"]
            low = child["rating"] is not None and child["rating"] < RULES["low_rating_below"]
            if not (no_rev or low):
                continue
            # never recommend merging into a weaker review history
            if (child["reviews"] or 0) >= (parent["reviews"] or 0):
                continue

            reason = ("No reviews — merge into stronger parent" if no_rev
                      else "Low rating — merge into higher-rated parent")
            dup = bool(child["vkey"]) and taken.get(child["vkey"], 0) > 1
            in_stock = (child["quantity"] or 0) > 0
            attr_txt = " · ".join(f"{k}: {v}" for k, v in child["attrs"].items()) or "NO COLOUR"

            report.append({
                "platform": "AMAZON",
                "account": child["account"],
                "base_sku": base_sku,
                "parent_asin": parent["asin"],
                "parent_rating_reviews": f"{parent['rating']:.1f} / {parent['reviews']}",
                "parent_title": (parent["title"] or "")[:90],
                "child_asin_sku": f"{child['asin']} / {child['sku']}",
                "child_asin": child["asin"],
                "child_colour_rating": (f"{attr_txt} / "
                                        f"{child['rating'] if child['rating'] is not None else 0.0:.1f}"),
                "child_reviews": child["reviews"],
                "merge_reason": reason,
                # Q7: out of stock is a WARNING the operator may override, never a rejection
                "stock_status": "In Stock" if in_stock else "Out of Stock",
                "duplicate_warning": "Yes" if dup else "No",
                "approved": "",                    # operator input — never pre-filled
                "operator_notes": "",              # operator input — never pre-filled
                "_blocked": dup,                   # only a duplicate blocks; stock only warns
                "_no_parent": False,
                "_flags": (["duplicate colour+size already under parent"] if dup else [])
                          + ([] if in_stock else ["out of stock — operator decides"]),
            })

    report.sort(key=lambda r: (r.get("_no_parent", False), r["_blocked"],
                               -(r["child_reviews"] == 0), r["base_sku"]))

    errors = [a for a, r in ratings.items() if r["status"].startswith("error")]
    payload = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "scope": {"platform": "Amazon", "site": SITE,
                  "accounts": [ACCOUNT_NAME[s_] for s_ in SUB_SOURCES],
                  "sub_sources": SUB_SOURCES},
        "rules": RULES,
        "coverage": {
            "asins_rating_collected": len(scraped_ok),
            "asins_rating_failed": len(errors),
            "families_examined": len(families),
            "families_no_qualifying_parent": skipped_no_parent,
        },
        "kpi": {
            "total_asins": len(scraped_ok),
            "no_review_or_low_rated": sum(
                1 for a, r in scraped_ok.items()
                if (r["reviews"] or 0) <= RULES["no_reviews_at_or_below"]
                or (r["rating"] is not None and r["rating"] < RULES["low_rating_below"])),
            "recommend_merge": sum(1 for r in report if not r["_blocked"]),
            "needs_review": sum(1 for r in report if r["_blocked"]),
        },
        "status_overview": {
            "recommend_merge": sum(1 for r in report if not r["_blocked"]),
            "needs_review": sum(1 for r in report if r["_blocked"]),
            "duplicate_warnings": sum(1 for r in report if r["duplicate_warning"] == "Yes"),
            "out_of_stock": sum(1 for r in report if r["stock_status"] == "Out of Stock"),
        },
        "rows": report,
    }
    PAYLOAD.write_text(json.dumps(payload, indent=1))
    print(f"payload -> {PAYLOAD.name}: {len(report)} merge candidates from "
          f"{len(families)} families | KPI {payload['kpi']}", flush=True)
    write_excel(payload)
    return 0


def write_excel(p):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    OUT.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    ws = wb.active
    ws.title = "Notes & Method"
    notes = [
        ("Amazon ASIN Rating Analysis & Variation Merging", True),
        (f"REQ-29-D01 · generated {p['generated_at']}", False),
        ("", False),
        ("⚠ PILOT / DRAFT — PROVISIONAL RULES PENDING PRASATH'S CONFIRMATION", True),
        ("Every threshold below is a documented default, not agreed business logic.", False),
        ("", False),
        ("SCOPE", True),
        (f"Platform Amazon · Site {p['scope']['site']} · Accounts "
         f"{', '.join(p['scope']['accounts'])} (sub_source {p['scope']['sub_sources']})", False),
        ("", False),
        ("WHERE THE DATA COMES FROM", True),
        ("Catalogue (ASIN, SKU, parent/child, colour, stock, title): listings.amazon_listings "
         "— live, read-only.", False),
        ("Star rating + review count: collected from the public Amazon UK product page. "
         "These exist in NEITHER company database (owner-confirmed 2026-08-18), so they cannot "
         "come from SQL.", False),
        ("", False),
        ("COVERAGE — this is a PILOT, not the full catalogue", True),
        (f"Ratings collected for {p['coverage']['asins_rating_collected']} ASINs "
         f"across {p['coverage']['families_examined']} variation families.", False),
        (f"Failed to collect: {p['coverage']['asins_rating_failed']} ASINs "
         "(recorded as errors — NOT counted as 'no reviews').", False),
        (f"Families skipped because no member met the parent bar: "
         f"{p['coverage']['families_no_qualifying_parent']}.", False),
        ("The full Amazon UK catalogue is 16,963 ASINs / 1,761 multi-child families. "
         "Scale up once the rules below are confirmed.", False),
        ("", False),
        ("PROVISIONAL RULES USED", True),
        (f"A child is a merge candidate if reviews <= {p['rules']['no_reviews_at_or_below']} "
         f"OR rating < {p['rules']['low_rating_below']}.", False),
        (f"The parent must have >= {p['rules']['parent_min_reviews']} reviews and "
         f"rating >= {p['rules']['parent_min_rating']}; chosen by most reviews, then highest rating.", False),
        (f"Duplicate colour check: {p['rules']['duplicate_match']} matching on "
         f"{', '.join(p['rules']['duplicate_attributes'])}.", False),
        (f"Out of stock: {p['rules']['out_of_stock']}.", False),
        ("", False),
        ("CONTROL", True),
        ("NO MERGE EXECUTES WITHOUT PH/OPERATOR APPROVAL. This report recommends only — "
         "it never re-parents a listing and never writes to Amazon.", False),
        ("Approved (Y/N) and Operator Notes are deliberately blank for the operator to fill.", False),
        ("Approved merges are executed by hand via the Seller Central flat-file process.", False),
    ]
    for i, (text, bold) in enumerate(notes, 1):
        c = ws.cell(row=i, column=1, value=text)
        c.font = Font(bold=bold, size=13 if i == 1 else 11)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 118

    ws2 = wb.create_sheet("ASIN Merge Task")
    headers = ["PLATFORM", "ACCOUNT", "BASE SKU", "PARENT ASIN", "PARENT RATING / REVIEWS",
               "CHILD ASIN / SKU", "CHILD COLOUR / RATING", "MERGE REASON", "STOCK STATUS",
               "DUPLICATE WARNING", "APPROVED (Y/N)", "OPERATOR NOTES"]
    hf = PatternFill("solid", fgColor="1F3864")
    for j, h in enumerate(headers, 1):
        c = ws2.cell(row=1, column=j, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = hf
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    warn = PatternFill("solid", fgColor="FCE4E4")
    for i, r in enumerate(p["rows"], 2):
        vals = [r["platform"], r["account"], r["base_sku"], r["parent_asin"],
                r["parent_rating_reviews"], r["child_asin_sku"], r["child_colour_rating"],
                r["merge_reason"], r["stock_status"], r["duplicate_warning"],
                r["approved"], r["operator_notes"]]
        for j, v in enumerate(vals, 1):
            c = ws2.cell(row=i, column=j, value=v)
            c.alignment = Alignment(vertical="top", wrap_text=j in (5, 6, 7, 8, 12))
            if r["_blocked"]:
                c.fill = warn
    for j, w in enumerate([10, 15, 16, 14, 20, 34, 26, 40, 13, 12, 14, 30], 1):
        ws2.column_dimensions[get_column_letter(j)].width = w
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:L{max(2, len(p['rows']) + 1)}"

    ws3 = wb.create_sheet("Field Reference")
    ref = [("FIELD", "PURPOSE"),
           ("Base SKU", "Product family identifier."),
           ("Parent ASIN", "Selected parent variation ASIN."),
           ("Parent Rating / Reviews", "Parent star rating and review count."),
           ("Child ASIN / SKU", "ASIN to merge as child and its SKU."),
           ("Child Colour / Rating", "Child variation colour and current rating."),
           ("Merge Reason", "Reason for merging, such as no reviews or low rating."),
           ("Stock Status", "In Stock / Out of Stock."),
           ("Duplicate Warning", "Yes/No — whether the colour/variation is duplicated."),
           ("Approved (Y/N)", "Operator approval decision."),
           ("Operator Notes", "Space for operator comments.")]
    for i, (a, b) in enumerate(ref, 1):
        for j, v in enumerate((a, b), 1):
            c = ws3.cell(row=i, column=j, value=v)
            if i == 1:
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = hf
    ws3.column_dimensions["A"].width = 26
    ws3.column_dimensions["B"].width = 62

    path = OUT / "REQ-29-D01_asin_variation_merge.xlsx"
    wb.save(path)
    print(f"excel  -> {path}", flush=True)


if __name__ == "__main__":
    sys.exit(main())

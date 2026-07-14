import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

data = json.load(open('/home/claude/frrc30.json'))
def rate(r): return (r['total_returns']/r['units_sold']) if r['units_sold'] else 999
data.sort(key=lambda r: (-r['total_returns'], -rate(r), r['asin']))

WIN_START, WIN_END, RUN_DATE = "2026-06-14", "2026-07-13", "2026-07-14"
ARIAL="Arial"
def F(sz=10,b=False,color="000000"): return Font(name=ARIAL,size=sz,bold=b,color=color)
thin=Side(style="thin",color="BFBFBF"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
HDR=PatternFill("solid",fgColor="1F3864"); TITLE=PatternFill("solid",fgColor="1F3864")
BLUE=PatternFill("solid",fgColor="DDEBF7")
WRAP=Alignment(wrap_text=True,vertical="top"); CTR=Alignment(horizontal="center",vertical="center",wrap_text=True)

wb=Workbook()

# ---------- Objective & Guide ----------
ws=wb.active; ws.title="Objective & Guide"; ws.sheet_view.showGridLines=False
ws.column_dimensions['A'].width=2; ws.column_dimensions['B'].width=30
ws.column_dimensions['C'].width=56; ws.column_dimensions['D'].width=56
def g(r,b=None,c=None,title=False,sub=False,bold=False,color="000000"):
    if b is not None:
        cell=ws.cell(row=r,column=2,value=b)
        if title: cell.font=F(14,True,"FFFFFF"); cell.fill=TITLE
        elif sub: cell.font=F(11,True,"1F3864")
        else: cell.font=F(10,bold,color); cell.alignment=WRAP
    if c is not None:
        cc=ws.cell(row=r,column=3,value=c); cc.font=F(10); cc.alignment=WRAP
g(2,b="FBA Returns Tracker & Root-Cause Analysis  (FRRC)",title=True)
g(3,b="Deliverable REQ-10-D01 · live from the Postgres analytics DB · run "+RUN_DATE+" · window = last 30 days")
g(5,b="WINDOW & SOURCE",sub=True)
g(6,b="Reporting window: last 30 days = "+WIN_START+" to "+WIN_END+" inclusive (run-day "+RUN_DATE+" excluded — no partial current-day data).")
g(7,b="Built from public.amazon_returns (FBA returns) joined to public.order_transaction (real Units Sold). Only ASINs with >=1 FBA return in the window appear; zero-return ASINs are omitted. This window: 91 ASINs, 105 returns.")
g(9,b="OBJECTIVE",sub=True)
g(10,b="Show which Amazon FBA products are being returned too often and why, and route each to the person who owns it. For every returning ASIN: real Units Sold, Total Returns, Return Rate %, the return split by reason bucket, and a computed Flag / Root Cause / Recommended Action — so a listing problem, a quality/supplier problem and ordinary buyer preference can be told apart and acted on.")
g(12,b="CROSS-CHECK vs REBECCA'S TRACKER (correctness proof)",sub=True)
for i,t in enumerate([
 "The source tracker holds real Amazon data (05-11 to 07-12). Running this exact method on that same window reproduces it:",
 "  • Returns: 95 of 101 ASINs match exactly. Total 123 vs 129 — the few gaps are ASINs whose returns were removed/reclassified since her manual snapshot.",
 "  • Units Sold: 65 of 101 match exactly; every other case is 1-3 higher in live data (a few more Completed orders recorded in that window since her pull).",
 "  • This confirms the Units Sold rule (Amazon+FBA+UK+Completed) and the returns rule (fulfilment='fba', request_date in window). No order_date re-alignment is needed.",
]):
    ws.cell(row=13+i,column=2,value=t).font=F(10,color=("1F3864" if i==0 else "000000")); ws.cell(row=13+i,column=2).alignment=WRAP
    ws.merge_cells(start_row=13+i,start_column=2,end_row=13+i,end_column=4)
r=18
g(r,b="HOW EACH COLUMN IS BUILT",sub=True); r+=1
for a,b in [
 ("SKU (inventory)","Canonical inventory SKU resolved from the returns ASIN via the listing_data bridge (which_channel=1, wrong_sku=0, mapped_sku>sku). 89/91 resolved; 2 ASINs not in listing_data fall back to the return SKU."),
 ("Return SKU (Amazon)","The listing/variant SKU exactly as filed on the Amazon return — kept beside the inventory SKU so every row is traceable."),
 ("Units Sold (Period)","order_transaction: source_name='AMAZON', fba_sales=TRUE, market_place='UK', order_status='Completed', order_date in window, per ASIN."),
 ("Total Returns","amazon_returns: fulfilment='fba', request_date in window, per ASIN (SUM of qty)."),
 ("Return Rate %","Total Returns / Units Sold. 'N/A' when the ASIN had returns but no in-window UK-FBA sales."),
 ("Reason bucket Qty","Each return's reason mapped to one bucket (see Thresholds). Buckets sum exactly to Total Returns."),
 ("Flag / Root Cause / Action","Computed live by formulas that read the Thresholds tab — nothing hardcoded in the row logic."),
 ("Responsible Person","order_transaction.user_name (portfolio holder) for that ASIN — who acts on the row."),
]:
    ws.cell(row=r,column=2,value=a).font=F(10,True); ws.cell(row=r,column=2).alignment=WRAP
    ws.cell(row=r,column=3,value=b).font=F(10); ws.cell(row=r,column=3).alignment=WRAP; r+=1
r+=1
g(r,b="OPEN ITEMS — flagged to Satheesvaran (not decided here)",sub=True); r+=1
for n in [
 "Unmapped reason codes MISSING_PARTS, SWITCHEROO, MISSED_ESTIMATED_DELIVERY, POOR_FIT, MISORDERED, UNAUTHORIZED_PURCHASE are not in the tracker's bucket map — currently counted under 'Unknown' so totals reconcile. Confirm the correct bucket for each.",
 "Order-status counting as a 'sale' = Completed (confirmed by the cross-check). Confirm if any other status should count.",
 "Multi-PH ASINs: where an ASIN maps to more than one user_name, the most frequent is shown; confirm the tie-break rule.",
]:
    ws.cell(row=r,column=2,value="•  "+n).font=F(10,color="C00000"); ws.cell(row=r,column=2).alignment=WRAP
    ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=4); r+=1

# ---------- Thresholds ----------
th=wb.create_sheet("Thresholds"); th.sheet_view.showGridLines=False
th.column_dimensions['A'].width=2; th.column_dimensions['B'].width=52
th.column_dimensions['C'].width=12; th.column_dimensions['D'].width=72
c=th.cell(row=2,column=2,value="Threshold Settings"); c.font=F(14,True,"FFFFFF"); c.fill=TITLE
th.cell(row=3,column=2,value="Edit the blue cells to change how the Tracker flags and classifies each ASIN. Return Rate % = Total Returns / Units Sold in the same window.").font=F(10)
th.cell(row=3,column=2).alignment=WRAP
for col,txt in ((2,"Threshold"),(3,"Value"),(4,"Used In / Rule")):
    cc=th.cell(row=5,column=col,value=txt); cc.font=F(10,True,"FFFFFF"); cc.fill=HDR; cc.border=BORDER; cc.alignment=CTR
thr=[("Critical Return Rate — above",0.2,"Flag = CRITICAL - URGENT REVIEW when Return Rate % exceeds this","0.0%"),
 ("High Return Rate — above",0.1,"Flag = HIGH RETURN - REVIEW when Return Rate % exceeds this (and below Critical)","0.0%"),
 ("Minimum Returns to Evaluate — at/above",2,"Root-cause logic runs only at/above this return count","0"),
 ("Listing Mismatch share — at/above",0.4,"Root Cause = Listing/Expectation Mismatch when Listing Qty / Total Returns >= this","0.0%"),
 ("Quality share — at/above",0.4,"Root Cause = Quality/Defect Issue when Quality Qty / Total Returns >= this","0.0%"),
 ("Buyer Preference share — at/above",0.5,"Root Cause = Buyer Preference when Buyer Pref Qty / Total Returns >= this","0.0%")]
r=6
for label,val,rule,fmt in thr:
    th.cell(row=r,column=2,value=label).font=F(10); th.cell(row=r,column=2).border=BORDER; th.cell(row=r,column=2).alignment=WRAP
    vc=th.cell(row=r,column=3,value=val); vc.font=F(10,True,"0000FF"); vc.fill=BLUE; vc.border=BORDER; vc.number_format=fmt; vc.alignment=CTR
    th.cell(row=r,column=4,value=rule).font=F(10); th.cell(row=r,column=4).border=BORDER; th.cell(row=r,column=4).alignment=WRAP; r+=1
r+=1
th.cell(row=r,column=2,value=("Reason buckets — Listing Mismatch: NOT_COMPATIBLE, NOT_AS_DESCRIBED | Quality: QUALITY_UNACCEPTABLE, DEFECTIVE, DAMAGED_BY_FC, DAMAGED_BY_CARRIER | "
 "Buyer Preference: UNWANTED_ITEM, FOUND_BETTER_PRICE, ORDERED_WRONG_ITEM | Shipping: UNDELIVERABLE_UNKNOWN, UNDELIVERABLE_REFUSED | Unknown: NO_REASON_GIVEN")).font=F(10)
th.cell(row=r,column=2).alignment=WRAP; th.merge_cells(start_row=r,start_column=2,end_row=r,end_column=4); r+=2
th.cell(row=r,column=2,value=("HELD (confirm with Satheesvaran): these live codes are NOT in the map and are currently under 'Unknown' so buckets reconcile — "
 "MISSING_PARTS, SWITCHEROO, MISSED_ESTIMATED_DELIVERY, POOR_FIT, MISORDERED, UNAUTHORIZED_PURCHASE.")).font=F(10,color="C00000")
th.cell(row=r,column=2).alignment=WRAP; th.merge_cells(start_row=r,start_column=2,end_row=r,end_column=4); r+=2
th.cell(row=r,column=2,value="Blue cells = editable inputs. Changing a value re-flags every ASIN on the Tracker automatically.").font=F(10,True)
th.cell(row=r,column=2).alignment=WRAP; th.merge_cells(start_row=r,start_column=2,end_row=r,end_column=4)

# ---------- Tracker ----------
tr=wb.create_sheet("Tracker"); tr.sheet_view.showGridLines=False
t=tr.cell(row=1,column=1,value="FBA RETURNS TRACKER · live · window "+WIN_START+" to "+WIN_END+" (last 30 days) · cross-checked vs Rebecca's tracker")
t.font=F(12,True,"FFFFFF"); t.fill=TITLE; tr.merge_cells(start_row=1,start_column=1,end_row=1,end_column=17)
headers=["SKU\n(inventory)","ASIN","Return SKU\n(Amazon)","Units Sold\n(Period)","Total\nReturns","Return Rate\n%",
 "Listing\nMismatch Qty","Quality\nIssue Qty","Buyer\nPreference Qty","Shipping\nIssue Qty","Unknown\nQty","Top Reason\n(Amazon)",
 "Flag Status\n(ours)","Root Cause\n(ours)","Recommended Action\n(ours)","Responsible\nPerson","Notes"]
for i,h in enumerate(headers,start=1):
    cc=tr.cell(row=3,column=i,value=h); cc.font=F(10,True,"FFFFFF"); cc.fill=HDR; cc.border=BORDER; cc.alignment=CTR
widths=[26,12,28,9,8,10,10,10,11,10,9,20,24,26,40,14,20]
for i,w in enumerate(widths,start=1): tr.column_dimensions[get_column_letter(i)].width=w
row=4
for d in data:
    tr.cell(row=row,column=1,value=d['sku']).font=F(10)
    tr.cell(row=row,column=2,value=d['asin']).font=F(10)
    tr.cell(row=row,column=3,value=d['return_sku']).font=F(10)
    tr.cell(row=row,column=4,value=d['units_sold']).font=F(10)
    tr.cell(row=row,column=5,value=d['total_returns']).font=F(10)
    e=tr.cell(row=row,column=6,value=f'=IFERROR(E{row}/D{row},"N/A")'); e.font=F(10); e.number_format="0.0%"
    tr.cell(row=row,column=7,value=d['listing_qty']).font=F(10)
    tr.cell(row=row,column=8,value=d['quality_qty']).font=F(10)
    tr.cell(row=row,column=9,value=d['buyer_qty']).font=F(10)
    tr.cell(row=row,column=10,value=d['shipping_qty']).font=F(10)
    tr.cell(row=row,column=11,value=d['unknown_qty']).font=F(10)
    tr.cell(row=row,column=12,value=d['top_reason']).font=F(10)
    tr.cell(row=row,column=13,value=(f'=IF(F{row}="N/A","N/A - No Sales Data",'
        f'IF(F{row}>Thresholds!$C$6,"CRITICAL - URGENT REVIEW",'
        f'IF(F{row}>Thresholds!$C$7,"HIGH RETURN - REVIEW","OK")))')).font=F(10)
    tr.cell(row=row,column=14,value=(f'=IF(E{row}<Thresholds!$C$8,"Too few returns to evaluate",'
        f'IF(G{row}/E{row}>=Thresholds!$C$9,"Listing/Expectation Mismatch",'
        f'IF(H{row}/E{row}>=Thresholds!$C$10,"Quality/Defect Issue",'
        f'IF(I{row}/E{row}>=Thresholds!$C$11,"Buyer Preference - not a product issue",'
        f'"Mixed reasons - no single dominant cause"))))')).font=F(10)
    tr.cell(row=row,column=15,value=(f'=IF(M{row}="OK","Monitor - no action needed",'
        f'IF(N{row}="Listing/Expectation Mismatch","Update title/images/description to match product; review A+ content",'
        f'IF(N{row}="Quality/Defect Issue","Raise with supplier/QC, inspect next inbound shipment",'
        f'IF(N{row}="Buyer Preference - not a product issue","Monitor only - not a product/listing issue",'
        f'IF(N{row}="Too few returns to evaluate","Monitor - insufficient data","Review manually - mixed signal")))))')).font=F(10)
    tr.cell(row=row,column=16,value=d['responsible_ph']).font=F(10)
    note = "reason code not in bucket map -> counted as Unknown (confirm)" if d['top_reason'] in (
        'MISSING_PARTS','SWITCHEROO','MISSED_ESTIMATED_DELIVERY','POOR_FIT','MISORDERED','UNAUTHORIZED_PURCHASE') else (
        "ASIN not in listing_data; SKU = return SKU" if not d['inv_sku'] else None)
    tr.cell(row=row,column=17,value=note).font=F(10,color="C00000")
    for col in range(1,18):
        cc=tr.cell(row=row,column=col); cc.border=BORDER
        cc.alignment=Alignment(horizontal="center") if col in (4,5,6,7,8,9,10,11) else Alignment(vertical="top",wrap_text=(col in (14,15,17)))
    row+=1
tr.freeze_panes="A4"; tr.auto_filter.ref=f"A3:Q{row-1}"
wb.save("/home/claude/FRRC_FBA_Returns_Tracker_30d.xlsx")
print("saved rows:",row-4)

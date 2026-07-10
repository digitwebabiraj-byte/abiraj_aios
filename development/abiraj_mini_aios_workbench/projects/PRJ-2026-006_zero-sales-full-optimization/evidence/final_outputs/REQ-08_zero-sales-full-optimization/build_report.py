"""Build the ZSFO template-matching .xlsx from the governed data.json.
Read-only: consumes data.json (produced by generate_dataset.sql via the Postgres MCP).
Run:  python build_report.py
"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(HERE, "data.json")
OUT  = os.path.join(HERE, "ZSFO_Zero_Sales_Full_Optimization_Utharsika.xlsx")

with open(DATA, encoding="utf-8") as f:
    d = json.load(f)
meta, rows = d["meta"], d["rows"]

wb = Workbook()
ws = wb.active
ws.title = "ZSFO"

# ---- palette ----
NAVY   = "1F2A44"
HEADBG = "2E3B55"
STRIPE = "F4F6FA"
REDBG  = "FCE4E4"
AMBBG  = "FFF4E0"
WHITE  = "FFFFFF"
thin = Side(style="thin", color="D6DBE6")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ---- title block ----
title = f"ZSFO — Zero Sales Full Optimization  |  {meta['portfolio_holder'].title()}  |  {meta['marketplace']}"
sub = (f"Run {meta['run_date']}  ·  window {meta['window_start']} → {meta['window_end']} "
       f"(last completed 30 days, current day excluded)  ·  "
       f"{meta['universe_asins']:,} UK ASINs → {meta['zero_sale_rows']:,} zero-sale  ·  "
       f"zero-sale = 0 units across FBA+FBM AND Vendor")
ws["A1"] = title
ws["A1"].font = Font(bold=True, size=14, color=WHITE)
ws["A2"] = sub
ws["A2"].font = Font(size=9, color="D6DBE6")

headers = [
    "ASIN", "SKU", "Last Month Sales", "Local UK Warehouse stock", "Amazon FBM Stock",
    "Impressions", "Clicks", "Conversion Rate", "Root-cause hint",
    "Last Amazon Sale", "Last Vendor Sale", "Vendor Units (lifetime)",
    "W1 Impr\n10-16 Jun", "W1 Clk", "W2 Impr\n17-23 Jun", "W2 Clk",
    "W3 Impr\n24-30 Jun", "W3 Clk", "W4 Impr\n01-07 Jul", "W4 Clk",
    "W5 Impr\n08-09 Jul", "W5 Clk",
]
HEAD_ROW = 4
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row=HEAD_ROW, column=c, value=h)
    cell.font = Font(bold=True, color=WHITE, size=9)
    cell.fill = PatternFill("solid", fgColor=HEADBG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

for i in range(1, 3):
    ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=len(headers))
ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor=NAVY)
ws.cell(row=2, column=1).fill = PatternFill("solid", fgColor=NAVY)

def row_vals(r):
    return [
        r["asin"], r["sku"], 0, r["uk_stock"], r["fbm_stock"],
        r["impr"], r["clk"], (r["cr"] if r["cr"] is not None else None), r["root_cause"],
        r["last_order"], r["last_vendor"], r["vlife"],
        r["w1i"], r["w1c"], r["w2i"], r["w2c"], r["w3i"], r["w3c"],
        r["w4i"], r["w4c"], r["w5i"], r["w5c"],
    ]

start = HEAD_ROW + 1
for ridx, r in enumerate(rows):
    excel_row = start + ridx
    vals = row_vals(r)
    oos = (r["uk_stock"] == 0 and r["fbm_stock"] == 0)
    zeroimpr = (r["impr"] == 0)
    for cidx, v in enumerate(vals, start=1):
        cell = ws.cell(row=excel_row, column=cidx, value=v)
        cell.border = border
        cell.font = Font(size=9)
        if cidx >= 3:
            cell.alignment = Alignment(horizontal="center")
        # base stripe
        fill = STRIPE if ridx % 2 else WHITE
        if oos:
            fill = REDBG
        elif zeroimpr:
            fill = AMBBG
        cell.fill = PatternFill("solid", fgColor=fill)
    # number formats
    ws.cell(row=excel_row, column=8).number_format = "0.00%"  # conversion rate as %
    for col in (4, 5, 6, 7, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22):
        ws.cell(row=excel_row, column=col).number_format = "#,##0"

# widths
widths = [14, 26, 9, 12, 10, 12, 9, 10, 40, 13, 13, 11] + [9] * 10
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = f"A{start}"
ws.auto_filter.ref = f"A{HEAD_ROW}:{get_column_letter(len(headers))}{start-1+len(rows)}"
ws.row_dimensions[HEAD_ROW].height = 34

wb.save(OUT)
print(f"wrote {OUT}  ({os.path.getsize(OUT)} bytes, {len(rows)} data rows)")

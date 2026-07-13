# -*- coding: utf-8 -*-
"""build_report.py — Paused Campaign Report (Utharsika), PRJ-2026-007 / REQ-09-D01.
Reads data.json (governed pull) and writes Paused_Campaign_Report_Utharsika.xlsx:
  - Report  : the 7 required columns, one row per still-paused ad target
  - Summary : reconciliation counts + pause-wave / rule breakdown
  - Notes   : rules, open items (A-E), provenance
Run:  python build_report.py   (from this folder)
"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
with open(os.path.join(HERE, "data.json"), encoding="utf-8") as f:
    payload = json.load(f)
meta, rows = payload["meta"], payload["rows"]

COLS = ["Campaign Name", "Ad Group Name", "ASIN", "SKU", "Pause Reason",
        "Campaign Pause Date", "Days Paused"]
KEYS = ["campaign_name", "ad_group_name", "asin", "sku", "pause_reason",
        "campaign_pause_date", "days_paused"]

NAVY = "1F2A44"; ACCENT = "2E5AAC"; LIGHT = "EEF2FA"; BORDER = "D6DEEA"
thin = Side(style="thin", color=BORDER)
box = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_font = Font(bold=True, color="FFFFFF", size=11)
hdr_fill = PatternFill("solid", fgColor=ACCENT)
title_font = Font(bold=True, color="FFFFFF", size=15)
title_fill = PatternFill("solid", fgColor=NAVY)
wrap = Alignment(vertical="top", wrap_text=True)
top = Alignment(vertical="top")
center = Alignment(vertical="top", horizontal="center")

wb = Workbook()

# ---- Report sheet ----
ws = wb.active
ws.title = "Report"
ws.merge_cells("A1:G1")
ws["A1"] = "Paused Campaign Report — Utharsika  (still paused as of %s)" % meta["run_date"]
ws["A1"].font = title_font; ws["A1"].fill = title_fill
ws["A1"].alignment = Alignment(vertical="center", horizontal="left")
ws.row_dimensions[1].height = 26
ws.merge_cells("A2:G2")
ws["A2"] = ("%d ad targets · %d distinct ASINs · Amazon PPC · automation pauses only · "
            "read-only from order_management_copy" % (meta["reconciliation"]["targets"],
                                                       meta["reconciliation"]["distinct_asins"]))
ws["A2"].font = Font(italic=True, color="55607A", size=10)
ws.row_dimensions[2].height = 16

hrow = 3
for j, c in enumerate(COLS, start=1):
    cell = ws.cell(row=hrow, column=j, value=c)
    cell.font = hdr_font; cell.fill = hdr_fill; cell.border = box
    cell.alignment = center
for i, r in enumerate(rows):
    rr = hrow + 1 + i
    for j, k in enumerate(KEYS, start=1):
        cell = ws.cell(row=rr, column=j, value=r[k])
        cell.border = box
        cell.alignment = wrap if k in ("pause_reason", "campaign_name") else top
        if k == "days_paused":
            cell.alignment = center
    if i % 2 == 1:
        for j in range(1, 8):
            ws.cell(row=rr, column=j).fill = PatternFill("solid", fgColor=LIGHT)
widths = [30, 22, 13, 26, 60, 16, 12]
for j, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(j)].width = w
ws.freeze_panes = "A4"
ws.auto_filter.ref = "A3:G%d" % (hrow + len(rows))

# ---- Summary sheet ----
sm = wb.create_sheet("Summary")
rec = meta["reconciliation"]
# rule + wave breakdown (computed)
def rule_of(reason):
    r = reason
    tags = []
    if "Rule 1" in r: tags.append("Rule 1 (ACOS)")
    if "Rule 2" in r: tags.append("Rule 2 (zero orders + spend)")
    if "Rule 3" in r: tags.append("Rule 3 (spend, orders dropped)")
    return " + ".join(tags) if tags else "Other"
from collections import Counter
rules = Counter(rule_of(r["pause_reason"]) for r in rows)
waves = Counter(r["campaign_pause_date"] for r in rows)
camps = Counter(r["campaign_name"] for r in rows)

sm.merge_cells("A1:B1")
sm["A1"] = "Reconciliation & Breakdown"
sm["A1"].font = title_font; sm["A1"].fill = title_fill
sm.row_dimensions[1].height = 24
block = [
    ("Run date", meta["run_date"]),
    ("Still-paused ad targets (report)", rec["targets"]),
    ("Distinct ASINs", rec["distinct_asins"]),
    ("Total automation pauses", rec["total_automation_pauses"]),
    ("Still paused", rec["still_paused"]),
    ("Re-activated (excluded)", rec["reactivated_excluded"]),
    ("", ""),
    ("Pause wave — by date", ""),
]
r = 2
for label, val in block:
    sm.cell(row=r, column=1, value=label).font = Font(bold=(val == "" and label != ""))
    sm.cell(row=r, column=2, value=val)
    r += 1
for d, n in sorted(waves.items()):
    sm.cell(row=r, column=1, value="   %s" % d); sm.cell(row=r, column=2, value=n); r += 1
sm.cell(row=r, column=1, value="Pause rule — by family").font = Font(bold=True); r += 1
for k, n in rules.most_common():
    sm.cell(row=r, column=1, value="   %s" % k); sm.cell(row=r, column=2, value=n); r += 1
sm.cell(row=r, column=1, value="Campaigns represented").font = Font(bold=True); r += 1
sm.cell(row=r, column=1, value="   distinct campaigns"); sm.cell(row=r, column=2, value=len(camps)); r += 1
sm.column_dimensions["A"].width = 40; sm.column_dimensions["B"].width = 16

# ---- Notes sheet ----
nt = wb.create_sheet("Notes")
notes = [
    "Paused Campaign Report — Utharsika  |  PRJ-2026-007 / REQ-09-D01 / PH-2026-07-UTHAR10",
    "",
    "WHAT: Utharsika's Amazon PPC ad targets that automation paused and that are STILL paused today.",
    "GRAIN: one row per paused ad target (per ASIN). Campaign Name = the parent campaign.",
    "SCOPE: campaign name contains 'Utharsika' (no owner column exists). Amazon only (source=1). SB excluded.",
    "PAUSE SOURCE: ppc_etl_automation_log — action_type='ad_pause_logs', status='success', applied_by='0' (automation).",
    "STILL PAUSED: current ppc.record_status='paused' at ad grain. 8 re-activated pauses were excluded.",
    "PAUSE REASON: verbatim from ppc_etl_automation_log.reason (Rule 1 ACOS / Rule 2 zero-orders+spend / Rule 3 spend).",
    "DAYS PAUSED: CURRENT_DATE - pause_date (pause_date = action_datetime::date). Moves with the run date.",
    "",
    "VALIDATION (2026-07-13): 33 targets / 32 ASINs; 41 total pauses / 33 still / 8 re-activated. 4/4 checks PASS.",
    "",
    "OPEN ITEMS — route to Satheesvaran, do NOT decide:",
    "  A. Scope key: name-token 'Utharsika' vs an upstream owner field.",
    "  B. Grain: one row per paused ASIN (current) vs one aggregated row per campaign.",
    "  C. Included set: still-paused only (33) vs every pause incl. re-activated (41).",
    "  D. Platform: Amazon only vs include eBay/SD/SB if pause automation is added there.",
    "  E. Manual pauses: automation-only vs also include manual pauses.",
    "",
    "The workbook's two sample rows (B0DH182H6J / B0CVKSQN9K, 2026-07-06, Days Paused 0) are illustrative only.",
    "SOURCE: order_management_copy (read-only). Rebuild: generate_report.sql -> data.json -> build_report.py.",
]
for i, line in enumerate(notes, start=1):
    c = nt.cell(row=i, column=1, value=line)
    if i == 1:
        c.font = Font(bold=True, size=12)
    elif line.endswith(":") or line.startswith("VALIDATION") or line.startswith("OPEN ITEMS"):
        c.font = Font(bold=True)
nt.column_dimensions["A"].width = 110

out = os.path.join(HERE, "Paused_Campaign_Report_Utharsika.xlsx")
wb.save(out)
print("xlsx written:", out, "|", len(rows), "rows")

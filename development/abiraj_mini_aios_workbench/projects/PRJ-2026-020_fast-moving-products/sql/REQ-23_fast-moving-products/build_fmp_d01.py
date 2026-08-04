# -*- coding: utf-8 -*-
"""
REQ-23-D01 Fast Moving Products (fmp) — Excel builder.
Data source: live warehouse public.order_transaction / listing_data / inv_products /
location_wise_inv_stock, market_place='Germany', pulled 2026-08-04 via the Postgres MCP.
Windows: 30d = last 30 complete days, 90d = last 90 complete days (ending 2026-08-03).
Derived rules (Trend / Action / Final Decision) are DEFAULTS — documented on the Notes tab,
pending Mahima's confirmation.
"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
payload = json.load(open(os.path.join(HERE, "fmp_payload.json"), encoding="utf-8"))
META = payload["meta"]

# ---------- derived helpers ----------
def avg_order_qty(qty30, orders30):
    return round(qty30 / orders30, 2) if orders30 else 0.0

def stock_cover_days(stock, qty30):
    daily = qty30 / 30.0
    if daily <= 0:
        return None  # no recent velocity -> undefined
    return round(stock / daily, 0)

def trend(qty30, qty90):
    # recent daily rate vs 90-day daily rate (90d window contains the 30d window)
    daily30 = qty30 / 30.0
    daily90 = qty90 / 90.0
    if daily90 <= 0:
        return "New"
    r = daily30 / daily90
    if r >= 1.30:
        return "↑ Growing"
    if r >= 0.80:
        return "Stable"
    return "↓ Slowing"

def action(stock, scd, tr):
    if stock == 0:
        return "Restock immediately"
    if scd is not None and scd < 30:
        return "Reorder soon"
    if scd is not None and scd <= 90:
        return "Promote / keep stock" if tr.startswith("↑") else "Maintain stock"
    if scd is not None and scd > 365:
        return "Overstocked – review"
    if tr.startswith("↓") and scd is not None and scd > 180:
        return "Slow – reduce buying"
    return "Monitor"

def final_decision(stock, scd, total_units):
    if stock == 0:
        return "Restock immediately"
    if scd is not None and scd < 30:
        return "Restock soon"
    if scd is not None and scd <= 90:
        return "Maintain stock"
    if scd is not None and scd > 365:
        return "Overstocked – review"
    return "Sufficient stock"

# ---------- styling ----------
FONT = "Arial"
TITLE_F   = Font(name=FONT, size=14, bold=True, color="FFFFFF")
SUB_F     = Font(name=FONT, size=9,  italic=True, color="44546A")
HEAD_F    = Font(name=FONT, size=10, bold=True, color="FFFFFF")
CELL_F    = Font(name=FONT, size=10)
HEAD_FILL = PatternFill("solid", fgColor="1F3864")
TITLE_FILL= PatternFill("solid", fgColor="2E5496")
BAND_FILL = PatternFill("solid", fgColor="EAF0F8")
OOS_FILL  = PatternFill("solid", fgColor="F8D7DA")  # out of stock highlight
thin = Side(style="thin", color="B8C4D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CEN = Alignment(horizontal="center", vertical="center")
LEFT= Alignment(horizontal="left", vertical="center", wrap_text=False)
RIG = Alignment(horizontal="right", vertical="center")

EUR = '#,##0.00 €'
INT = '#,##0'

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEAD_F; cell.fill = HEAD_FILL; cell.alignment = CEN; cell.border = BORDER

def channel_sheet(wb, name, rows, id_label):
    ws = wb.create_sheet(name)
    headers = ["Rank","SKU",id_label,"Product Name","Category",
               "Sold Qty (30 Days)","Sold Qty (90 Days)","Sales Revenue €","Orders",
               "Avg Order Qty","Current Stock","Stock Cover Days","Trend","Action"]
    ncols = len(headers)
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=ncols)
    t = ws.cell(row=1,column=1,value=f"Fast Moving Products – {name} (Germany)")
    t.font=TITLE_F; t.fill=TITLE_FILL; t.alignment=Alignment(horizontal="left",vertical="center")
    ws.row_dimensions[1].height=24
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=ncols)
    ws.cell(row=2,column=1,
            value=(f"Report window: 30-day {META['win30_start']} → {META['win_end']}  |  "
                   f"90-day {META['win90_start']} → {META['win_end']}  |  Currency € (EUR)  |  "
                   f"Live data pulled {META['generated']}  |  Ranked by 30-day units")).font=SUB_F
    hr = 3
    for i,h in enumerate(headers,1):
        ws.cell(row=hr,column=i,value=h)
    style_header(ws, hr, ncols)
    r = hr+1
    for rank,row in enumerate(rows,1):
        scd = stock_cover_days(row["current_stock"], row["qty30"])
        tr  = trend(row["qty30"], row["qty90"])
        vals = [rank, row["sku"], row["product_id"], row["title"] or "",
                row["category"], row["qty30"], row["qty90"], float(row["rev30"]), row["orders30"],
                avg_order_qty(row["qty30"], row["orders30"]), row["current_stock"],
                (scd if scd is not None else "n/a"), tr, action(row["current_stock"], scd, tr)]
        for i,v in enumerate(vals,1):
            ws.cell(row=r,column=i,value=v)
        r += 1
    _finish(ws, hr, r-1, ncols, money_cols=[8], int_cols=[6,7,9,11,12], center_cols=[1,3,13,14],
            band=True, stock_col=11)
    return ws

def combined_sheet(wb, rows):
    ws = wb.create_sheet("Combined")
    headers = ["Overall Rank","SKU","Product Name","Category","Amazon sold Qty","eBay sold Qty",
               "Shopify sold Qty","Total Units Sold","Total Revenue (€)","Current Stock",
               "Stock Cover Days","Final Decision"]
    ncols=len(headers)
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=ncols)
    t=ws.cell(row=1,column=1,value="Final Combined Top Products – All Channels (Germany)")
    t.font=TITLE_F;t.fill=TITLE_FILL;t.alignment=Alignment(horizontal="left",vertical="center")
    ws.row_dimensions[1].height=24
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=ncols)
    ws.cell(row=2,column=1,value=(f"Per shared internal SKU, last 30 days {META['win30_start']} → "
            f"{META['win_end']}  |  Currency €  |  Stock Cover Days = Current Stock ÷ "
            f"(30-day units ÷ 30)  |  Live data {META['generated']}")).font=SUB_F
    hr=3
    for i,h in enumerate(headers,1): ws.cell(row=hr,column=i,value=h)
    style_header(ws,hr,ncols)
    r=hr+1
    for rank,row in enumerate(rows,1):
        scd=stock_cover_days(row["current_stock"], row["total_units"])
        vals=[rank,row["sku"],row["title"] or "",row["category"],row["amz"],row["ebay"],row["shop"],
              row["total_units"],float(row["total_rev"]),row["current_stock"],
              (scd if scd is not None else "n/a"),
              final_decision(row["current_stock"],scd,row["total_units"])]
        for i,v in enumerate(vals,1): ws.cell(row=r,column=i,value=v)
        r+=1
    _finish(ws,hr,r-1,ncols,money_cols=[9],int_cols=[5,6,7,8,10,11],center_cols=[1],
            band=True,stock_col=10)
    return ws

def _finish(ws,hr,last,ncols,money_cols,int_cols,center_cols,band,stock_col):
    for rr in range(hr+1,last+1):
        for c in range(1,ncols+1):
            cell=ws.cell(row=rr,column=c); cell.border=BORDER; cell.font=CELL_F
            if c in money_cols: cell.number_format=EUR; cell.alignment=RIG
            elif c in int_cols: cell.number_format=INT; cell.alignment=CEN
            elif c in center_cols: cell.alignment=CEN
            else: cell.alignment=LEFT
        if band and (rr-hr)%2==0:
            for c in range(1,ncols+1):
                if ws.cell(row=rr,column=c).fill.fgColor.rgb in (None,"00000000"):
                    ws.cell(row=rr,column=c).fill=BAND_FILL
        if ws.cell(row=rr,column=stock_col).value==0:
            for c in range(1,ncols+1): ws.cell(row=rr,column=c).fill=OOS_FILL
    ws.freeze_panes=ws.cell(row=hr+1,column=1)

def notes_sheet(wb):
    ws=wb.create_sheet("Notes & Method",0)
    ws.column_dimensions['A'].width=26; ws.column_dimensions['B'].width=110
    def h(t):
        r=ws.max_row+1 if ws.max_row>1 or ws.cell(1,1).value else 1
        c=ws.cell(row=r,column=1,value=t); c.font=Font(name=FONT,size=12,bold=True,color="1F3864")
        return r
    rows=[
        ("REQ-23-D01 Fast Moving Products","Channel-wise top-selling products for Germany (DE) across Shopify, Amazon and eBay, plus a combined all-channel roll-up. Prepared for Mahima."),
        ("Scope","Germany (DE) only. Currency = EUR (€). Order status = Completed. Ranked by 30-day units sold (top 25 per channel / combined)."),
        ("Windows","30-day = last 30 complete days; 90-day = last 90 complete days; both ending the last complete day. This run: 30d "+META['win30_start']+" → "+META['win_end']+", 90d "+META['win90_start']+" → "+META['win_end']+". Live data pulled "+META['generated']+"."),
        ("Data sources","Sales/units/orders: public.order_transaction (source_name, market_place='Germany'). Product Name: inv_products.title / listing_data.title by SKU. Category: latest non-null order_transaction.category_name per SKU (~74% coverage; rest 'Uncategorised'). Current Stock: location_wise_inv_stock, location='Germany'."),
        ("Sales Revenue €","= SUM(item_price × quantity) in the marketplace currency (EUR). This is gross item revenue and is the per-product figure; it intentionally differs from order_total (which includes shipping/fees and cannot be attributed to a single SKU in multi-item orders)."),
        ("Avg Order Qty","= 30-day units ÷ 30-day orders."),
        ("Stock Cover Days","= Current Stock ÷ Average Daily Sales, where Average Daily Sales = 30-day units ÷ 30. 'n/a' when there is no 30-day velocity."),
        ("Trend [DEFAULT RULE – confirm]","Compares recent daily rate (30d÷30) to the 90-day daily rate (90d÷90). ≥ 1.30 → '↑ Growing'; 0.80–1.30 → 'Stable'; < 0.80 → '↓ Slowing'."),
        ("Action [DEFAULT RULE – confirm]","Stock=0 → Restock immediately; cover<30d → Reorder soon; cover≤90d → Promote/keep (if Growing) else Maintain; cover>365d → Overstocked – review; Slowing & cover>180d → Slow – reduce buying; otherwise Monitor."),
        ("Final Decision [DEFAULT RULE – confirm]","Combined-tab equivalent using combined stock cover: Stock=0 → Restock immediately; <30d → Restock soon; ≤90d → Maintain stock; >365d → Overstocked – review; else Sufficient stock."),
        ("Known data caveats","(1) Trend/Action/Final Decision thresholds above are Claude's documented defaults, NOT yet agreed by Mahima. (2) Some eBay/Shopify variant SKUs carry only a variant label as their title (e.g. '50W', '2'); the best available title was used. (3) Category coverage ~74%. (4) Stock is live 'as of today', not as-of the sales window. (5) Combo SKUs (containing '+') are ranked as sold; their inv_products title may read 'Combo Default Title.'."),
        ("Rows in red","Highlighted rows are currently out of stock (Current Stock = 0)."),
    ]
    for a,b in rows:
        h(a); ws.cell(row=ws.max_row,column=2,value=b).alignment=Alignment(wrap_text=True,vertical="top")
        ws.row_dimensions[ws.max_row].height=max(15,14*(1+len(b)//95))
    return ws

wb=Workbook(); wb.remove(wb.active)
channel_sheet(wb,"Shopify DE",payload["shopify"],"Product ID (Shopify)")
channel_sheet(wb,"Amazon DE", payload["amazon"], "Product ID (ASIN)")
channel_sheet(wb,"eBay DE",   payload["ebay"],   "Product ID (eBay Listing ID)")
combined_sheet(wb,payload["combined"])
notes_sheet(wb)

# column widths
widths={"Notes & Method":None}
for ws in wb.worksheets:
    if ws.title=="Notes & Method": continue
    for col in range(1,ws.max_column+1):
        L=get_column_letter(col)
        hdr=str(ws.cell(row=3,column=col).value or "")
        if hdr in ("Product Name","Product Name "): ws.column_dimensions[L].width=48
        elif "SKU" in hdr: ws.column_dimensions[L].width=30
        elif "Product ID" in hdr: ws.column_dimensions[L].width=18
        elif "Category" in hdr: ws.column_dimensions[L].width=20
        elif hdr in ("Trend","Action","Final Decision"): ws.column_dimensions[L].width=20
        elif hdr in ("Sales Revenue €","Total Revenue (€)"): ws.column_dimensions[L].width=15
        else: ws.column_dimensions[L].width=13

out=os.path.join(HERE,"REQ-23-D01_fast_moving_products.xlsx")
wb.save(out)
print("saved:",out)
for k in ("shopify","amazon","ebay","combined"): print(k, len(payload[k]))

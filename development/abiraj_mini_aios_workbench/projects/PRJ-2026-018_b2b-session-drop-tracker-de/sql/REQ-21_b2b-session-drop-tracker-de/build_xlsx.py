#!/usr/bin/env python3
"""REQ-21-D01 — build the deliverable workbook from the governed bsdt_data.json.
Sheets: Tracker (12 cols) · Thresholds (editable, drives Tier via formula) · Summary · Data Notes."""
import json
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
PROJ = HERE.parent.parent
DATA = json.load(open(HERE / "bsdt_data.json", encoding="utf-8"))
OUTDIR = PROJ / "evidence/final_outputs/REQ-21_b2b-session-drop-tracker-de"
OUTDIR.mkdir(parents=True, exist_ok=True)
OUT = OUTDIR / "REQ-21-D01_b2b_session_drop_tracker_DE.xlsx"

m = DATA["meta"]
NAVY = "1F2A44"; SLATE = "334155"; LIGHT = "F1F5F9"
T3F = "FDE2E2"; T2F = "FEF3C7"; T1F = "E2E8F0"
tier_fill = {"Tier 3 - High": T3F, "Tier 2 - Moderate": T2F, "Tier 1 - Low": T1F}
thin = Side(style="thin", color="CBD5E1")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hcell(c, text, fill=NAVY, color="FFFFFF"):
    c.value = text; c.font = Font(bold=True, color=color, size=10)
    c.fill = PatternFill("solid", fgColor=fill)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border

wb = openpyxl.Workbook()

# ---------- Tracker ----------
ws = wb.active; ws.title = "Tracker"
title = ws.cell(1, 1, f"B2B Session Drop Tracker — Amazon.de (Germany)   ·   End user: {m['end_user']}")
title.font = Font(bold=True, color="FFFFFF", size=13); title.fill = PatternFill("solid", fgColor=NAVY)
ws.merge_cells("A1:L1"); ws.row_dimensions[1].height = 24
sub = ws.cell(2, 1, f"Current: {m['current_window']['label']}   |   Previous: {m['previous_window']['label']}   |   "
                    f"Tier by MAX(prev,current) B2B Sessions  (Tier 2 ≥ {m['thresholds']['tier2_min_sessions']}, "
                    f"Tier 3 ≥ {m['thresholds']['tier3_min_sessions']})   |   {DATA['row_count']} ASINs")
sub.font = Font(italic=True, color="FFFFFF", size=9); sub.fill = PatternFill("solid", fgColor=SLATE)
ws.merge_cells("A2:L2"); ws.row_dimensions[2].height = 16

headers = ["ASIN", "Prev B2B Sessions", "Prev B2B Page Views", "Prev B2B Orders",
           "Current B2B Sessions", "Current B2B Page Views", "Current B2B Orders",
           "Buy Box % (Current)", "Session Change", "Tier", "Status", "Action"]
hr = 3
for j, h in enumerate(headers, 1): hcell(ws.cell(hr, j), h)
ws.freeze_panes = "A4"

# Tier is a live formula referencing the editable Thresholds sheet
for i, r in enumerate(DATA["rows"]):
    rr = hr + 1 + i
    vals = [r["asin"], r["prev_sessions"], r["prev_page_views"], r["prev_orders"],
            r["curr_sessions"], r["curr_page_views"], r["curr_orders"],
            (None if r["buy_box_pct"] is None else r["buy_box_pct"] / 100.0),
            r["session_change"], None, r["status"], r["action"]]
    for j, v in enumerate(vals, 1):
        c = ws.cell(rr, j, v); c.border = border
        c.alignment = Alignment(horizontal="center" if j <= 10 else "left", vertical="center",
                                wrap_text=(j == 12))
        if j == 8 and v is not None: c.number_format = "0.0%"
    # Session Change formula + colour
    ws.cell(rr, 9).value = f"=E{rr}-B{rr}"
    # Tier formula: Tier 2 lower bound in Thresholds!B4, Tier 3 lower bound in Thresholds!B5 (MAX of prev,current)
    ws.cell(rr, 10).value = (f'=IF(MAX(B{rr},E{rr})>=Thresholds!$B$5,"Tier 3 - High",'
                             f'IF(MAX(B{rr},E{rr})>=Thresholds!$B$4,"Tier 2 - Moderate","Tier 1 - Low"))')
    fill = tier_fill[r["tier"]]
    ws.cell(rr, 10).fill = PatternFill("solid", fgColor=fill)
    ws.cell(rr, 11).fill = PatternFill("solid", fgColor=fill)

widths = [14, 12, 12, 11, 13, 13, 11, 12, 11, 16, 22, 60]
for j, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(j)].width = w
ws.auto_filter.ref = f"A{hr}:L{hr + DATA['row_count']}"

# ---------- Thresholds (editable) ----------
th = wb.create_sheet("Thresholds")
th.cell(1, 1, "Threshold Settings — edit the blue cells; the Tracker's Tier column recalculates automatically").font = Font(bold=True, size=11)
th.merge_cells("A1:C1")
for j, h in enumerate(["Tier", "B2B Sessions (lower bound)", "Action"], 1): hcell(th.cell(2, j), h)
BLUE = "DBEAFE"
tri = [("Tier 1 - Low", None, DATA["tier_action"]["Tier 1 - Low"]),
       ("Tier 2 - Moderate", m["thresholds"]["tier2_min_sessions"], DATA["tier_action"]["Tier 2 - Moderate"]),
       ("Tier 3 - High", m["thresholds"]["tier3_min_sessions"], DATA["tier_action"]["Tier 3 - High"])]
for i, (t, b, a) in enumerate(tri):
    rr = 3 + i
    th.cell(rr, 1, t).border = border
    bc = th.cell(rr, 2, ("< Tier 2" if b is None else b)); bc.border = border
    bc.alignment = Alignment(horizontal="center")
    if b is not None: bc.fill = PatternFill("solid", fgColor=BLUE); bc.font = Font(bold=True)
    ac = th.cell(rr, 3, a); ac.border = border; ac.alignment = Alignment(wrap_text=True, vertical="top")
th.column_dimensions["A"].width = 20; th.column_dimensions["B"].width = 24; th.column_dimensions["C"].width = 90
th.cell(7, 1, "Note: Tier 2 lower bound = cell B4 ; Tier 3 lower bound = cell B5. Tier 1 = below Tier 2. The Tracker's Tier column reads B4/B5 and recalculates automatically.").font = Font(italic=True, size=9)

# ---------- Summary ----------
sm = wb.create_sheet("Summary")
sm.cell(1, 1, "Summary — B2B Session Drop Tracker (Amazon.de)").font = Font(bold=True, size=12)
rows = [
    ("End user", m["end_user"]), ("Scope", m["scope"]),
    ("Source of record", m["source_of_record"]),
    ("Current window", m["current_window"]["label"]), ("Previous window", m["previous_window"]["label"]),
    ("Total ASINs", DATA["row_count"]),
    ("Tier 3 - High", DATA["tier_distribution"].get("Tier 3 - High", 0)),
    ("Tier 2 - Moderate", DATA["tier_distribution"].get("Tier 2 - Moderate", 0)),
    ("Tier 1 - Low", DATA["tier_distribution"].get("Tier 1 - Low", 0)),
    ("Engine", m["engine_note"]),
]
for i, (k, v) in enumerate(rows, 3):
    sm.cell(i, 1, k).font = Font(bold=True); sm.cell(i, 2, v)
sm.column_dimensions["A"].width = 20; sm.column_dimensions["B"].width = 100

# ---------- Data Notes ----------
dn = wb.create_sheet("Data Notes")
notes = [
    "B2B Session Drop Tracker — Amazon.de (Germany) — REQ-21-D01",
    "",
    "Source of record: owner-supplied Amazon.de Seller Central Business Report export (Detail Page Sales and Traffic",
    "by Child Item), B2B-only columns (Sessions·Total·B2B, Page Views·Total·B2B, Units ordered·B2B). The internal",
    "database is NOT the source — its business_reports.amz_traffic_by_asin mirror is incomplete for .de B2B (May 2026",
    "missing; ~half the ASINs absent), so it cannot reproduce this report. Confirmed with the owner 2026-07-31.",
    "",
    "Included ASINs: only those with some B2B Sessions or B2B Page Views in at least one window (zero-both excluded).",
    "Tier logic: Tier = MAX(Prev, Current) B2B Sessions vs the editable Thresholds tab (Tier 2 ≥ 5, Tier 3 ≥ 10).",
    "Session Change = Current − Previous B2B Sessions (context/trend only — does NOT change Tier or Action).",
    "Units Orders and Buy Box % are context only; they never change Tier/Status/Action.",
    "Buy Box % shown for the Current window.",
    "",
    "Engine verification: all rows re-derived independently and matched the source — 0 mismatches on Session Change,",
    "Tier, Status and Action.",
]
for i, t in enumerate(notes, 1):
    c = dn.cell(i, 1, t)
    if i == 1: c.font = Font(bold=True, size=12)
dn.column_dimensions["A"].width = 120

wb.save(OUT)
print("Wrote", OUT)

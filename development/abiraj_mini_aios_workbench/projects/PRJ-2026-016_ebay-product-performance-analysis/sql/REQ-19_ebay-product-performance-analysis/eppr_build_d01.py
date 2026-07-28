# -*- coding: utf-8 -*-
"""
REQ-19-D01  eBay Product Performance Analysis  (project code: eppr)  — LEDSONE-PRIMARY build
Source lock (owner, 2026): use the two ledsone MCPs' data only —
  - RAW ledsone Postgres (mcp.ledsone.co.uk) = every column EXCEPT organic traffic
  - AIOS knowledge base (docs.ledsone.co.uk) = the query rules (all_list=1, source_id=2, VARCHAR casts)
The WAREHOUSE is used for ONE feed only: eBay ORGANIC traffic (Impressions/Views/Conversion),
which has no source in ledsone (the ESNM two-DB pattern). Nothing else touches the warehouse.

Grain: one row per eBay LISTING (item_id), UK + Germany, all active eBay accounts.
Window: rolling 30 days ending on the last COMPLETE day (anchor = today-1).
Honesty: every filled column traces to a real table.column; unsourceable columns render 'NO DATA'
(Cost Price + Gross/Net/Margin — no COGS in either DB; Watch Count — eBay API only; Sales Trend —
undefined rule). Money is per marketplace currency (UK GBP £, DE EUR €); never blended.

fetch_records() is the single data layer shared by the xlsx builder (this file) and the dashboard
renderer, so the two outputs can never drift.
"""
import os, psycopg2
from datetime import date, datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Credentials come from the shared global env store (see 05_documentation/capability/shared_db_credentials/);
# NEVER hardcode passwords in tracked code. LED_* = raw ledsone (read-only), PG* = warehouse (temp_user).
LED = dict(host=os.getenv("LED_PGHOST","207.148.78.148"), port=os.getenv("LED_PGPORT","5432"),
           dbname=os.getenv("LED_PGDATABASE","ledsone"), user=os.getenv("LED_PGUSER","dbhub_readonly"),
           password=os.getenv("LED_PGPASSWORD",""), connect_timeout=30)
WH = dict(host=os.getenv("PGHOST","149.28.134.54"), port=os.getenv("PGPORT","5435"),
          dbname=os.getenv("PGDATABASE","order_management_copy"),
          user=os.getenv("PGUSER","temp_user"), password=os.getenv("PGPASSWORD",""), connect_timeout=30)

OUT = os.path.abspath(os.path.join(os.path.dirname(__file__),
      "..","..","evidence","final_outputs","REQ-19_ebay-product-performance-analysis",
      "REQ-19-D01_ebay_product_performance_v4_final.xlsx"))

ND = "NO DATA"
VAT_RATE = {"UK":0.20, "Germany":0.19}
# Brand = the account's eBay store brand (own-brand listings); values as recorded in the business.
BRAND_MAP = {
    "led_sone":"LEDSone","ledsonede":"LEDSone DE","electricalsone":"Electricalsone","so_926407":"Sunsone",
    "dctransformer":"DC Transformer","coventrylights":"Coventry Lights","lighting_sone":"Lightingsone",
    "re6865":"Retroled","vintageinterior":"Vintageinterior","huettenlampen":"Huettenlampen",
    "bestbringer":"Bestbringer","electro_shine":"Electroshine","uk-lightsway":"UK Lightsway",
    "homin_gmbh":"Homin GMBH","neighbourmarket":"Neighbour Market","cottagelighting":"Cottage Lighting",
    "ledpedia":"Ledpedia",
}
HEADERS = ["Product Image","SKU","Parent SKU","eBay Item ID","Product Title","Brand","Category",
           "Marketplace","Account","Listing Date","Listing Status","Selling Price (£/€)","Cost Price (£/€)",
           "Shipping Cost (£/€)","eBay Fees (£/€)","Ad Cost (£/€)","VAT (£/€)","Available Stock","Units Sold","Orders",
           "Revenue (£/€)","Gross Profit (£/€)","Net Profit (£/€)","Profit Margin %","Impressions","Views","Clicks",
           "CTR %","Conversion Rate %","Last Sold Date","Days Active",
           "Promotion Status","PPC Campaign","Sales Trend"]
MONEY_COLS = (12,13,14,15,16,17,21,22,23)

LED_SQL = """
WITH base AS (   -- real listable SKUs (all_list=1): price, stock, listing date, rep SKU
  SELECT el.item_id, el.site, el.sub_source, MIN(el.sku) AS rep_sku, COUNT(*) AS vc,
         MAX(NULLIF(el.category_id,'')) AS cat_id, MIN(el.created_at)::date AS listing_date,
         MAX(el.price) AS price, SUM(el.quantity) AS stock
  FROM listings.ebay_listings el
  WHERE el.all_list=1 AND el.is_ended=0 AND COALESCE(el.wrong_sku,0)=0 AND el.site IN ('UK','Germany')
  GROUP BY el.item_id, el.site, el.sub_source
),
meta AS (   -- title/image/type/parent live on the PARENT row (all_list=0); take over ALL rows per item_id
  SELECT el.item_id, MAX(NULLIF(el.title,'')) AS title, MAX(NULLIF(el.main_image_url,'')) AS img,
         MAX(NULLIF(el.product_type,'')) AS ptype, MAX(NULLIF(el.parent_sku,'')) AS parent_sku
  FROM listings.ebay_listings el
  WHERE el.is_ended=0 AND el.site IN ('UK','Germany')
  GROUP BY el.item_id
),
sales AS (
  SELECT oii.item_id,
         SUM(CAST(NULLIF(oii.item_quantity,'') AS numeric))                                    AS units,
         COUNT(DISTINCT o.id)                                                                  AS orders,
         ROUND(SUM(CAST(NULLIF(oii.item_price,'') AS numeric)*CAST(NULLIF(oii.item_quantity,'') AS numeric)),2) AS revenue,
         MAX(o.order_date)::date                                                               AS last_sold
  FROM order_management.order_item_info oii
  JOIN order_management.orders o     ON o.id=oii.order_id
  JOIN order_management.sub_source ss ON ss.id=o.sub_source_id AND ss.source_id=2
  WHERE o.order_date >= %(t0)s AND o.order_date < %(t1)s AND o.status NOT IN ('Cancelled','Deleted')
  GROUP BY oii.item_id
),
ship AS (
  SELECT item_id, ROUND(SUM(cc),2) AS shipping FROM (
    SELECT DISTINCT oii.item_id, o.id AS oid, o.shipping_cost AS cc
    FROM order_management.order_item_info oii
    JOIN order_management.orders o     ON o.id=oii.order_id
    JOIN order_management.sub_source ss ON ss.id=o.sub_source_id AND ss.source_id=2
    WHERE o.order_date >= %(t0)s AND o.order_date < %(t1)s AND o.status NOT IN ('Cancelled','Deleted')
      AND o.shipping_cost IS NOT NULL
  ) x GROUP BY item_id
),
fees AS (
  SELECT e.item_id::text AS iid, ROUND(SUM(e.fee),2) AS fee
  FROM accounting.ebay_order_expenses e
  WHERE e.transaction_date BETWEEN %(d0)s AND %(d1)s AND COALESCE(e.fee_type,'') NOT IN ('AD_FEE','PREMIUM_AD_FEES')
  GROUP BY e.item_id
),
adcps AS (
  SELECT e.item_id::text AS iid, ROUND(SUM(e.fee),2) AS cps
  FROM accounting.ebay_order_expenses e
  WHERE e.transaction_date BETWEEN %(d0)s AND %(d1)s AND e.fee_type IN ('AD_FEE','PREMIUM_AD_FEES')
  GROUP BY e.item_id
),
adcpc AS (
  SELECT pd.ebay_listing_id::text AS iid, ROUND(SUM(pd.ad_fees_payout_currency),2) AS cpc
  FROM ebay_campaigns.performance_data pd
  WHERE pd.date BETWEEN %(d0)s AND %(d1)s
  GROUP BY pd.ebay_listing_id
),
camp AS (
  SELECT a.ebay_listing_id::text AS iid, MAX(c.campaign_name) AS cname,
         BOOL_OR(NOT COALESCE(c.deleted,false)) AS promoted
  FROM ebay_campaigns.ads a JOIN ebay_campaigns.campaigns c ON c.campaign_id=a.campaign_id
  WHERE COALESCE(a.deleted,false)=false
  GROUP BY a.ebay_listing_id
)
SELECT b.item_id, b.site, ss.name AS account, m.parent_sku, b.rep_sku, b.vc,
       m.title, COALESCE(m.ptype,b.cat_id) AS category, b.listing_date, b.price, b.stock, m.img,
       s.units, s.orders, s.revenue, s.last_sold,
       f.fee AS ebay_fees, COALESCE(ac.cps,0)+COALESCE(ap.cpc,0) AS ad_cost,
       (ac.cps IS NOT NULL OR ap.cpc IS NOT NULL OR COALESCE(cm.promoted,false)) AS is_promoted,
       sh.shipping, cm.cname AS ppc_campaign
FROM base b
JOIN order_management.sub_source ss ON ss.id=b.sub_source
JOIN meta m ON m.item_id=b.item_id
LEFT JOIN sales s  ON s.item_id=b.item_id
LEFT JOIN ship  sh ON sh.item_id=b.item_id
LEFT JOIN fees  f  ON f.iid=b.item_id
LEFT JOIN adcps ac ON ac.iid=b.item_id
LEFT JOIN adcpc ap ON ap.iid=b.item_id
LEFT JOIN camp  cm ON cm.iid=b.item_id
ORDER BY s.revenue DESC NULLS LAST, b.item_id;
"""

TRAFFIC_SQL = """
SELECT ref_id, SUM(impression) AS impr, SUM(click) AS clicks, SUM(conversion) AS conv
FROM public.traffic_data
WHERE which_channel=2 AND market_place IN ('UK','Germany') AND date::date BETWEEN %(d0)s AND %(d1)s
GROUP BY ref_id;
"""

def fetch_records():
    anchor = date.today() - timedelta(days=1)
    d0 = anchor - timedelta(days=29)
    params = {"t0": datetime.combine(d0, datetime.min.time()),
              "t1": datetime.combine(anchor + timedelta(days=1), datetime.min.time()),
              "d0": d0, "d1": anchor}
    # 1) everything from raw ledsone
    lc = psycopg2.connect(**LED); cur = lc.cursor(); cur.execute(LED_SQL, params)
    rows = cur.fetchall(); lc.close()
    # 2) organic traffic — warehouse only (no ledsone source)
    traffic = {}
    try:
        wc = psycopg2.connect(**WH); wcur = wc.cursor(); wcur.execute(TRAFFIC_SQL, {"d0": d0, "d1": anchor})
        for ref, impr, clk, conv in wcur.fetchall():
            traffic[str(ref)] = (int(impr) if impr is not None else None,
                                 int(clk) if clk is not None else None,
                                 float(conv) if conv is not None else None)
        wc.close()
    except Exception as e:
        print("WARN traffic feed unavailable:", str(e).splitlines()[0])

    recs = []
    for r in rows:
        (item_id, site, account, parent_sku, rep_sku, vc, title, category, ldate, price, stock, img,
         units, orders, revenue, last_sold, ebay_fees, ad_cost, is_promoted, shipping, ppc) = r
        mkt = site  # 'UK' / 'Germany'
        cur_sym = "£" if mkt == "UK" else "€"
        revenue = float(revenue) if revenue is not None else 0.0
        vat = round(revenue - revenue/(1+VAT_RATE[mkt]), 2) if revenue else 0
        tr = traffic.get(str(item_id))
        impr = tr[0] if tr else None
        views = tr[1] if tr else None
        clicks = tr[1] if tr else None
        conv = tr[2] if tr else None
        ctr = round(clicks/impr*100, 2) if impr and clicks is not None and impr > 0 else None
        cvr = round(conv/clicks*100, 2) if clicks and conv is not None and clicks > 0 else None
        # Profit stack. Cost Price is an ESTIMATE = 20% of selling price (owner-agreed proxy; no real
        # COGS exists in either database). Gross/Net/Margin are derived from it and are therefore
        # estimates, not booked figures — flagged on every artefact.
        price_v = round(float(price), 2) if price is not None else None
        cost_v = round(price_v*0.20, 2) if price_v is not None else None
        ship_v = float(shipping) if shipping is not None else 0
        fees_v = float(ebay_fees) if ebay_fees is not None else 0
        ad_v = float(ad_cost) if ad_cost is not None else 0
        units_v = int(units) if units is not None else 0
        gross_v = round(revenue - cost_v*units_v, 2) if cost_v is not None else None
        net_v = round(gross_v - fees_v - ad_v - ship_v - vat, 2) if gross_v is not None else None
        margin_v = round(net_v/revenue*100, 2) if (net_v is not None and revenue) else None
        v = [
            img or None,
            (rep_sku or ND) + ("" if vc == 1 else " (+%d)" % (vc-1)),
            parent_sku or None, str(item_id), title or None,
            BRAND_MAP.get(account, account.title() if account else None), category or None,
            mkt, account, ldate.isoformat() if ldate else None, "Active",
            price_v,
            cost_v,                                        # Cost Price = est. 20% of selling price
            ship_v, fees_v, ad_v,
            vat, int(stock) if stock is not None else None,
            units_v, int(orders) if orders is not None else 0, revenue,
            gross_v, net_v, margin_v,                      # derived from the 20% cost estimate
            impr, views, clicks, ctr, cvr,                 # traffic (Watch Count column removed — no source)
            last_sold.isoformat() if last_sold else None,
            (anchor - ldate).days if ldate else None,
            "Promoted" if is_promoted else "Not Promoted", ppc or None, None,   # Sales Trend
        ]
        recs.append({"c": cur_sym, "v": v})
    return recs, d0, anchor

def main():
    recs, d0, anchor = fetch_records()
    wb = Workbook(); ws = wb.active; ws.title = "eBay Product Performance"
    note = ("REQ-19-D01 eBay Product Performance Analysis | one row per eBay listing (item_id) | "
            "SOURCE: raw ledsone (mcp.ledsone.co.uk) for all columns; warehouse used ONLY for organic traffic | "
            "window %s..%s (30 days) | money per marketplace currency: UK £ / DE € (never blended) | "
            "Views=Clicks (eBay one organic click/view metric) | Brand=account store brand | "
            "VAT=std output VAT 20%% UK / 19%% DE of revenue | "
            "⚠ ESTIMATE: Cost Price = 20%% of selling price (no real COGS exists); Gross/Net Profit & Margin "
            "are derived from it and are ESTIMATES, not booked figures | "
            "'NO DATA' = no source: Sales Trend (undefined rule). (Watch Count column removed — eBay API only, in no DB.)" % (d0, anchor))
    ws["A1"] = note; ws["A1"].font = Font(name="Arial", size=9, italic=True, color="555555")
    ws.append([]); hr = 3
    ws.append(HEADERS)
    for c in range(1, len(HEADERS)+1):
        cell = ws.cell(row=hr, column=c)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    rev = {"UK":0.0, "Germany":0.0}
    for rec in recs:
        v = rec["v"]; mkt = v[7]
        if isinstance(v[20], (int, float)): rev[mkt] += v[20]
        ws.append([("NO DATA" if x is None else x) for x in v])
        rownum = ws.max_row
        cur_fmt = u'£#,##0.00' if mkt == "UK" else u'€#,##0.00'
        for col in MONEY_COLS:
            cell = ws.cell(row=rownum, column=col)
            if isinstance(cell.value, (int, float)): cell.number_format = cur_fmt
    ws.freeze_panes = "E4"
    widths=[16,20,18,15,40,15,20,11,14,12,12,12,12,12,12,12,10,9,9,8,12,12,12,11,11,9,9,8,12,12,10,14,16,11]
    for i,w in enumerate(widths,1):
        ws.column_dimensions[ws.cell(row=hr,column=i).column_letter].width = w
    wb.save(OUT)
    filled = sum(1 for x in recs[0]["v"] if x is not None) if recs else 0
    print("ROWS (live D count):", len(recs))
    print("Revenue  UK £%.2f  DE €%.2f" % (rev["UK"], rev["Germany"]))
    print("Saved:", OUT)

if __name__ == "__main__":
    main()
